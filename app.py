from flask import Flask, render_template, request, jsonify, send_file
from models import get_db_connection, init_database
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import os

app = Flask(__name__)

# 初始化数据库
init_database()

# ==================== 首页 ====================
@app.route('/')
def index():
    return render_template('index.html')

# ==================== 指标管理 ====================
@app.route('/api/indicators', methods=['GET', 'POST'])
def indicators():
    conn = get_db_connection()

    if request.method == 'POST':
        data = request.json
        cursor = conn.cursor()
        cursor.execute(
            'INSERT INTO indicators (name, unit, description) VALUES (?, ?, ?)',
            (data['name'], data.get('unit'), data.get('description'))
        )
        conn.commit()
        indicator_id = cursor.lastrowid
        conn.close()
        return jsonify({'id': indicator_id, 'message': '指标添加成功'}), 201

    indicators = conn.execute('SELECT * FROM indicators ORDER BY created_at DESC').fetchall()
    conn.close()
    return jsonify([dict(row) for row in indicators])

@app.route('/api/indicators/<int:id>', methods=['PUT', 'DELETE'])
def indicator_detail(id):
    conn = get_db_connection()

    if request.method == 'DELETE':
        conn.execute('DELETE FROM indicators WHERE id = ?', (id,))
        conn.commit()
        conn.close()
        return jsonify({'message': '指标删除成功'})

    if request.method == 'PUT':
        data = request.json
        conn.execute(
            'UPDATE indicators SET name = ?, unit = ?, description = ? WHERE id = ?',
            (data['name'], data.get('unit'), data.get('description'), id)
        )
        conn.commit()
        conn.close()
        return jsonify({'message': '指标更新成功'})

# ==================== 检测方法管理 ====================
@app.route('/api/detection-methods', methods=['GET', 'POST'])
def detection_methods():
    conn = get_db_connection()

    if request.method == 'POST':
        data = request.json
        cursor = conn.cursor()
        cursor.execute(
            'INSERT INTO detection_methods (indicator_id, method_name, standard_code, description) VALUES (?, ?, ?, ?)',
            (data['indicator_id'], data['method_name'], data.get('standard_code'), data.get('description'))
        )
        conn.commit()
        method_id = cursor.lastrowid
        conn.close()
        return jsonify({'id': method_id, 'message': '检测方法添加成功'}), 201

    indicator_id = request.args.get('indicator_id')
    if indicator_id:
        methods = conn.execute(
            'SELECT * FROM detection_methods WHERE indicator_id = ? ORDER BY created_at DESC',
            (indicator_id,)
        ).fetchall()
    else:
        methods = conn.execute(
            'SELECT dm.*, i.name as indicator_name FROM detection_methods dm '
            'LEFT JOIN indicators i ON dm.indicator_id = i.id '
            'ORDER BY dm.created_at DESC'
        ).fetchall()

    conn.close()
    return jsonify([dict(row) for row in methods])

@app.route('/api/detection-methods/<int:id>', methods=['DELETE'])
def detection_method_detail(id):
    conn = get_db_connection()
    conn.execute('DELETE FROM detection_methods WHERE id = ?', (id,))
    conn.commit()
    conn.close()
    return jsonify({'message': '检测方法删除成功'})

# ==================== 限值标准管理 ====================
@app.route('/api/limit-standards', methods=['GET', 'POST'])
def limit_standards():
    conn = get_db_connection()

    if request.method == 'POST':
        data = request.json
        cursor = conn.cursor()
        cursor.execute(
            'INSERT INTO limit_standards (indicator_id, standard_name, min_value, max_value, exact_value, description) '
            'VALUES (?, ?, ?, ?, ?, ?)',
            (data['indicator_id'], data['standard_name'], data.get('min_value'),
             data.get('max_value'), data.get('exact_value'), data.get('description'))
        )
        conn.commit()
        standard_id = cursor.lastrowid
        conn.close()
        return jsonify({'id': standard_id, 'message': '限值标准添加成功'}), 201

    indicator_id = request.args.get('indicator_id')
    if indicator_id:
        standards = conn.execute(
            'SELECT * FROM limit_standards WHERE indicator_id = ? ORDER BY created_at DESC',
            (indicator_id,)
        ).fetchall()
    else:
        standards = conn.execute(
            'SELECT ls.*, i.name as indicator_name FROM limit_standards ls '
            'LEFT JOIN indicators i ON ls.indicator_id = i.id '
            'ORDER BY ls.created_at DESC'
        ).fetchall()

    conn.close()
    return jsonify([dict(row) for row in standards])

@app.route('/api/limit-standards/<int:id>', methods=['DELETE'])
def limit_standard_detail(id):
    conn = get_db_connection()
    conn.execute('DELETE FROM limit_standards WHERE id = ?', (id,))
    conn.commit()
    conn.close()
    return jsonify({'message': '限值标准删除成功'})

# ==================== 报告管理 ====================
@app.route('/api/reports', methods=['GET', 'POST'])
def reports():
    conn = get_db_connection()

    if request.method == 'POST':
        data = request.json
        cursor = conn.cursor()
        cursor.execute(
            'INSERT INTO reports (report_name, sample_location, sample_date, sampler, weather, temperature, remark) '
            'VALUES (?, ?, ?, ?, ?, ?, ?)',
            (data['report_name'], data.get('sample_location'), data.get('sample_date'),
             data.get('sampler'), data.get('weather'), data.get('temperature'), data.get('remark'))
        )
        conn.commit()
        report_id = cursor.lastrowid

        # 添加报告数据
        if 'data' in data:
            for item in data['data']:
                cursor.execute(
                    'INSERT INTO report_data (report_id, indicator_id, detection_method_id, measured_value, '
                    'limit_standard_id, is_qualified, remark) VALUES (?, ?, ?, ?, ?, ?, ?)',
                    (report_id, item['indicator_id'], item.get('detection_method_id'),
                     item.get('measured_value'), item.get('limit_standard_id'),
                     item.get('is_qualified'), item.get('remark'))
                )

        conn.commit()
        conn.close()
        return jsonify({'id': report_id, 'message': '报告创建成功'}), 201

    reports = conn.execute('SELECT * FROM reports ORDER BY created_at DESC').fetchall()
    conn.close()
    return jsonify([dict(row) for row in reports])

@app.route('/api/reports/<int:id>', methods=['GET', 'DELETE'])
def report_detail(id):
    conn = get_db_connection()

    if request.method == 'DELETE':
        conn.execute('DELETE FROM reports WHERE id = ?', (id,))
        conn.commit()
        conn.close()
        return jsonify({'message': '报告删除成功'})

    report = conn.execute('SELECT * FROM reports WHERE id = ?', (id,)).fetchone()
    if not report:
        conn.close()
        return jsonify({'error': '报告不存在'}), 404

    data = conn.execute(
        'SELECT rd.*, i.name as indicator_name, i.unit, '
        'dm.method_name, dm.standard_code, '
        'ls.standard_name, ls.min_value, ls.max_value, ls.exact_value '
        'FROM report_data rd '
        'LEFT JOIN indicators i ON rd.indicator_id = i.id '
        'LEFT JOIN detection_methods dm ON rd.detection_method_id = dm.id '
        'LEFT JOIN limit_standards ls ON rd.limit_standard_id = ls.id '
        'WHERE rd.report_id = ?',
        (id,)
    ).fetchall()

    conn.close()

    result = dict(report)
    result['data'] = [dict(row) for row in data]
    return jsonify(result)

# ==================== Excel导出 ====================
@app.route('/api/reports/<int:id>/export')
def export_report(id):
    conn = get_db_connection()

    report = conn.execute('SELECT * FROM reports WHERE id = ?', (id,)).fetchone()
    if not report:
        conn.close()
        return jsonify({'error': '报告不存在'}), 404

    data = conn.execute(
        'SELECT rd.*, i.name as indicator_name, i.unit, '
        'dm.method_name, dm.standard_code, '
        'ls.standard_name, ls.min_value, ls.max_value, ls.exact_value '
        'FROM report_data rd '
        'LEFT JOIN indicators i ON rd.indicator_id = i.id '
        'LEFT JOIN detection_methods dm ON rd.detection_method_id = dm.id '
        'LEFT JOIN limit_standards ls ON rd.limit_standard_id = ls.id '
        'WHERE rd.report_id = ?',
        (id,)
    ).fetchall()

    conn.close()

    # 创建Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "水质检测报告"

    # 设置样式
    title_font = Font(name='宋体', size=16, bold=True)
    header_font = Font(name='宋体', size=11, bold=True)
    normal_font = Font(name='宋体', size=10)

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 标题
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = report['report_name']
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # 报告信息
    row = 3
    info_items = [
        ('采样地点', report['sample_location']),
        ('采样日期', report['sample_date']),
        ('采样人员', report['sampler']),
        ('天气', report['weather']),
        ('温度', report['temperature'])
    ]

    for label, value in info_items:
        if value:
            ws[f'A{row}'] = label + '：'
            ws[f'B{row}'] = value
            ws.merge_cells(f'B{row}:H{row}')
            row += 1

    row += 1

    # 表头
    headers = ['序号', '检测指标', '单位', '检测方法', '标准代号', '限值标准', '检测值', '是否合格']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # 数据行
    for idx, item in enumerate(data, start=1):
        row += 1

        # 限值显示
        limit_text = ''
        if item['standard_name']:
            if item['exact_value'] is not None:
                limit_text = f"{item['standard_name']}: {item['exact_value']}"
            elif item['min_value'] is not None and item['max_value'] is not None:
                limit_text = f"{item['standard_name']}: {item['min_value']}-{item['max_value']}"
            elif item['min_value'] is not None:
                limit_text = f"{item['standard_name']}: ≥{item['min_value']}"
            elif item['max_value'] is not None:
                limit_text = f"{item['standard_name']}: ≤{item['max_value']}"
            else:
                limit_text = item['standard_name']

        row_data = [
            idx,
            item['indicator_name'],
            item['unit'] or '',
            item['method_name'] or '',
            item['standard_code'] or '',
            limit_text,
            item['measured_value'] or '',
            '合格' if item['is_qualified'] else '不合格' if item['is_qualified'] is not None else ''
        ]

        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.font = normal_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

    # 调整列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12

    # 保存文件
    os.makedirs('exports', exist_ok=True)
    filename = f"exports/水质检测报告_{report['report_name']}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    wb.save(filename)

    return send_file(filename, as_attachment=True, download_name=f"{report['report_name']}.xlsx")

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
