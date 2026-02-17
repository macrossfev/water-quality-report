from flask import Blueprint, request, jsonify, session, send_file
from auth import login_required, log_operation
from models_v2 import get_db
from datetime import datetime
import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

customer_bp = Blueprint('customer_bp', __name__)

# ==================== 客户管理 API ====================
@customer_bp.route('/api/customers', methods=['GET', 'POST'])
@login_required
def api_customers():
    """客户管理"""
    with get_db() as conn:

        if request.method == 'POST':
            data = request.json
            inspected_unit = data.get('inspected_unit', '').strip()
            water_plant = data.get('water_plant', '').strip()
            unit_address = data.get('unit_address', '').strip()
            contact_person = data.get('contact_person', '').strip()
            contact_phone = data.get('contact_phone', '').strip()
            email = data.get('email', '').strip()
            remark = data.get('remark', '').strip()

            if not inspected_unit:
                return jsonify({'error': '被检单位不能为空'}), 400

            try:
                cursor = conn.cursor()
                cursor.execute('''
                    INSERT INTO customers (inspected_unit, water_plant, unit_address,
                                          contact_person, contact_phone, email, remark)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (inspected_unit, water_plant, unit_address, contact_person,
                      contact_phone, email, remark))
                customer_id = cursor.lastrowid

                log_operation('添加客户', f'添加客户: {inspected_unit}', conn=conn)

                return jsonify({'id': customer_id, 'message': '客户添加成功'}), 201
            except Exception as e:
                return jsonify({'error': f'添加客户失败: {str(e)}'}), 400

        # GET请求
        customers = conn.execute('''
            SELECT id, inspected_unit, water_plant, unit_address,
                   contact_person, contact_phone, email, remark,
                   created_at, updated_at
            FROM customers
            ORDER BY created_at DESC
        ''').fetchall()

        return jsonify([dict(customer) for customer in customers])

@customer_bp.route('/api/customers/export', methods=['GET'])
@login_required
def api_customers_export():
    """导出客户信息到Excel"""
    with get_db() as conn:
        try:
            customers = conn.execute('''
                SELECT inspected_unit, water_plant, unit_address,
                       contact_person, contact_phone, email, remark
                FROM customers ORDER BY created_at DESC
            ''').fetchall()

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '客户信息'

            headers = ['被检单位', '水厂名称', '单位地址', '联系人', '联系电话', '邮箱', '备注']
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

            for row_idx, customer in enumerate(customers, 2):
                row_data = [customer['inspected_unit'], customer['water_plant'],
                            customer['unit_address'], customer['contact_person'],
                            customer['contact_phone'], customer['email'], customer['remark']]
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value or '')
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center')

            # Auto-adjust column widths
            for col_idx, header in enumerate(headers, 1):
                max_length = len(header) * 2
                for row in range(2, len(customers) + 2):
                    val = ws.cell(row=row, column=col_idx).value
                    if val:
                        max_length = max(max_length, len(str(val)) * 1.2)
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max(max_length + 2, 10), 50)

            os.makedirs('temp', exist_ok=True)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filepath = os.path.join('temp', f'客户信息_{timestamp}.xlsx')
            wb.save(filepath)

            log_operation('导出客户', f'导出客户信息，共{len(customers)}条记录')

            return send_file(filepath, as_attachment=True,
                             download_name=f'客户信息_{timestamp}.xlsx',
                             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        except Exception as e:
            return jsonify({'error': f'导出失败: {str(e)}'}), 500


@customer_bp.route('/api/customers/download-template', methods=['GET'])
@login_required
def api_customers_download_template():
    """下载客户导入模板"""
    try:
        wb = openpyxl.Workbook()

        # Sheet 1: 客户信息
        ws_data = wb.active
        ws_data.title = '客户信息'

        headers = ['被检单位', '水厂名称', '单位地址', '联系人', '联系电话', '邮箱', '备注']
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # Example rows
        example_data = [
            ['XX市自来水公司', '第一水厂', 'XX市XX区XX路100号', '张三', '010-12345678', 'zhangsan@example.com', '重点客户'],
            ['XX县供水有限公司', '城南水厂', 'XX县XX镇XX街50号', '李四', '13800138000', 'lisi@example.com', ''],
        ]
        example_font = Font(name='微软雅黑', color='808080', italic=True, size=10)
        for row_idx, row_data in enumerate(example_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_data.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.font = example_font
                cell.alignment = Alignment(vertical='center')

        # Empty rows with borders
        for row_idx in range(4, 14):
            for col_idx in range(1, len(headers) + 1):
                cell = ws_data.cell(row=row_idx, column=col_idx, value='')
                cell.border = thin_border

        # Auto-adjust column widths
        col_widths = [20, 18, 30, 10, 18, 25, 20]
        for col_idx, width in enumerate(col_widths, 1):
            ws_data.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        # Sheet 2: 填写说明
        ws_help = wb.create_sheet(title='填写说明')
        instructions = [
            ['客户信息导入模板 - 填写说明'],
            [''],
            ['1. 请在"客户信息"工作表中填写客户数据，从第2行开始（第1行为表头，请勿修改）。'],
            ['2. "被检单位"为必填项，不能为空。'],
            ['3. "联系电话"支持手机号和座机号格式，如：13800138000 或 010-12345678，长度7-15位。'],
            ['4. "邮箱"需包含@符号，如：example@mail.com。'],
            ['5. 示例数据（灰色斜体行）可直接覆盖或删除，导入时会正常处理。'],
            ['6. 导入时如遇到"被检单位+水厂名称"重复的记录，可选择跳过、覆盖或中止导入。'],
        ]
        title_font = Font(name='微软雅黑', bold=True, size=14)
        body_font = Font(name='微软雅黑', size=11)
        for row_idx, row in enumerate(instructions, 1):
            cell = ws_help.cell(row=row_idx, column=1, value=row[0] if row else '')
            cell.font = title_font if row_idx == 1 else body_font
        ws_help.column_dimensions['A'].width = 80

        os.makedirs('temp', exist_ok=True)
        filepath = os.path.join('temp', '客户导入模板.xlsx')
        wb.save(filepath)

        return send_file(filepath, as_attachment=True,
                         download_name='客户导入模板.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'error': f'下载模板失败: {str(e)}'}), 500


@customer_bp.route('/api/customers/import', methods=['POST'])
@login_required
def api_customers_import():
    """从Excel导入客户信息"""
    import re

    if 'file' not in request.files:
        return jsonify({'error': '未选择文件'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '未选择文件'}), 400

    if not ('.' in file.filename and file.filename.rsplit('.', 1)[1].lower() in {'xlsx', 'xls'}):
        return jsonify({'error': '文件格式不支持，仅支持.xlsx和.xls格式'}), 400

    on_duplicate = request.form.get('on_duplicate', 'skip')  # skip, overwrite, abort

    try:
        # Save uploaded file
        os.makedirs('temp/uploads', exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        filepath = os.path.join('temp/uploads', f'{timestamp}_customer_import.xlsx')
        file.save(filepath)

        wb = openpyxl.load_workbook(filepath, data_only=True)

        # Try sheet "客户信息", fallback to first sheet
        if '客户信息' in wb.sheetnames:
            ws = wb['客户信息']
        else:
            ws = wb.worksheets[0]

        # Read header row
        header_map = {}
        field_mapping = {
            '被检单位': 'inspected_unit', '水厂名称': 'water_plant',
            '单位地址': 'unit_address', '联系人': 'contact_person',
            '联系电话': 'contact_phone', '邮箱': 'email', '备注': 'remark'
        }
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col_idx).value
            if val and str(val).strip() in field_mapping:
                header_map[col_idx] = field_mapping[str(val).strip()]

        if 'inspected_unit' not in header_map.values():
            os.remove(filepath)
            return jsonify({'error': '未找到"被检单位"列，请检查模板格式'}), 400

        # Load existing customers for duplicate check
        with get_db() as conn:
            existing = {}
            rows = conn.execute('SELECT id, inspected_unit, water_plant FROM customers').fetchall()
            for r in rows:
                key = (r['inspected_unit'] or '', r['water_plant'] or '')
                existing[key] = r['id']

            errors = []
            warnings = []
            success_count = 0
            skip_count = 0
            total_rows = 0

            phone_pattern = re.compile(r'^[\d\-]{7,15}$')

            data_rows = []
            for row_idx in range(2, ws.max_row + 1):
                row_data = {}
                has_data = False
                for col_idx, field_name in header_map.items():
                    val = ws.cell(row=row_idx, column=col_idx).value
                    row_data[field_name] = str(val).strip() if val is not None else ''
                    if row_data[field_name]:
                        has_data = True
                if not has_data:
                    continue
                data_rows.append((row_idx, row_data))

            total_rows = len(data_rows)

            for row_idx, row_data in data_rows:
                row_errors = []

                # Validate required field
                if not row_data.get('inspected_unit'):
                    row_errors.append(f'第{row_idx}行: 被检单位不能为空')

                # Validate phone
                phone = row_data.get('contact_phone', '')
                if phone and not phone_pattern.match(phone):
                    row_errors.append(f'第{row_idx}行: 联系电话格式不正确(应为数字和横线，7-15位)')

                # Validate email
                email_val = row_data.get('email', '')
                if email_val and '@' not in email_val:
                    row_errors.append(f'第{row_idx}行: 邮箱格式不正确(需包含@)')

                if row_errors:
                    errors.extend(row_errors)
                    continue

                # Duplicate check
                dup_key = (row_data.get('inspected_unit', ''), row_data.get('water_plant', ''))
                if dup_key in existing:
                    if on_duplicate == 'abort':
                        errors.append(f'第{row_idx}行: 发现重复记录({dup_key[0]} + {dup_key[1]})，中止导入')
                        os.remove(filepath)
                        return jsonify({
                            'success': False,
                            'message': f'导入中止：第{row_idx}行发现重复记录',
                            'total_rows': total_rows,
                            'success_count': success_count,
                            'skip_count': skip_count,
                            'errors': errors,
                            'warnings': warnings
                        })
                    elif on_duplicate == 'overwrite':
                        existing_id = existing[dup_key]
                        conn.execute('''
                            UPDATE customers SET unit_address=?, contact_person=?,
                                contact_phone=?, email=?, remark=?, updated_at=CURRENT_TIMESTAMP
                            WHERE id=?
                        ''', (row_data.get('unit_address', ''), row_data.get('contact_person', ''),
                              row_data.get('contact_phone', ''), row_data.get('email', ''),
                              row_data.get('remark', ''), existing_id))
                        success_count += 1
                        warnings.append(f'第{row_idx}行: 已覆盖更新({dup_key[0]} + {dup_key[1]})')
                    else:
                        skip_count += 1
                        warnings.append(f'第{row_idx}行: 跳过重复记录({dup_key[0]} + {dup_key[1]})')
                else:
                    conn.execute('''
                        INSERT INTO customers (inspected_unit, water_plant, unit_address,
                                              contact_person, contact_phone, email, remark)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''', (row_data.get('inspected_unit', ''), row_data.get('water_plant', ''),
                          row_data.get('unit_address', ''), row_data.get('contact_person', ''),
                          row_data.get('contact_phone', ''), row_data.get('email', ''),
                          row_data.get('remark', '')))
                    existing[dup_key] = True
                    success_count += 1

            log_operation('导入客户', f'导入客户信息: 共{total_rows}行, 成功{success_count}, 跳过{skip_count}, 错误{len(errors)}', conn=conn)

            # Clean up
            try:
                os.remove(filepath)
            except Exception:
                pass

            return jsonify({
                'success': True,
                'message': f'导入完成: 成功{success_count}条, 跳过{skip_count}条, 错误{len(errors)}条',
                'total_rows': total_rows,
                'success_count': success_count,
                'skip_count': skip_count,
                'errors': errors,
                'warnings': warnings
            })
    except Exception as e:
        return jsonify({'error': f'导入失败: {str(e)}'}), 500


@customer_bp.route('/api/customers/<int:id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def api_customer_detail(id):
    """客户详情操作"""
    with get_db() as conn:

        # GET请求 - 获取单个客户详情
        if request.method == 'GET':
            customer = conn.execute('''
                SELECT id, inspected_unit, water_plant, unit_address,
                       contact_person, contact_phone, email, remark,
                       created_at, updated_at
                FROM customers WHERE id = ?
            ''', (id,)).fetchone()

            if not customer:
                return jsonify({'error': '客户不存在'}), 404

            return jsonify(dict(customer))

        # DELETE请求
        if request.method == 'DELETE':
            customer = conn.execute('SELECT inspected_unit FROM customers WHERE id = ?', (id,)).fetchone()

            if not customer:
                return jsonify({'error': '客户不存在'}), 404

            conn.execute('DELETE FROM customers WHERE id = ?', (id,))

            log_operation('删除客户', f'删除客户: {customer["inspected_unit"]}', conn=conn)

            return jsonify({'message': '客户删除成功'})

        # PUT请求
        if request.method == 'PUT':
            data = request.json
            inspected_unit = data.get('inspected_unit', '').strip()
            water_plant = data.get('water_plant', '').strip()
            unit_address = data.get('unit_address', '').strip()
            contact_person = data.get('contact_person', '').strip()
            contact_phone = data.get('contact_phone', '').strip()
            email = data.get('email', '').strip()
            remark = data.get('remark', '').strip()

            if not inspected_unit:
                return jsonify({'error': '被检单位不能为空'}), 400

            try:
                conn.execute('''
                    UPDATE customers
                    SET inspected_unit = ?, water_plant = ?, unit_address = ?,
                        contact_person = ?, contact_phone = ?, email = ?, remark = ?,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                ''', (inspected_unit, water_plant, unit_address, contact_person,
                      contact_phone, email, remark, id))

                log_operation('更新客户', f'更新客户: {inspected_unit}', conn=conn)

                return jsonify({'message': '客户更新成功'})
            except Exception as e:
                return jsonify({'error': f'更新客户失败: {str(e)}'}), 400
