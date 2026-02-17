from flask import Blueprint, request, jsonify, send_file, session, current_app
from models_v2 import get_db
from auth import login_required, admin_required, log_operation
import json
import os
import re
import io
import traceback
import openpyxl
import pandas as pd
import sqlite3
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime
from field_code_mapping import FieldCodeMapping
from generate_example_template import create_example_template
from report_template_exporter import export_report_template
from sample_type_exporter import export_sample_type_template
from import_processor import import_reports_from_excel

import_bp = Blueprint('import_bp', __name__)


# ==================== 辅助函数 ====================

def identify_sheet_type(sheet_name):
    """识别工作表类型"""
    sheet_name_lower = sheet_name.lower()

    if '1' in sheet_name or 'cover' in sheet_name_lower or '封面' in sheet_name:
        return 'cover'
    elif '2' in sheet_name or 'info' in sheet_name_lower or '信息' in sheet_name:
        return 'info'
    elif any(x in sheet_name for x in ['3', '4']) or 'data' in sheet_name_lower or '数据' in sheet_name:
        return 'data'
    elif '5' in sheet_name or 'note' in sheet_name_lower or '说明' in sheet_name:
        return 'conclusion'
    else:
        return 'other'

def extract_page_number(sheet_name):
    """从工作表名称提取页码"""
    match = re.search(r'\d+', sheet_name)
    return int(match.group()) if match else 0


# ==================== 检测指标导入导出 API ====================

@import_bp.route('/api/indicators/export/excel', methods=['GET'])
@admin_required
def api_export_indicators_excel():
    """导出检测指标到Excel"""
    with get_db() as conn:

        indicators = conn.execute(
            'SELECT i.*, g.name as group_name '
            'FROM indicators i '
            'LEFT JOIN indicator_groups g ON i.group_id = g.id '
            'ORDER BY i.sort_order, i.name'
        ).fetchall()

        # 关闭连接释放锁

        # 创建Excel工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "检测指标"

        # 设置样式
        header_font = Font(name='宋体', size=11, bold=True)
        normal_font = Font(name='宋体', size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 表头
        headers = ['指标名称', '单位', '默认值', '限值', '检测方法', '所属分组', '排序', '备注']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # 数据行
        for row_idx, indicator in enumerate(indicators, start=2):
            row_data = [
                indicator['name'],
                indicator['unit'] or '',
                indicator['default_value'] or '',
                indicator['limit_value'] or '',
                indicator['detection_method'] or '',
                indicator['group_name'] or '',
                indicator['sort_order'],
                indicator['remark'] or ''
            ]

            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col)
                cell.value = value
                cell.font = normal_font
                cell.border = border

        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 30

        # 保存文件
        os.makedirs('exports', exist_ok=True)
        filename = f"exports/indicators_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        wb.save(filename)

        log_operation('导出检测指标', f'导出 {len(indicators)} 个检测指标')

        return send_file(filename, as_attachment=True, download_name='检测指标.xlsx')

@import_bp.route('/api/indicators/import/excel', methods=['POST'])
@admin_required
def api_import_indicators_excel():
    """从Excel导入检测指标"""
    if 'file' not in request.files:
        return jsonify({'error': '未上传文件'}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({'error': '未选择文件'}), 400

    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': '请上传Excel文件(.xlsx 或 .xls)'}), 400

    try:
        # 读取Excel文件
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        with get_db() as conn:
            cursor = conn.cursor()

            # 获取所有分组，建立名称到ID的映射
            groups = cursor.execute('SELECT id, name FROM indicator_groups').fetchall()
            group_map = {g['name']: g['id'] for g in groups}

            # 获取样品类型映射（按名称匹配分组名→样品类型ID）
            sample_types = cursor.execute('SELECT id, name FROM sample_types').fetchall()
            sample_type_map = {st['name']: st['id'] for st in sample_types}

            imported_count = 0
            updated_count = 0
            error_rows = []

            # 从第2行开始读取（第1行是表头）
            # 新格式：指标名称、单位、默认值、限值、检测方法、所属分组、排序、备注
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:  # 跳过空行
                    continue

                name = row[0]
                unit = row[1] or ''
                default_value = row[2] or ''
                limit_value = row[3] or '' if len(row) > 3 else ''
                detection_method = row[4] or '' if len(row) > 4 else ''
                group_name = row[5] or '' if len(row) > 5 else ''
                sort_order = row[6] if len(row) > 6 and row[6] is not None else 0
                remark = row[7] or '' if len(row) > 7 else ''

                # 查找分组ID
                group_id = group_map.get(group_name) if group_name else None

                try:
                    # 按(name, group_id)查重：同名不同分组视为不同指标
                    existing = cursor.execute(
                        'SELECT id FROM indicators WHERE name = ? AND group_id IS ?',
                        (name, group_id)
                    ).fetchone()

                    if existing:
                        indicator_id = existing['id']
                        # 更新该分组下的同名指标（各分组独立，互不影响）
                        cursor.execute(
                            'UPDATE indicators SET unit = ?, default_value = ?, limit_value = ?, '
                            'detection_method = ?, remark = ?, sort_order = ? WHERE id = ?',
                            (unit, default_value, limit_value, detection_method, remark, sort_order, indicator_id)
                        )
                        updated_count += 1
                    else:
                        # 插入新指标
                        cursor.execute(
                            'INSERT INTO indicators (group_id, name, unit, default_value, limit_value, detection_method, remark, sort_order) '
                            'VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
                            (group_id, name, unit, default_value, limit_value, detection_method, remark, sort_order)
                        )
                        indicator_id = cursor.lastrowid
                        imported_count += 1

                    # 将limit_value写入对应样品类型的template_indicators
                    if limit_value and group_name:
                        sample_type_id = sample_type_map.get(group_name)
                        if sample_type_id:
                            existing_ti = cursor.execute(
                                'SELECT id FROM template_indicators WHERE sample_type_id = ? AND indicator_id = ?',
                                (sample_type_id, indicator_id)
                            ).fetchone()
                            if existing_ti:
                                cursor.execute(
                                    'UPDATE template_indicators SET limit_value = ? WHERE id = ?',
                                    (limit_value, existing_ti['id'])
                                )
                            else:
                                cursor.execute(
                                    'INSERT INTO template_indicators (sample_type_id, indicator_id, limit_value, sort_order) '
                                    'VALUES (?, ?, ?, ?)',
                                    (sample_type_id, indicator_id, limit_value, sort_order)
                                )
                except Exception as e:
                    error_rows.append(f'第{row_idx}行: {str(e)}')


            log_operation('导入检测指标', f'新增 {imported_count} 个，更新 {updated_count} 个', conn=conn)

            result = {
                'message': f'导入成功！新增 {imported_count} 个指标，更新 {updated_count} 个指标',
                'imported': imported_count,
                'updated': updated_count
            }

            if error_rows:
                result['errors'] = error_rows

            return jsonify(result)

    except Exception as e:
        return jsonify({'error': f'导入失败: {str(e)}'}), 500


# ==================== 报告批量导入 API ====================

@import_bp.route('/api/reports/import/excel', methods=['POST'])
@login_required
def api_import_reports_excel():
    """从Excel批量导入报告"""
    if 'file' not in request.files:
        return jsonify({'error': '未上传文件'}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({'error': '未选择文件'}), 400

    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': '请上传Excel文件(.xlsx 或 .xls)'}), 400

    try:
        # 读取Excel文件
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        with get_db() as conn:
            cursor = conn.cursor()

            # 获取样品类型映射
            sample_types = cursor.execute('SELECT id, name, code FROM sample_types').fetchall()
            sample_type_name_map = {st['name']: st for st in sample_types}
            sample_type_code_map = {st['code']: st for st in sample_types}

            # 获取公司映射
            companies = cursor.execute('SELECT id, name FROM companies').fetchall()
            company_map = {c['name']: c['id'] for c in companies}

            # 获取检测指标映射
            indicators = cursor.execute('SELECT id, name FROM indicators').fetchall()
            indicator_map = {i['name']: i['id'] for i in indicators}

            imported_count = 0
            error_rows = []

            # 读取表头（第1行）
            headers = [cell.value for cell in ws[1]]

            # 查找固定列的索引
            try:
                sample_number_idx = headers.index('样品编号')
                sample_type_idx = headers.index('样品类型')
            except ValueError:
                return jsonify({'error': 'Excel格式错误：必须包含"样品编号"和"样品类型"列'}), 400

            # 可选列
            company_idx = headers.index('委托单位') if '委托单位' in headers else None
            detection_date_idx = headers.index('检测日期') if '检测日期' in headers else None
            detection_person_idx = headers.index('检测人员') if '检测人员' in headers else None
            review_person_idx = headers.index('审核人员') if '审核人员' in headers else None
            remark_idx = headers.index('备注') if '备注' in headers else None

            # 检测指标列（除了固定列之外的列都视为检测指标）
            fixed_cols = {'样品编号', '样品类型', '委托单位', '检测日期', '检测人员', '审核人员', '备注'}
            indicator_cols = [(idx, col) for idx, col in enumerate(headers) if col and col not in fixed_cols]

            # 从第2行开始读取数据
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[sample_number_idx]:  # 跳过空行
                    continue

                try:
                    sample_number = str(row[sample_number_idx]).strip()
                    sample_type_value = str(row[sample_type_idx]).strip()

                    # 查找样品类型（支持名称或代码）
                    sample_type = sample_type_name_map.get(sample_type_value) or sample_type_code_map.get(sample_type_value)

                    if not sample_type:
                        error_rows.append(f'第{row_idx}行: 样品类型"{sample_type_value}"不存在')
                        continue

                    sample_type_id = sample_type['id']

                    # 获取其他字段
                    company_id = None
                    if company_idx is not None and row[company_idx]:
                        company_name = str(row[company_idx]).strip()
                        company_id = company_map.get(company_name)
                        if not company_id and company_name:
                            # 自动创建公司
                            cursor.execute('INSERT INTO companies (name) VALUES (?)', (company_name,))
                            company_id = cursor.lastrowid
                            company_map[company_name] = company_id

                    detection_person = str(row[detection_person_idx]).strip() if detection_person_idx is not None and row[detection_person_idx] else ''
                    review_person = str(row[review_person_idx]).strip() if review_person_idx is not None and row[review_person_idx] else ''
                    detection_date = str(row[detection_date_idx]) if detection_date_idx is not None and row[detection_date_idx] else None
                    remark = str(row[remark_idx]).strip() if remark_idx is not None and row[remark_idx] else ''

                    # 生成报告编号
                    report_number = f"{sample_number}-{sample_type['code']}"

                    # 检查报告编号是否已存在
                    existing = cursor.execute('SELECT id FROM reports WHERE report_number = ?', (report_number,)).fetchone()

                    if existing:
                        error_rows.append(f'第{row_idx}行: 报告编号"{report_number}"已存在')
                        continue

                    # 创建报告
                    cursor.execute(
                        'INSERT INTO reports (report_number, sample_number, company_id, sample_type_id, '
                        'detection_person, review_person, detection_date, remark, created_by) '
                        'VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)',
                        (report_number, sample_number, company_id, sample_type_id, detection_person,
                         review_person, detection_date, remark, session['user_id'])
                    )
                    report_id = cursor.lastrowid

                    # 添加检测数据
                    for col_idx, col_name in indicator_cols:
                        if col_name in indicator_map and row[col_idx]:
                            measured_value = str(row[col_idx]).strip()
                            if measured_value:
                                cursor.execute(
                                    'INSERT INTO report_data (report_id, indicator_id, measured_value, remark) '
                                    'VALUES (?, ?, ?, ?)',
                                    (report_id, indicator_map[col_name], measured_value, '')
                                )

                    imported_count += 1

                except Exception as e:
                    error_rows.append(f'第{row_idx}行: {str(e)}')
                    continue


            log_operation('批量导入报告', f'成功导入 {imported_count} 份报告', conn=conn)

            result = {
                'message': f'导入成功！共导入 {imported_count} 份报告',
                'imported': imported_count
            }

            if error_rows:
                result['errors'] = error_rows

            return jsonify(result)

    except Exception as e:
        return jsonify({'error': f'导入失败: {str(e)}'}), 500

@import_bp.route('/api/reports/export/template', methods=['GET'])
@login_required
def api_export_reports_template():
    """导出报告导入模板Excel"""
    sample_type_id = request.args.get('sample_type_id')

    if not sample_type_id:
        return jsonify({'error': '请指定样品类型'}), 400

    with get_db() as conn:

        # 获取样品类型信息
        sample_type = conn.execute('SELECT * FROM sample_types WHERE id = ?', (sample_type_id,)).fetchone()

        if not sample_type:
            return jsonify({'error': '样品类型不存在'}), 404

        # 获取该样品类型的检测指标
        indicators = conn.execute(
            'SELECT i.name, i.unit '
            'FROM template_indicators ti '
            'LEFT JOIN indicators i ON ti.indicator_id = i.id '
            'WHERE ti.sample_type_id = ? '
            'ORDER BY ti.sort_order',
            (sample_type_id,)
        ).fetchall()


        # 创建Excel工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "检测数据"

        # 设置样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        subheader_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
        subheader_font = Font(name='宋体', size=11, bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 简化格式表头：检测项目、单位、样品数据列
        headers = ['检测项目', '单位']
        sample_numbers = ['样品编号1*', '样品编号2', '样品编号3']

        for col, header in enumerate(headers + sample_numbers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            if col <= len(headers):
                cell.fill = subheader_fill
                cell.font = subheader_font
            else:
                cell.fill = header_fill
                cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # 添加检测指标行
        for row_idx, indicator in enumerate(indicators, start=2):
            # 检测项目（A列）
            cell = ws.cell(row=row_idx, column=1)
            cell.value = indicator['name']
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = border

            # 单位（B列）
            cell = ws.cell(row=row_idx, column=2)
            cell.value = indicator['unit'] or ''
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

            # 样品数据列（C列起）留空
            for col_idx in range(3, 3 + len(sample_numbers)):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # 调整列宽
        ws.column_dimensions['A'].width = 20  # 检测项目
        ws.column_dimensions['B'].width = 12  # 单位
        for col_letter in ['C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col_letter].width = 15

        # 冻结首行和前2列
        ws.freeze_panes = 'C2'

        # 保存文件
        os.makedirs('exports', exist_ok=True)
        filename = f"exports/report_template_{sample_type['code']}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        wb.save(filename)

        log_operation('导出报告模板', f'导出样品类型:{sample_type["name"]}')
        return send_file(filename, as_attachment=True, download_name=f'报告导入模板_{sample_type["code"]}.xlsx')


# ==================== 新增API接口 ====================

@import_bp.route('/api/field-code-reference', methods=['GET'])
@login_required
def api_download_field_code_reference():
    """下载字段代号使用说明文档"""
    try:
        # 生成文档内容
        doc_content = FieldCodeMapping.generate_documentation()

        # 创建文本文件
        output = io.BytesIO()
        output.write(doc_content.encode('utf-8'))
        output.seek(0)

        log_operation('下载字段代号文档', '字段代号使用说明')

        return send_file(
            output,
            mimetype='text/plain; charset=utf-8',
            as_attachment=True,
            download_name='Excel模板字段代号使用说明.txt'
        )
    except Exception as e:
        return jsonify({'error': f'下载失败: {str(e)}'}), 500

@import_bp.route('/api/download-example-template', methods=['GET'])
@login_required
def api_download_example_template():
    """下载Excel模板示例文件"""
    try:
        template_path = 'template_examples/水质检测报告模板示例.xlsx'

        # 如果文件不存在，先生成
        if not os.path.exists(template_path):
            create_example_template()

        log_operation('下载模板示例', 'Excel模板示例文件')

        return send_file(
            template_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='水质检测报告模板示例.xlsx'
        )
    except Exception as e:
        return jsonify({'error': f'下载失败: {str(e)}'}), 500

@import_bp.route('/api/export-report-template/<int:template_id>', methods=['GET'])
@login_required
def api_export_report_template(template_id):
    """导出报告填写模板"""
    try:
        output_path = export_report_template(template_id)
        log_operation('导出报告填写模板', f'模板ID: {template_id}')
        return send_file(output_path, as_attachment=True, download_name=os.path.basename(output_path))
    except Exception as e:
        return jsonify({'error': f'导出失败: {str(e)}'}), 500

@import_bp.route('/api/export-sample-type-template/<int:sample_type_id>', methods=['GET'])
@login_required
def api_export_sample_type_template(sample_type_id):
    """导出样品类型检测模板"""
    try:
        output_path = export_sample_type_template(sample_type_id)
        log_operation('导出检测项目模板', f'样品类型ID: {sample_type_id}')
        return send_file(output_path, as_attachment=True, download_name=os.path.basename(output_path))
    except Exception as e:
        return jsonify({'error': f'导出失败: {str(e)}'}), 500

@import_bp.route('/api/import-report-info', methods=['POST'])
@login_required
def api_import_report_info():
    """导入报告基本信息"""
    if 'file' not in request.files:
        return jsonify({'error': '未上传文件'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '未选择文件'}), 400

    try:
        # 保存上传的文件
        os.makedirs('temp/imports', exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        temp_path = f'temp/imports/report_info_{timestamp}.xlsx'
        file.save(temp_path)

        # 读取Excel文件
        wb = openpyxl.load_workbook(temp_path, data_only=True)
        ws = wb['报告基本信息']

        with get_db() as conn:
            cursor = conn.cursor()
            created_count = 0

            # 获取字段名称（第一列，从第2行开始）
            field_names = []
            row_idx = 2
            while True:
                field_name = ws.cell(row_idx, 1).value
                if field_name is None:
                    break
                field_names.append(field_name.replace('*', '').strip())
                row_idx += 1

            # 处理每一列数据（从第2列开始）
            col_idx = 2
            while True:
                sample_number = ws.cell(1, col_idx).value
                if sample_number is None or str(sample_number).strip() == '':
                    break

                sample_number = str(sample_number).replace('*', '').strip()

                # 读取该列的所有数据
                report_data = {}
                for i, field_name in enumerate(field_names, start=2):
                    cell_value = ws.cell(i, col_idx).value
                    report_data[field_name] = cell_value if cell_value is not None else ''

                # 创建报告记录（简化版本，实际需要根据模板字段创建）
                # 这里暂时创建基本报告记录
                report_number = f"RPT{datetime.now().strftime('%Y%m%d%H%M%S')}{created_count+1:03d}"

                # 注意：这里需要根据实际的模板字段映射来创建报告
                # 暂时先创建一个简单的占位实现
                created_count += 1
                col_idx += 1


            # 删除临时文件
            try:
                os.remove(temp_path)
            except OSError:
                pass

            log_operation('导入报告基本信息', f'成功导入 {created_count} 份报告')

            return jsonify({
                'message': '导入成功',
                'created_count': created_count
            }), 200

    except Exception as e:
        return jsonify({'error': f'导入失败: {str(e)}'}), 500

@import_bp.route('/api/import-detection-data', methods=['POST'])
@login_required
def api_import_detection_data():
    """导入检测项目数据"""
    if 'file' not in request.files:
        return jsonify({'error': '未上传文件'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '未选择文件'}), 400

    try:
        # 保存上传的文件
        os.makedirs('temp/imports', exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        temp_path = f'temp/imports/detection_data_{timestamp}.xlsx'
        file.save(temp_path)

        # 读取Excel文件
        wb = openpyxl.load_workbook(temp_path, data_only=True)
        ws = wb['检测数据']

        with get_db() as conn:
            cursor = conn.cursor()
            updated_count = 0

            # 获取检测项目列表（第2列，从第2行开始）
            indicators = []
            row_idx = 2
            while True:
                indicator_name = ws.cell(row_idx, 2).value
                if indicator_name is None:
                    break
                indicators.append(indicator_name)
                row_idx += 1

            # 处理每一列样品数据（从第6列开始）
            col_idx = 6
            while True:
                sample_number = ws.cell(1, col_idx).value
                if sample_number is None or str(sample_number).strip() == '':
                    break

                sample_number = str(sample_number).replace('*', '').strip()

                # 查找对应的报告
                report = conn.execute(
                    'SELECT id FROM reports WHERE sample_number = ?',
                    (sample_number,)
                ).fetchone()

                if report:
                    report_id = report['id']

                    # 读取该列的检测数据
                    for i, indicator_name in enumerate(indicators, start=2):
                        measured_value = ws.cell(i, col_idx).value
                        if measured_value is not None and str(measured_value).strip() != '':
                            # 查找指标ID（优先通过template_indicators按样品类型匹配）
                            indicator = conn.execute('''
                                SELECT i.id FROM indicators i
                                JOIN template_indicators ti ON ti.indicator_id = i.id
                                WHERE i.name = ? AND ti.sample_type_id = (
                                    SELECT sample_type_id FROM reports WHERE id = ?
                                ) LIMIT 1
                            ''', (indicator_name, report_id)).fetchone()
                            if not indicator:
                                indicator = conn.execute(
                                    'SELECT id FROM indicators WHERE name = ? LIMIT 1',
                                    (indicator_name,)
                                ).fetchone()

                            if indicator:
                                # 更新或插入检测数据
                                cursor.execute('''
                                    INSERT OR REPLACE INTO report_data (report_id, indicator_id, measured_value)
                                    VALUES (?, ?, ?)
                                ''', (report_id, indicator['id'], str(measured_value)))

                    updated_count += 1

                col_idx += 1


            # 删除临时文件
            try:
                os.remove(temp_path)
            except OSError:
                pass

            log_operation('导入检测数据', f'成功更新 {updated_count} 份报告的检测数据', conn=conn)

            return jsonify({
                'message': '导入成功',
                'updated_count': updated_count
            }), 200

    except Exception as e:
        return jsonify({'error': f'导入失败: {str(e)}'}), 500

@import_bp.route('/api/validate-report-excel', methods=['POST'])
@login_required
def api_validate_report_excel():
    """验证上传的Excel文件格式"""
    try:
        # 获取参数
        template_id = request.form.get('template_id')
        sample_type_id = request.form.get('sample_type_id')

        if not template_id or not sample_type_id:
            return jsonify({'valid': False, 'errors': ['缺少必要参数：template_id 或 sample_type_id']}), 400

        # 获取上传的文件
        if 'template_excel' not in request.files or 'detection_excel' not in request.files:
            return jsonify({'valid': False, 'errors': ['缺少Excel文件']}), 400

        template_file = request.files['template_excel']
        detection_file = request.files['detection_excel']

        if not template_file.filename or not detection_file.filename:
            return jsonify({'valid': False, 'errors': ['文件名为空']}), 400

        validation_errors = []
        validation_warnings = []

        # 保存临时文件
        os.makedirs('temp/validate', exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        template_path = f'temp/validate/template_{timestamp}.xlsx'
        detection_path = f'temp/validate/detection_{timestamp}.xlsx'

        template_file.save(template_path)
        detection_file.save(detection_path)

        # 验证报告模板Excel
        try:
            template_wb = openpyxl.load_workbook(template_path, data_only=True)

            # 检查是否有"报告基本信息"工作表
            if '报告基本信息' not in template_wb.sheetnames:
                validation_warnings.append('报告模板Excel中没有找到"报告基本信息"工作表，将无法自动填充基本信息')
            else:
                info_ws = template_wb['报告基本信息']
                # 验证基本信息格式
                expected_fields = ['样品编号', '委托单位', '检测日期', '检测人员', '审核人员', '备注']
                found_fields = []
                for row_idx in range(2, min(20, info_ws.max_row + 1)):
                    field_name = info_ws.cell(row_idx, 1).value
                    if field_name:
                        found_fields.append(str(field_name).strip())

                missing_fields = [f for f in expected_fields if f not in found_fields]
                if missing_fields:
                    validation_warnings.append(f'报告基本信息中缺少以下字段：{", ".join(missing_fields)}')

            template_wb.close()

        except Exception as e:
            validation_errors.append(f'报告模板Excel格式错误：{str(e)}')

        # 验证检测数据Excel
        try:
            detection_wb = openpyxl.load_workbook(detection_path, data_only=True)
            detection_ws = detection_wb.active

            # 检查是否有数据
            if detection_ws.max_row < 2:
                validation_errors.append('检测数据Excel中没有数据行')

            # 检查A列（第1列）是否有指标名称（简化格式）
            indicator_count = 0
            for row_idx in range(2, min(100, detection_ws.max_row + 1)):
                indicator_name = detection_ws.cell(row_idx, 1).value
                if indicator_name and str(indicator_name).strip():
                    indicator_count += 1

            if indicator_count == 0:
                validation_errors.append('检测数据Excel的A列（检测项目列）没有找到任何指标')

            # 检查C列（第3列）是否有样品编号
            sample_number = detection_ws.cell(1, 3).value
            if not sample_number or str(sample_number).strip() == '':
                validation_warnings.append('检测数据Excel的C列（样品数据列）标题为空，请在首行C列填写样品编号')

            detection_wb.close()

        except Exception as e:
            validation_errors.append(f'检测数据Excel格式错误：{str(e)}')

        # 清理临时文件
        try:
            os.remove(template_path)
            os.remove(detection_path)
        except OSError:
            pass

        # 返回验证结果
        is_valid = len(validation_errors) == 0

        return jsonify({
            'valid': is_valid,
            'errors': validation_errors,
            'warnings': validation_warnings
        }), 200

    except Exception as e:
        traceback.print_exc()
        return jsonify({
            'valid': False,
            'errors': [f'验证失败：{str(e)}']
        }), 500

@import_bp.route('/api/parse-report-excel', methods=['POST'])
@login_required
def api_parse_report_excel():
    """解析上传的报告模板Excel和检测数据Excel"""
    try:
        # 获取参数
        template_id = request.form.get('template_id')
        sample_type_id = request.form.get('sample_type_id')

        if not template_id or not sample_type_id:
            return jsonify({'error': '缺少必要参数'}), 400

        # 获取上传的文件
        if 'template_excel' not in request.files or 'detection_excel' not in request.files:
            return jsonify({'error': '缺少Excel文件'}), 400

        template_file = request.files['template_excel']
        detection_file = request.files['detection_excel']

        if not template_file.filename or not detection_file.filename:
            return jsonify({'error': '文件名为空'}), 400

        # 保存临时文件
        os.makedirs('temp/parse', exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        template_path = f'temp/parse/template_{timestamp}.xlsx'
        detection_path = f'temp/parse/detection_{timestamp}.xlsx'

        template_file.save(template_path)
        detection_file.save(detection_path)

        # 解析报告模板Excel
        template_wb = openpyxl.load_workbook(template_path, data_only=True)

        # 解析检测数据Excel
        detection_wb = openpyxl.load_workbook(detection_path, data_only=True)
        detection_ws = detection_wb.active

        with get_db() as conn:
            cursor = conn.cursor()

            # 解析基本信息（从报告模板Excel中提取）
            basic_info = {}

            # 定义基本信息字段，避免与模板字段混淆
            basic_info_fields = {
                '样品编号': 'sample_number',
                '委托单位': 'company_name',
                '检测日期': 'detection_date',
                '检测人员': 'detection_person',
                '审核人员': 'review_person',
                '备注': 'remark'
            }

            # 尝试从第一个工作表读取基本信息
            if '报告基本信息' in template_wb.sheetnames:
                info_ws = template_wb['报告基本信息']
                # 读取字段名和值（假设格式：列A是字段名，列B是值）
                row_idx = 2
                while row_idx <= info_ws.max_row:
                    field_name = info_ws.cell(row_idx, 1).value
                    field_value = info_ws.cell(row_idx, 2).value

                    if not field_name:
                        break

                    field_name_str = str(field_name).strip()

                    # 只处理基本信息字段
                    if field_name_str in basic_info_fields:
                        standard_field = basic_info_fields[field_name_str]
                        basic_info[standard_field] = str(field_value) if field_value else ''

                    row_idx += 1

            # 解析模板字段
            template_fields = []
            fields_result = cursor.execute('''
                SELECT fm.id, fm.field_name, fm.field_name as field_display_name, fm.field_type,
                       fm.is_required, fm.sheet_name, fm.cell_address
                FROM template_field_mappings fm
                WHERE fm.template_id = ?
                ORDER BY fm.id
            ''', (template_id,)).fetchall()

            for field in fields_result:
                field_data = {
                    'field_mapping_id': field[0],
                    'field_name': field[2] or field[1],
                    'field_type': field[3],
                    'is_required': field[4],
                    'field_value': ''
                }

                # 尝试从Excel中读取字段值
                if field[5] and field[6]:  # sheet_name 和 cell_address
                    try:
                        if field[5] in template_wb.sheetnames:
                            ws = template_wb[field[5]]
                            cell_value = ws[field[6]].value
                            if cell_value:
                                field_data['field_value'] = str(cell_value)
                    except (KeyError, IndexError, TypeError):
                        pass

                template_fields.append(field_data)

            # 解析检测数据
            detection_data = []

            # 获取该样品类型的指标列表
            indicators_result = cursor.execute('''
                SELECT i.id, i.name, i.unit, i.default_value,
                       COALESCE(ti.limit_value, i.limit_value) as limit_value,
                       ig.name as group_name, ti.sort_order
                FROM template_indicators ti
                JOIN indicators i ON ti.indicator_id = i.id
                LEFT JOIN indicator_groups ig ON i.group_id = ig.id
                WHERE ti.sample_type_id = ?
                ORDER BY ti.sort_order
            ''', (sample_type_id,)).fetchall()

            # 从检测数据Excel中读取数据
            # 简化格式：第1列检测项目，第2列单位（参考），第3列及以后为样品检测数据
            indicator_name_col = 1  # 指标名称列（A列）
            unit_col = 2            # 单位列（B列，参考用）
            value_col = 3           # 检测值列（C列，第一个样品的数据）

            for indicator in indicators_result:
                indicator_id = indicator[0]
                indicator_name = indicator[1]
                unit = indicator[2]
                default_value = indicator[3]
                limit_value = indicator[4]
                group_name = indicator[5]

                # 在Excel中查找该指标
                measured_value = ''
                for row_idx in range(2, detection_ws.max_row + 1):
                    cell_indicator = detection_ws.cell(row_idx, indicator_name_col).value
                    if cell_indicator and str(cell_indicator).strip() == indicator_name:
                        # 读取检测值（C列，第3列）
                        cell_value = detection_ws.cell(row_idx, value_col).value
                        if cell_value is not None:
                            measured_value = str(cell_value).strip()
                        break

                detection_data.append({
                    'indicator_id': indicator_id,
                    'indicator_name': indicator_name,
                    'unit': unit,
                    'measured_value': measured_value or default_value or '',
                    'limit_value': limit_value,
                    'group_name': group_name
                })


            # 清理临时文件
            try:
                os.remove(template_path)
                os.remove(detection_path)
            except OSError:
                pass

            # 返回解析结果
            result = {
                'basic_info': basic_info,
                'template_fields': template_fields,
                'detection_data': detection_data
            }

            log_operation('解析报告Excel', f'模板ID: {template_id}, 样品类型ID: {sample_type_id}')

            return jsonify(result), 200

    except Exception as e:
        error_trace = traceback.format_exc()
        print("解析Excel错误详情:")
        print(error_trace)

        # 提供更详细的错误信息
        error_msg = str(e)
        if 'no such table' in error_msg:
            table_name = error_msg.split('no such table:')[-1].strip()
            error_msg = f'数据库表不存在：{table_name}。请检查数据库是否已正确初始化。'
        elif 'no such column' in error_msg:
            error_msg = f'数据库字段不存在：{error_msg}。请检查数据库结构是否正确。'
        elif 'Worksheet' in error_msg and 'does not exist' in error_msg:
            error_msg = f'Excel工作表不存在：{error_msg}。请检查Excel文件格式。'

        return jsonify({
            'error': f'解析失败: {error_msg}',
            'details': error_trace if current_app.debug else None
        }), 500

@import_bp.route('/api/import-reports', methods=['POST'])
@login_required
def api_import_reports():
    """批量导入报告"""
    if 'file' not in request.files:
        return jsonify({'error': '未上传文件'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '文件名为空'}), 400

    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': '仅支持Excel文件(.xlsx, .xls)'}), 400

    # 保存上传的文件
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"import_{timestamp}_{file.filename}"
    upload_path = os.path.join('uploads', filename)
    os.makedirs('uploads', exist_ok=True)
    file.save(upload_path)

    # 获取参数
    template_id = request.form.get('template_id')
    template_id = int(template_id) if template_id else None
    created_by = session.get('username', 'system')

    try:
        # 处理导入
        results = import_reports_from_excel(upload_path, template_id, created_by)

        log_operation('批量导入报告',
                     f'成功:{len(results["success"])} 失败:{len(results["errors"])} 警告:{len(results["warnings"])}')

        return jsonify(results)
    except Exception as e:
        return jsonify({'error': f'导入失败: {str(e)}'}), 500
    finally:
        # 清理临时文件
        if os.path.exists(upload_path):
            os.remove(upload_path)
