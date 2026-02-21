"""
原始数据管理 API Blueprint
从 app_v2.py 提取的原始数据管理相关路由
"""
from flask import Blueprint, request, jsonify, send_file, session
from models_v2 import get_db, get_db_connection
from auth import login_required, admin_required, log_operation
from raw_data_importer import RawDataImporter
from raw_data_converter import convert_raw_excel
from raw_data_validator import RawDataValidator, validate_samples, validate_from_database
from raw_data_template_generator import generate_raw_data_template
from werkzeug.utils import secure_filename
import os
import json
import re
import openpyxl
import pandas as pd
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime

raw_data_bp = Blueprint('raw_data_bp', __name__)

UPLOAD_FOLDER = 'temp/uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def allowed_file(filename):
    """检查文件扩展名是否允许"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@raw_data_bp.route('/api/raw-data/upload', methods=['POST'])
@login_required
def api_raw_data_upload():
    """上传并导入Excel原始数据"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': '未选择文件'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': '未选择文件'}), 400

        if not allowed_file(file.filename):
            return jsonify({'error': '文件格式不支持，仅支持.xlsx和.xls格式'}), 400

        # 保存上传的文件
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        saved_filename = f"{timestamp}_{filename}"
        filepath = os.path.join(UPLOAD_FOLDER, saved_filename)
        file.save(filepath)

        # 获取处理选项
        on_duplicate = request.form.get('on_duplicate', 'skip')
        duplicate_decisions = None
        dd_json = request.form.get('duplicate_decisions', '')
        if dd_json:
            try:
                duplicate_decisions = json.loads(dd_json)
            except (json.JSONDecodeError, TypeError):
                pass

        # 导入数据
        importer = RawDataImporter()
        result = importer.import_excel(filepath, on_duplicate=on_duplicate,
                                       duplicate_decisions=duplicate_decisions)

        # pause模式返回重复列表时保留文件路径
        if result.get('paused'):
            result['saved_file'] = saved_filename
            return jsonify(result)

        # 删除临时文件
        try:
            os.remove(filepath)
        except OSError:
            pass

        # 记录操作日志
        if result['success']:
            log_operation(
                '导入原始数据',
                f"导入成功: {result['success_count']}条，跳过: {result['skip_count']}条"
            )

        return jsonify(result)

    except Exception as e:
        return jsonify({'error': f'上传失败: {str(e)}'}), 500

CONVERT_FOLDER = 'temp/convert'
os.makedirs(CONVERT_FOLDER, exist_ok=True)


@raw_data_bp.route('/api/raw-data/convert-preview', methods=['POST'])
@login_required
def api_raw_data_convert_preview():
    """上传原始检测Excel，预览转换结果"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': '未选择文件'}), 400

        file = request.files['file']
        if file.filename == '' or not allowed_file(file.filename):
            return jsonify({'error': '文件格式不支持，仅支持.xlsx格式'}), 400

        skip_blank = request.form.get('skip_blank', 'true') == 'true'

        # 保存上传文件
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        saved_filename = f"convert_{timestamp}_{filename}"
        filepath = os.path.join(CONVERT_FOLDER, saved_filename)
        file.save(filepath)

        # 执行转换
        output_path = os.path.join(CONVERT_FOLDER, f"import_{timestamp}_{filename}")
        result = convert_raw_excel(filepath, output_path=output_path, skip_blank_samples=skip_blank)

        # 删除源文件
        try:
            os.remove(filepath)
        except OSError:
            pass

        if not result['success']:
            return jsonify({'error': result['message']}), 400

        # 构建预览数据
        preview_samples = []
        for s in result['samples']:
            sid = s['样品编号']
            sample_data = result['data'].get(sid, {})
            # 取前 5 个有值的指标作为预览
            preview_indicators = {}
            count = 0
            for p in result['parameters']:
                v = sample_data.get(p)
                if v is not None and count < 5:
                    preview_indicators[p] = v
                    count += 1
            preview_samples.append({
                'sample_number': sid,
                'company_name': s.get('被检单位', ''),
                'plant_name': s.get('被检水厂', ''),
                'sample_type': s.get('样品类型', ''),
                'sampling_date': s.get('采样日期', ''),
                'indicator_count': len([p for p in result['parameters'] if sample_data.get(p)]),
                'preview_indicators': preview_indicators,
                'all_indicators': {p: sample_data.get(p, '') for p in result['parameters']},
            })

        # 执行校核（方案C：转换预览阶段预校核）
        validation_results = []
        try:
            validation_results = validate_samples(result['samples'], result['data'])
        except Exception:
            pass  # 校核失败不影响转换预览

        return jsonify({
            'success': True,
            'message': result['message'],
            'converted_file': saved_filename,
            'output_file': os.path.basename(output_path),
            'sample_count': result['sample_count'],
            'param_count': result['param_count'],
            'parameters': result['parameters'],
            'samples': preview_samples,
            'validation': validation_results,
        })

    except Exception as e:
        return jsonify({'error': f'转换失败: {str(e)}'}), 500


@raw_data_bp.route('/api/raw-data/convert-import', methods=['POST'])
@login_required
def api_raw_data_convert_import():
    """将预览确认的转换结果导入系统"""
    try:
        data = request.get_json()
        if not data or 'output_file' not in data:
            return jsonify({'error': '缺少转换文件信息'}), 400

        output_file = secure_filename(data['output_file'])
        filepath = os.path.join(CONVERT_FOLDER, output_file)

        if not os.path.exists(filepath):
            return jsonify({'error': '转换文件已过期，请重新上传'}), 404

        on_duplicate = data.get('on_duplicate', 'skip')
        sample_edits = data.get('sample_edits', {})
        selected_samples = data.get('selected_samples')

        # 如果有编辑数据，先修改转换后的 Excel
        if sample_edits:
            try:
                wb = openpyxl.load_workbook(filepath)
                ws = wb.active
                # 构建样品编号→列索引的映射（第1行 B列起为样品编号）
                sid_col_map = {}
                for c in range(2, ws.max_column + 1):
                    val = ws.cell(1, c).value
                    if val:
                        sid_col_map[str(val).strip()] = c
                # 构建字段名→行索引的映射（A列第2行起为字段名）
                field_row_map = {}
                for r in range(2, ws.max_row + 1):
                    val = ws.cell(r, 1).value
                    if val:
                        field_row_map[str(val).strip()] = r
                # 字段名映射: 前端key → Excel字段名
                key_to_field = {
                    'company_name': '被检单位',
                    'plant_name': '被检水厂',
                    'sample_type': '样品类型',
                    'sampling_date': '采样日期',
                }
                for sid, edits in sample_edits.items():
                    col = sid_col_map.get(sid)
                    if not col:
                        continue
                    # 样品类型 "名称|代码" 格式写入Excel时只保留名称
                    if 'sample_type' in edits and '|' in str(edits.get('sample_type', '')):
                        edits['sample_type'] = edits['sample_type'].split('|', 1)[0]
                    for key, field_name in key_to_field.items():
                        row = field_row_map.get(field_name)
                        if row and key in edits:
                            ws.cell(row, col).value = edits[key]
                    # 应用检测指标编辑
                    ind_edits = edits.get('indicators', {})
                    for param, value in ind_edits.items():
                        row = field_row_map.get(param)
                        if row:
                            ws.cell(row, col).value = value if value != '' else None
                # 删除未勾选的样品列（从右往左删避免索引偏移）
                if selected_samples is not None:
                    cols_to_delete = sorted(
                        [c for sid, c in sid_col_map.items() if sid not in selected_samples],
                        reverse=True
                    )
                    for c in cols_to_delete:
                        ws.delete_cols(c)

                wb.save(filepath)
                wb.close()
            except Exception as e:
                return jsonify({'error': f'应用编辑失败: {str(e)}'}), 500

        # 使用现有导入器导入（宽松模式，不要求字段完全匹配）
        importer = RawDataImporter()
        result = importer.import_excel(filepath, on_duplicate=on_duplicate, strict_columns=False)

        # 清理转换文件
        try:
            os.remove(filepath)
        except OSError:
            pass
        # 也清理同时间戳的其他文件
        for f in os.listdir(CONVERT_FOLDER):
            fpath = os.path.join(CONVERT_FOLDER, f)
            try:
                if os.path.isfile(fpath):
                    os.remove(fpath)
            except OSError:
                pass

        if result['success']:
            log_operation(
                '转换导入原始数据',
                f"转换导入成功: {result['success_count']}条，跳过: {result['skip_count']}条"
            )

        return jsonify(result)

    except Exception as e:
        return jsonify({'error': f'导入失败: {str(e)}'}), 500


@raw_data_bp.route('/api/raw-data/columns', methods=['GET'])
@login_required
def api_raw_data_columns():
    """获取当前系统的列名配置"""
    try:
        importer = RawDataImporter()
        columns = importer.get_column_list()

        if columns is None:
            return jsonify({'columns': None, 'message': '系统尚未初始化，请先导入数据'})

        return jsonify({'columns': columns})

    except Exception as e:
        return jsonify({'error': f'获取列名失败: {str(e)}'}), 500

@raw_data_bp.route('/api/raw-data/download-template', methods=['GET'])
@login_required
def api_raw_data_download_template():
    """下载原始数据导入模板"""
    try:
        # 生成模板
        template_path = generate_raw_data_template()

        # 生成下载文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        download_name = f'原始数据导入模板_{timestamp}.xlsx'

        # 记录操作日志
        log_operation('下载导入模板', '下载原始数据导入模板')

        # 发送文件
        return send_file(
            template_path,
            as_attachment=True,
            download_name=download_name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({'error': f'生成模板失败: {str(e)}'}), 500

@raw_data_bp.route('/api/raw-data/search', methods=['POST'])
@login_required
def api_raw_data_search():
    """根据样品编号精确查询原始数据"""
    try:
        data = request.json
        sample_number = data.get('sample_number', '').strip()

        if not sample_number:
            return jsonify({'error': '样品编号不能为空'}), 400

        with get_db() as conn:
            cursor = conn.cursor()

            # 查询主记录
            cursor.execute('''
                SELECT id, sample_number, report_number, company_name, plant_name, sample_type, sampling_date,
                       created_at, updated_at
                FROM raw_data_records
                WHERE sample_number = ?
            ''', (sample_number,))

            record = cursor.fetchone()

            if not record:
                return jsonify({'found': False, 'message': '未找到该样品编号的数据'})

            record_id = record[0]
            record_data = {
                'id': record[0],
                'sample_number': record[1],
                'report_number': record[2],
                'company_name': record[3],
                'plant_name': record[4],
                'sample_type': record[5],
                'sampling_date': record[6],
                'created_at': record[7],
                'updated_at': record[8]
            }

            # 查询检测指标数据
            cursor.execute('''
                SELECT column_name, value
                FROM raw_data_values
                WHERE record_id = ?
                ORDER BY id
            ''', (record_id,))

            indicators = {}
            for row in cursor.fetchall():
                indicators[row[0]] = row[1]


            return jsonify({
                'found': True,
                'data': record_data,
                'indicators': indicators
            })

    except Exception as e:
        return jsonify({'error': f'查询失败: {str(e)}'}), 500

@raw_data_bp.route('/api/raw-data/search-by-company', methods=['POST'])
@login_required
def api_raw_data_search_by_company():
    """根据被检单位模糊查询原始数据列表"""
    try:
        data = request.json
        company_name = data.get('company_name', '').strip()

        if not company_name:
            return jsonify({'error': '被检单位不能为空'}), 400

        with get_db() as conn:
            cursor = conn.cursor()

            # 模糊查询主记录
            cursor.execute('''
                SELECT id, sample_number, company_name, plant_name, sample_type, sampling_date,
                       created_at, updated_at
                FROM raw_data_records
                WHERE company_name LIKE ?
                ORDER BY company_name, plant_name, sampling_date DESC
            ''', (f'%{company_name}%',))

            records = cursor.fetchall()

            if not records:
                return jsonify({'found': False, 'message': '未找到匹配的数据', 'records': []})

            result_list = []
            for record in records:
                result_list.append({
                    'id': record[0],
                    'sample_number': record[1],
                    'company_name': record[2],
                    'plant_name': record[3],
                    'sample_type': record[4],
                    'sampling_date': record[5],
                    'created_at': record[6],
                    'updated_at': record[7]
                })

            return jsonify({
                'found': True,
                'count': len(result_list),
                'records': result_list
            })

    except Exception as e:
        return jsonify({'error': f'查询失败: {str(e)}'}), 500

@raw_data_bp.route('/api/raw-data/search-by-plant', methods=['POST'])
@login_required
def api_raw_data_search_by_plant():
    """根据被检水厂模糊查询原始数据列表"""
    try:
        data = request.json
        plant_name = data.get('plant_name', '').strip()

        if not plant_name:
            return jsonify({'error': '被检水厂不能为空'}), 400

        with get_db() as conn:
            cursor = conn.cursor()

            # 模糊查询主记录
            cursor.execute('''
                SELECT id, sample_number, company_name, plant_name, sample_type, sampling_date,
                       created_at, updated_at
                FROM raw_data_records
                WHERE plant_name LIKE ?
                ORDER BY company_name, plant_name, sampling_date DESC
            ''', (f'%{plant_name}%',))

            records = cursor.fetchall()

            if not records:
                return jsonify({'found': False, 'message': '未找到匹配的数据', 'records': []})

            result_list = []
            for record in records:
                result_list.append({
                    'id': record[0],
                    'sample_number': record[1],
                    'company_name': record[2],
                    'plant_name': record[3],
                    'sample_type': record[4],
                    'sampling_date': record[5],
                    'created_at': record[6],
                    'updated_at': record[7]
                })

            return jsonify({
                'found': True,
                'count': len(result_list),
                'records': result_list
            })

    except Exception as e:
        return jsonify({'error': f'查询失败: {str(e)}'}), 500

@raw_data_bp.route('/api/raw-data/search-companies', methods=['POST'])
@login_required
def api_raw_data_search_companies():
    """根据关键词模糊查找被检单位列表"""
    try:
        data = request.json
        keyword = data.get('keyword', '').strip()

        if not keyword:
            return jsonify({'error': '搜索关键词不能为空'}), 400

        with get_db() as conn:
            cursor = conn.cursor()

            # 模糊查询所有匹配的单位
            cursor.execute('''
                SELECT DISTINCT company_name
                FROM raw_data_records
                WHERE company_name LIKE ?
                ORDER BY company_name
            ''', (f'%{keyword}%',))

            companies = [row[0] for row in cursor.fetchall() if row[0]]

            return jsonify({
                'companies': companies,
                'count': len(companies)
            })

    except Exception as e:
        return jsonify({'error': f'查询失败: {str(e)}'}), 500

@raw_data_bp.route('/api/raw-data/get-plants', methods=['POST'])
@login_required
def api_raw_data_get_plants():
    """根据被检单位获取水厂列表"""
    try:
        data = request.json
        company_name = data.get('company_name', '').strip()

        if not company_name:
            return jsonify({'error': '被检单位不能为空'}), 400

        with get_db() as conn:
            cursor = conn.cursor()

            # 查询该单位下的所有水厂
            cursor.execute('''
                SELECT DISTINCT plant_name
                FROM raw_data_records
                WHERE company_name = ?
                ORDER BY plant_name
            ''', (company_name,))

            plants = [row[0] for row in cursor.fetchall() if row[0]]

            return jsonify({
                'plants': plants,
                'count': len(plants)
            })

    except Exception as e:
        return jsonify({'error': f'查询失败: {str(e)}'}), 500

@raw_data_bp.route('/api/raw-data/get-sample-types', methods=['POST'])
@login_required
def api_raw_data_get_sample_types():
    """根据被检单位和水厂获取样品类型列表"""
    try:
        data = request.json
        company_name = data.get('company_name', '').strip()
        plant_names = data.get('plant_names', [])  # 可以是多个水厂

        if not company_name:
            return jsonify({'error': '被检单位不能为空'}), 400

        with get_db() as conn:
            cursor = conn.cursor()

            if plant_names:
                # 查询指定水厂的样品类型
                placeholders = ','.join(['?' for _ in plant_names])
                query = f'''
                    SELECT DISTINCT sample_type
                    FROM raw_data_records
                    WHERE company_name = ? AND plant_name IN ({placeholders})
                    ORDER BY sample_type
                '''
                cursor.execute(query, [company_name] + plant_names)
            else:
                # 查询该单位下所有样品类型
                cursor.execute('''
                    SELECT DISTINCT sample_type
                    FROM raw_data_records
                    WHERE company_name = ?
                    ORDER BY sample_type
                ''', (company_name,))

            sample_types = [row[0] for row in cursor.fetchall() if row[0]]

            return jsonify({
                'sample_types': sample_types,
                'count': len(sample_types)
            })

    except Exception as e:
        return jsonify({'error': f'查询失败: {str(e)}'}), 500

@raw_data_bp.route('/api/raw-data/search-by-filters', methods=['POST'])
@login_required
def api_raw_data_search_by_filters():
    """根据单位、水厂、样品类型组合查询（支持录入时间筛选）"""
    try:
        data = request.json
        company_name = data.get('company_name', '').strip()
        plant_names = data.get('plant_names', [])
        sample_types = data.get('sample_types', [])
        created_start = data.get('created_start', '').strip()  # 录入开始日期
        created_end = data.get('created_end', '').strip()      # 录入结束日期

        if not company_name:
            return jsonify({'error': '被检单位不能为空'}), 400

        with get_db() as conn:
            cursor = conn.cursor()

            # 构建查询条件
            conditions = ['company_name = ?']
            params = [company_name]

            if plant_names:
                placeholders = ','.join(['?' for _ in plant_names])
                conditions.append(f'plant_name IN ({placeholders})')
                params.extend(plant_names)

            if sample_types:
                placeholders = ','.join(['?' for _ in sample_types])
                conditions.append(f'sample_type IN ({placeholders})')
                params.extend(sample_types)

            # 添加录入时间筛选条件
            if created_start:
                conditions.append('DATE(created_at) >= ?')
                params.append(created_start)

            if created_end:
                conditions.append('DATE(created_at) <= ?')
                params.append(created_end)

            where_clause = ' AND '.join(conditions)

            query = f'''
                SELECT id, sample_number, report_number, company_name, plant_name, sample_type, sampling_date,
                       created_at, updated_at
                FROM raw_data_records
                WHERE {where_clause}
                ORDER BY created_at DESC, company_name, plant_name, sampling_date DESC
            '''

            cursor.execute(query, params)
            records = cursor.fetchall()

            if not records:
                return jsonify({'found': False, 'message': '未找到匹配的数据', 'records': []})

            result_list = []
            for record in records:
                result_list.append({
                    'id': record[0],
                    'sample_number': record[1],
                    'report_number': record[2],
                    'company_name': record[3],
                    'plant_name': record[4],
                    'sample_type': record[5],
                    'sampling_date': record[6],
                    'created_at': record[7],
                    'updated_at': record[8]
                })

            return jsonify({
                'found': True,
                'count': len(result_list),
                'records': result_list
            })

    except Exception as e:
        return jsonify({'error': f'查询失败: {str(e)}'}), 500

@raw_data_bp.route('/api/raw-data/search-by-time', methods=['POST'])
@login_required
def api_raw_data_search_by_time():
    """按录入时间查询所有样品数据"""
    try:
        data = request.json
        created_start = data.get('created_start', '').strip()
        created_end = data.get('created_end', '').strip()

        if not created_start and not created_end:
            return jsonify({'error': '请至少选择开始日期或结束日期'}), 400

        with get_db() as conn:
            cursor = conn.cursor()

            # 构建查询条件
            conditions = []
            params = []

            if created_start:
                conditions.append('DATE(created_at) >= ?')
                params.append(created_start)

            if created_end:
                conditions.append('DATE(created_at) <= ?')
                params.append(created_end)

            where_clause = ' AND '.join(conditions)

            query = f'''
                SELECT id, sample_number, report_number, company_name, plant_name, sample_type, sampling_date,
                       created_at, updated_at
                FROM raw_data_records
                WHERE {where_clause}
                ORDER BY created_at DESC, sample_number
            '''

            cursor.execute(query, params)
            records = cursor.fetchall()

            if not records:
                return jsonify({'found': False, 'message': '未找到匹配的数据', 'records': []})

            result_list = []
            for record in records:
                result_list.append({
                    'id': record[0],
                    'sample_number': record[1],
                    'report_number': record[2],
                    'company_name': record[3],
                    'plant_name': record[4],
                    'sample_type': record[5],
                    'sampling_date': record[6],
                    'created_at': record[7],
                    'updated_at': record[8]
                })

            return jsonify({
                'found': True,
                'count': len(result_list),
                'records': result_list
            })

    except Exception as e:
        return jsonify({'error': f'查询失败: {str(e)}'}), 500

@raw_data_bp.route('/api/raw-data/detail/<int:record_id>', methods=['GET'])
@login_required
def api_raw_data_detail(record_id):
    """获取原始数据详情"""
    with get_db() as conn:
        try:
            cursor = conn.cursor()

            # 查询主记录
            cursor.execute('''
                SELECT id, sample_number, report_number, company_name, plant_name, sample_type, sampling_date,
                       created_at, updated_at
                FROM raw_data_records
                WHERE id = ?
            ''', (record_id,))

            record = cursor.fetchone()

            if not record:
                return jsonify({'error': '记录不存在'}), 404

            record_data = {
                'id': record[0],
                'sample_number': record[1],
                'report_number': record[2],
                'company_name': record[3],
                'plant_name': record[4],
                'sample_type': record[5],
                'sampling_date': record[6],
                'created_at': record[7],
                'updated_at': record[8]
            }

            # 查询检测指标数据
            cursor.execute('''
                SELECT column_name, value
                FROM raw_data_values
                WHERE record_id = ?
                ORDER BY id
            ''', (record_id,))

            indicators = {}
            for row in cursor.fetchall():
                indicators[row[0]] = row[1]


            return jsonify({
                'data': record_data,
                'indicators': indicators
            })

        except Exception as e:
            return jsonify({'error': f'获取详情失败: {str(e)}'}), 500

@raw_data_bp.route('/api/raw-data/update/<int:record_id>', methods=['PUT'])
@login_required
def api_raw_data_update(record_id):
    """更新原始数据记录"""
    try:
        data = request.json

        # 提取基础字段
        sample_number = data.get('sample_number', '').strip()
        company_name = data.get('company_name', '').strip()
        plant_name = data.get('plant_name', '').strip()
        sample_type = data.get('sample_type', '').strip()
        sampling_date = data.get('sampling_date', '').strip()
        indicators = data.get('indicators', {})

        if not sample_number:
            return jsonify({'error': '样品编号不能为空'}), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        # 检查记录是否存在
        cursor.execute('SELECT id FROM raw_data_records WHERE id = ?', (record_id,))
        if not cursor.fetchone():
            conn.close()
            return jsonify({'error': '记录不存在'}), 404

        # 检查样品编号是否与其他记录重复
        cursor.execute('SELECT id FROM raw_data_records WHERE sample_number = ? AND id != ?',
                      (sample_number, record_id))
        if cursor.fetchone():
            conn.close()
            return jsonify({'error': f'样品编号"{sample_number}"已被其他记录使用'}), 400

        # 更新主记录
        cursor.execute('''
            UPDATE raw_data_records
            SET sample_number = ?, company_name = ?, plant_name = ?,
                sample_type = ?, sampling_date = ?, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (sample_number, company_name, plant_name, sample_type, sampling_date, record_id))

        # 删除旧的检测值数据
        cursor.execute('DELETE FROM raw_data_values WHERE record_id = ?', (record_id,))

        # 插入新的检测值数据
        for column_name, value in indicators.items():
            if value is not None and str(value).strip():
                cursor.execute('''
                    INSERT INTO raw_data_values (record_id, column_name, value)
                    VALUES (?, ?, ?)
                ''', (record_id, column_name, str(value).strip()))

        conn.commit()
        conn.close()

        return jsonify({'message': '更新成功'})

    except Exception as e:
        if conn:
            conn.rollback()
            conn.close()
        return jsonify({'error': f'更新失败: {str(e)}'}), 500

@raw_data_bp.route('/api/raw-data/delete/<int:record_id>', methods=['DELETE'])
@login_required
def api_raw_data_delete(record_id):
    """删除原始数据记录"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # 检查记录是否存在
        cursor.execute('SELECT sample_number FROM raw_data_records WHERE id = ?', (record_id,))
        record = cursor.fetchone()

        if not record:
            conn.close()
            return jsonify({'error': '记录不存在'}), 404

        # 删除记录（级联删除会自动删除关联的检测值）
        cursor.execute('DELETE FROM raw_data_records WHERE id = ?', (record_id,))

        conn.commit()
        conn.close()

        return jsonify({'message': f'已删除样品编号"{record[0]}"的记录'})

    except Exception as e:
        if conn:
            conn.rollback()
            conn.close()
        return jsonify({'error': f'删除失败: {str(e)}'}), 500

@raw_data_bp.route('/api/raw-data/export-single', methods=['POST'])
@login_required
def api_raw_data_export_single():
    """导出单条原始数据记录"""
    try:
        data = request.json
        sample_number = data.get('sample_number', '').strip()
        export_format = data.get('format', 'excel')  # excel 或 csv

        if not sample_number:
            return jsonify({'error': '样品编号不能为空'}), 400

        with get_db() as conn:
            cursor = conn.cursor()

            # 查询数据
            cursor.execute('''
                SELECT id, sample_number, company_name, plant_name, sample_type, sampling_date
                FROM raw_data_records
                WHERE sample_number = ?
            ''', (sample_number,))

            record = cursor.fetchone()

            if not record:
                return jsonify({'error': '未找到该样品编号的数据'}), 404

            record_id = record[0]

            # 获取列名顺序
            cursor.execute('''
                SELECT column_name
                FROM raw_data_column_schema
                ORDER BY column_order
            ''')
            columns = [row[0] for row in cursor.fetchall()]

            # 构建数据行
            data_row = {
                '样品编号': record[1],
                '所属公司': record[2],
                '所属水厂': record[3],
                '水样类型': record[4],
                '采样时间': record[5]
            }

            # 获取检测指标值
            cursor.execute('''
                SELECT column_name, value
                FROM raw_data_values
                WHERE record_id = ?
            ''', (record_id,))

            for row in cursor.fetchall():
                data_row[row[0]] = row[1]


            # 创建DataFrame
            df = pd.DataFrame([data_row], columns=columns)

            # 生成文件
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            if export_format == 'csv':
                filename = f'{sample_number}_{timestamp}.csv'
                filepath = os.path.join('exports', filename)
                df.to_csv(filepath, index=False, encoding='utf-8-sig')
            else:
                filename = f'{sample_number}_{timestamp}.xlsx'
                filepath = os.path.join('exports', filename)
                df.to_excel(filepath, index=False, engine='openpyxl')

            log_operation('导出单条原始数据', f'样品编号: {sample_number}')

            return send_file(filepath, as_attachment=True, download_name=filename)

    except Exception as e:
        return jsonify({'error': f'导出失败: {str(e)}'}), 500

@raw_data_bp.route('/api/raw-data/companies', methods=['GET'])
@login_required
def api_raw_data_companies():
    """获取所有被检单位"""
    with get_db() as conn:
        try:
            cursor = conn.cursor()

            cursor.execute('''
                SELECT DISTINCT company_name
                FROM raw_data_records
                WHERE company_name IS NOT NULL AND company_name != ''
                ORDER BY company_name
            ''')

            companies = [row[0] for row in cursor.fetchall()]

            return jsonify({'companies': companies})

        except Exception as e:
            return jsonify({'error': f'获取被检单位失败: {str(e)}'}), 500

@raw_data_bp.route('/api/raw-data/plants', methods=['POST'])
@login_required
def api_raw_data_plants():
    """获取指定被检单位的所有水厂"""
    try:
        data = request.json
        company_name = data.get('company_name', '').strip()

        if not company_name:
            return jsonify({'error': '被检单位不能为空'}), 400

        with get_db() as conn:
            cursor = conn.cursor()

            cursor.execute('''
                SELECT DISTINCT plant_name
                FROM raw_data_records
                WHERE company_name = ? AND plant_name IS NOT NULL AND plant_name != ''
                ORDER BY plant_name
            ''', (company_name,))

            plants = [row[0] for row in cursor.fetchall()]

            return jsonify({'plants': plants})

    except Exception as e:
        return jsonify({'error': f'获取水厂失败: {str(e)}'}), 500

@raw_data_bp.route('/api/raw-data/samples', methods=['POST'])
@login_required
def api_raw_data_samples():
    """获取指定条件下的所有样品"""
    try:
        data = request.json
        company_name = (data.get('company_name') or '').strip()
        plant_name = (data.get('plant_name') or '').strip()
        sample_type_id = data.get('sample_type_id')

        with get_db() as conn:
            cursor = conn.cursor()

            # 构建查询
            query = '''
                SELECT DISTINCT rdr.id, rdr.sample_number, rdr.report_number,
                       rdr.sampling_date, st.name as sample_type_name
                FROM raw_data_records rdr
                LEFT JOIN sample_types st ON rdr.sample_type = st.name
                WHERE 1=1
            '''
            params = []

            if company_name:
                query += ' AND rdr.company_name = ?'
                params.append(company_name)

            if plant_name:
                query += ' AND rdr.plant_name = ?'
                params.append(plant_name)

            if sample_type_id:
                # 根据sample_type_id获取样品类型名称
                cursor.execute('SELECT name FROM sample_types WHERE id = ?', (sample_type_id,))
                result = cursor.fetchone()
                if result:
                    query += ' AND rdr.sample_type = ?'
                    params.append(result[0])

            query += ' ORDER BY rdr.sampling_date DESC, rdr.sample_number'

            cursor.execute(query, params)

            samples = []
            for row in cursor.fetchall():
                samples.append({
                    'id': row[0],
                    'sample_number': row[1],
                    'report_number': row[2],
                    'sampling_date': row[3],
                    'sample_type_name': row[4]
                })

            return jsonify({'samples': samples})

    except Exception as e:
        return jsonify({'error': f'获取样品失败: {str(e)}'}), 500

@raw_data_bp.route('/api/raw-data/filter-export', methods=['POST'])
@login_required
def api_raw_data_filter_export():
    """筛选并导出原始数据 - 新版：行为样品编号，列为检测项目"""
    try:
        data = request.json
        template_id = data.get('template_id')
        selected_sample_ids = data.get('selected_sample_ids', [])  # 选中的样品ID列表

        if not template_id:
            return jsonify({'error': '请选择导出模板'}), 400

        if not selected_sample_ids:
            return jsonify({'error': '请至少选择一个样品'}), 400

        with get_db() as conn:
            cursor = conn.cursor()

            # 获取模板配置的检测指标
            cursor.execute('''
                SELECT column_name
                FROM export_template_columns
                WHERE template_id = ?
                ORDER BY column_order
            ''', (template_id,))

            template_indicators = [row[0] for row in cursor.fetchall()]

            if not template_indicators:
                return jsonify({'error': '模板配置错误：未包含任何检测指标'}), 400

            # 查询选中的样品记录
            placeholders = ','.join(['?'] * len(selected_sample_ids))
            query = f'''
                SELECT id, sample_number, report_number, company_name, plant_name,
                       sample_type, sampling_date
                FROM raw_data_records
                WHERE id IN ({placeholders})
                ORDER BY sampling_date, sample_number
            '''

            cursor.execute(query, selected_sample_ids)
            records = cursor.fetchall()

            if not records:
                return jsonify({'error': '未找到选中的样品数据'}), 404

            # 复用 raw_data_field_mapping + for-report 同款匹配逻辑
            # 反向查找：indicator_name → raw列名

            # 1) 获取指标信息（name → id, unit）
            cursor.execute('SELECT id, name, unit FROM indicators')
            ind_rows = cursor.fetchall()
            ind_name_to_id = {row[1]: row[0] for row in ind_rows}
            ind_id_to_name = {row[0]: row[1] for row in ind_rows}
            ind_name_to_unit = {row[1]: row[2] for row in ind_rows}

            # 2) 加载已固化的字段映射（raw_field_name → indicator_id），构建反向索引
            cursor.execute('SELECT raw_field_name, indicator_id FROM raw_data_field_mapping')
            raw_to_ind_id = {row[0]: row[1] for row in cursor.fetchall()}
            # 反向：indicator_id → raw_field_name
            ind_id_to_raw = {}
            for raw_name, ind_id in raw_to_ind_id.items():
                if ind_id not in ind_id_to_raw:
                    ind_id_to_raw[ind_id] = raw_name

            # 3) 获取所有raw列名，构建模糊匹配索引
            cursor.execute('SELECT DISTINCT column_name FROM raw_data_values')
            all_raw_columns = [row[0] for row in cursor.fetchall()]
            # raw列名 → 逐层去括号的基础名列表
            raw_base_index = {}  # base_name → raw_col
            for rc in all_raw_columns:
                name = rc
                while True:
                    stripped = re.sub(r'\([^)]*\)\s*$', '', name).strip()
                    if stripped != name:
                        raw_base_index[stripped] = rc
                        name = stripped
                    else:
                        raw_base_index[name] = rc
                        break

            # 与 for-report 一致的别名映射（raw常见名 → 系统指标名）
            alias_map = {
                '六价铬': '铬(六价)',
                '挥发酚': '挥发酚类(以苯酚计)',
                '总α': '总α放射性',
                '总β': '总β放射性',
            }
            # 反向别名：系统指标名 → raw常见名
            reverse_alias = {v: k for k, v in alias_map.items()}

            # 4) 为每个模板指标找到对应的raw列名
            indicator_to_raw = {}
            indicator_display = {}

            for indicator in template_indicators:
                matched_raw = None
                ind_id = ind_name_to_id.get(indicator)

                # 策略1：通过 raw_data_field_mapping 反查
                if ind_id and ind_id in ind_id_to_raw:
                    matched_raw = ind_id_to_raw[ind_id]

                # 策略2：精确匹配raw列名
                if not matched_raw and indicator in all_raw_columns:
                    matched_raw = indicator

                # 策略3：逐层去括号匹配（与for-report的match_indicator同逻辑）
                if not matched_raw:
                    name = indicator
                    while True:
                        stripped = re.sub(r'\([^)]*\)\s*$', '', name).strip()
                        if stripped in raw_base_index:
                            matched_raw = raw_base_index[stripped]
                            break
                        if stripped == name:
                            break
                        name = stripped

                # 策略4：反向别名匹配
                if not matched_raw:
                    base = re.sub(r'\([^)]*\)', '', indicator).strip()
                    alias_key = reverse_alias.get(indicator) or reverse_alias.get(base)
                    if alias_key and alias_key in raw_base_index:
                        matched_raw = raw_base_index[alias_key]

                if matched_raw:
                    indicator_to_raw[indicator] = matched_raw
                    indicator_display[indicator] = matched_raw
                else:
                    unit = ind_name_to_unit.get(indicator, '')
                    if unit and unit != '/':
                        indicator_display[indicator] = f'{indicator}({unit})'
                    else:
                        indicator_display[indicator] = indicator

            # 准备导出数据：转置格式 - 行为检测项目，列为样品编号
            sample_numbers = []
            sample_values = {}

            for record in records:
                record_id = record[0]
                sample_number = record[1]
                sample_numbers.append(sample_number)

                cursor.execute('''
                    SELECT column_name, value
                    FROM raw_data_values
                    WHERE record_id = ?
                ''', (record_id,))

                raw_values = dict(cursor.fetchall())

                values = {}
                for indicator in template_indicators:
                    raw_col = indicator_to_raw.get(indicator, indicator)
                    values[indicator] = raw_values.get(raw_col, '')
                sample_values[sample_number] = values

            # 构建转置DataFrame：行=检测项目（带单位），列=样品编号
            transposed_data = []
            for indicator in template_indicators:
                row = {'检测项目': indicator_display[indicator]}
                for sn in sample_numbers:
                    row[sn] = sample_values[sn].get(indicator, '')
                transposed_data.append(row)

            columns = ['检测项目'] + sample_numbers
            df = pd.DataFrame(transposed_data, columns=columns)

            # 生成文件
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            filename = f'导出数据_{timestamp}.xlsx'
            filepath = os.path.join('exports', filename)
            df.to_excel(filepath, index=False, engine='openpyxl')

            log_operation('筛选导出原始数据', f'导出{len(records)}条记录，包含{len(template_indicators)}个检测指标')

            return send_file(filepath, as_attachment=True, download_name=filename)

    except Exception as e:
        return jsonify({'error': f'导出失败: {str(e)}'}), 500

@raw_data_bp.route('/api/raw-data/filter-preview', methods=['POST'])
@login_required
def api_raw_data_filter_preview():
    """预览筛选结果（返回符合条件的样品编号列表）"""
    try:
        data = request.json
        filter_field = data.get('filter_field')
        filter_value = data.get('filter_value', '').strip()
        date_start = data.get('date_start')
        date_end = data.get('date_end')

        with get_db() as conn:
            cursor = conn.cursor()

            # 构建查询
            query = 'SELECT sample_number, company_name, plant_name, sample_type, sampling_date FROM raw_data_records WHERE 1=1'
            params = []

            if filter_field == 'company_name' and filter_value:
                query += ' AND company_name LIKE ?'
                params.append(f'%{filter_value}%')
            elif filter_field == 'plant_name' and filter_value:
                query += ' AND plant_name LIKE ?'
                params.append(f'%{filter_value}%')
            elif filter_field == 'sample_type' and filter_value:
                query += ' AND sample_type LIKE ?'
                params.append(f'%{filter_value}%')
            elif filter_field == 'date_range' and date_start and date_end:
                query += ' AND sampling_date BETWEEN ? AND ?'
                params.append(date_start)
                params.append(date_end)

            query += ' ORDER BY sampling_date DESC'

            cursor.execute(query, params)
            results = []

            for row in cursor.fetchall():
                results.append({
                    'sample_number': row[0],
                    'company_name': row[1],
                    'plant_name': row[2],
                    'sample_type': row[3],
                    'sampling_date': row[4]
                })


            return jsonify({
                'total': len(results),
                'results': results
            })

    except Exception as e:
        return jsonify({'error': f'查询失败: {str(e)}'}), 500

@raw_data_bp.route('/api/raw-data/sample-numbers', methods=['GET'])
@login_required
def api_raw_data_sample_numbers():
    """获取原始数据样品编号列表（用于自动完成/下拉选择）"""
    try:
        search = request.args.get('search', '').strip()
        with get_db() as conn:
            cursor = conn.cursor()

            query = '''
                SELECT sample_number, company_name, plant_name, sample_type, sampling_date
                FROM raw_data_records
            '''
            params = []

            if search:
                query += ' WHERE sample_number LIKE ? OR company_name LIKE ? OR plant_name LIKE ?'
                params = [f'%{search}%', f'%{search}%', f'%{search}%']

            query += ' ORDER BY sampling_date DESC LIMIT 50'

            cursor.execute(query, params)
            results = []
            for row in cursor.fetchall():
                results.append({
                    'sample_number': row[0],
                    'company_name': row[1],
                    'plant_name': row[2],
                    'sample_type': row[3],
                    'sampling_date': row[4]
                })

            return jsonify(results)

    except Exception as e:
        return jsonify({'error': f'查询失败: {str(e)}'}), 500


@raw_data_bp.route('/api/raw-data/for-report', methods=['GET'])
@login_required
def api_raw_data_for_report():
    """根据样品编号获取原始数据用于创建报告"""
    try:
        sample_number = request.args.get('sample_number', '').strip()
        if not sample_number:
            return jsonify({'error': '请提供样品编号'}), 400

        with get_db() as conn:
            cursor = conn.cursor()

            # 1. 查询原始数据记录
            cursor.execute('''
                SELECT id, sample_number, company_name, plant_name, sample_type, sampling_date
                FROM raw_data_records
                WHERE sample_number = ?
            ''', (sample_number,))

            record = cursor.fetchone()
            if not record:
                return jsonify({'error': f'未找到样品编号为 {sample_number} 的原始数据'}), 404

            record_id = record[0]
            company_name = record[2]
            plant_name = record[3]
            sample_type = record[4]
            sampling_date = record[5]

            # 2. 查询原始数据检测值
            cursor.execute('''
                SELECT column_name, value
                FROM raw_data_values
                WHERE record_id = ?
                ORDER BY id
            ''', (record_id,))

            raw_values = cursor.fetchall()

            # 3. 匹配客户信息 -> customer_id
            #    优先按 被检单位+水厂 精确匹配 customers 表，回退按被检单位匹配
            customer_id = None
            if company_name and plant_name:
                cursor.execute('SELECT id FROM customers WHERE inspected_unit = ? AND water_plant = ?',
                               (company_name, plant_name))
                cust_row = cursor.fetchone()
                if cust_row:
                    customer_id = cust_row[0]
            if not customer_id and company_name:
                cursor.execute('SELECT id FROM customers WHERE inspected_unit = ? LIMIT 1',
                               (company_name,))
                cust_row = cursor.fetchone()
                if cust_row:
                    customer_id = cust_row[0]
            # 兼容旧 companies 表
            company_id = None
            if not customer_id and company_name:
                cursor.execute('SELECT id FROM companies WHERE name = ?', (company_name,))
                company_row = cursor.fetchone()
                if company_row:
                    company_id = company_row[0]

            # 4. 匹配样品类型 -> sample_type_id
            #    支持 "名称|代码" 格式（优先按代码精确匹配）和纯名称格式
            sample_type_id = None
            if sample_type and '|' in sample_type:
                st_name, st_code = sample_type.split('|', 1)
                cursor.execute('SELECT id FROM sample_types WHERE name = ? AND code = ?', (st_name, st_code))
                type_row = cursor.fetchone()
                if type_row:
                    sample_type_id = type_row[0]
                sample_type = st_name  # 存入数据库时只保留名称
            elif sample_type:
                cursor.execute('SELECT id FROM sample_types WHERE name = ?', (sample_type,))
                type_row = cursor.fetchone()
                if type_row:
                    sample_type_id = type_row[0]

            # 5. 预加载指标，优先使用该样品类型关联的指标
            all_indicators = {}
            if sample_type_id:
                cursor.execute('''
                    SELECT i.id, i.name, i.unit,
                        COALESCE(ti.limit_value, i.limit_value) as limit_value,
                        i.detection_method
                    FROM template_indicators ti
                    JOIN indicators i ON ti.indicator_id = i.id
                    WHERE ti.sample_type_id = ?
                ''', (sample_type_id,))
            else:
                cursor.execute('''
                    SELECT id, name, unit, limit_value, detection_method
                    FROM indicators
                ''')
            for ind_row in cursor.fetchall():
                all_indicators[ind_row[1]] = {
                    'indicator_id': ind_row[0],
                    'indicator_name': ind_row[1],
                    'unit': ind_row[2],
                    'limit_value': ind_row[3],
                    'detection_method': ind_row[4]
                }

            # 构建模糊匹配索引：去除单位括号后的名称 -> 指标信息
            import re as _re
            fuzzy_index = {}
            for ind_name, ind_info in all_indicators.items():
                # 去掉末尾的(单位)部分，如 "硝酸盐(以N计)" -> "硝酸盐"
                base = _re.sub(r'\([^)]*\)\s*$', '', ind_name).strip()
                if base and base not in fuzzy_index:
                    fuzzy_index[base] = ind_info
                # 也保留原名
                fuzzy_index[ind_name] = ind_info

            # 已知别名映射：原始记录常见名称 -> 系统指标名称
            alias_map = {
                '六价铬': '铬(六价)',
                '挥发酚': '挥发酚类(以苯酚计)',
                '总α放射性': '总α放射性',
                '总β放射性': '总β放射性',
                '总α': '总α放射性',
                '总β': '总β放射性',
            }

            def match_indicator(col_name):
                """尝试将原始数据字段名匹配到系统指标"""
                # 1. 精确匹配
                if col_name in all_indicators:
                    return all_indicators[col_name]
                # 2. 去掉原始字段名中的单位括号后匹配
                base = _re.sub(r'\([^)]*\)\s*$', '', col_name).strip()
                # 可能有多层括号，如 "硝酸盐(以N计)(mg/L)" -> "硝酸盐(以N计)" -> "硝酸盐"
                while base != col_name:
                    if base in all_indicators:
                        return all_indicators[base]
                    if base in fuzzy_index:
                        return fuzzy_index[base]
                    col_name = base
                    base = _re.sub(r'\([^)]*\)\s*$', '', base).strip()
                # 3. 别名映射
                if base in alias_map and alias_map[base] in all_indicators:
                    return all_indicators[alias_map[base]]
                # 4. Unicode下标归一化后匹配 (₃->3, ₂->2)
                normalized = base.replace('₃', '3').replace('₂', '2').replace('₁', '1')
                if normalized in all_indicators:
                    return all_indicators[normalized]
                if normalized in fuzzy_index:
                    return fuzzy_index[normalized]
                return None

            # 6. 加载已固化的字段映射
            cursor.execute('SELECT raw_field_name, indicator_id, indicator_name FROM raw_data_field_mapping')
            saved_mappings = {row[0]: row[1] for row in cursor.fetchall()}

            # 7. 为每个检测值尝试匹配指标
            detection_items = []
            unmatched_items = []
            new_mappings = []  # 新发现的映射，待保存

            for rv in raw_values:
                col_name = rv[0]
                value = rv[1]

                ind_info = None
                # 优先查已固化映射
                if col_name in saved_mappings:
                    ind_id = saved_mappings[col_name]
                    for info in all_indicators.values():
                        if info['indicator_id'] == ind_id:
                            ind_info = info
                            break

                # 未命中则走模糊匹配
                if not ind_info:
                    ind_info = match_indicator(col_name)
                    # 模糊匹配成功且非精确匹配，保存映射
                    if ind_info and col_name not in all_indicators:
                        new_mappings.append((col_name, ind_info['indicator_id'], ind_info['indicator_name']))

                if ind_info:
                    detection_items.append({
                        'indicator_name': ind_info['indicator_name'],
                        'indicator_id': ind_info['indicator_id'],
                        'measured_value': value or '',
                        'unit': ind_info['unit'] or '',
                        'limit_value': ind_info['limit_value'] or '',
                        'detection_method': ind_info['detection_method'] or ''
                    })
                else:
                    unmatched_items.append({
                        'indicator_name': col_name,
                        'indicator_id': None,
                        'measured_value': value or '',
                        'unit': '',
                        'limit_value': '',
                        'detection_method': ''
                    })

            # 8. 持久化新发现的映射
            for raw_name, ind_id, ind_name in new_mappings:
                try:
                    cursor.execute(
                        'INSERT OR IGNORE INTO raw_data_field_mapping (raw_field_name, indicator_id, indicator_name) VALUES (?, ?, ?)',
                        (raw_name, ind_id, ind_name)
                    )
                except Exception:
                    pass

            # 收集结果到局部变量，确保在 with 块内完成所有数据库操作
            result_data = {
                'sample_number': sample_number,
                'company_name': company_name,
                'company_id': company_id,
                'customer_id': customer_id,
                'plant_name': plant_name,
                'sample_type': sample_type,
                'sample_type_id': sample_type_id,
                'sampling_date': sampling_date,
                'detection_items': detection_items,
                'unmatched_items': unmatched_items
            }

        # 在 with 块外返回响应，连接已安全关闭
        return jsonify(result_data)

    except Exception as e:
        return jsonify({'error': f'查询失败: {str(e)}'}), 500


# ── 数据校核 API ─────────────────────────────────────────────────────────

@raw_data_bp.route('/api/raw-data/validate', methods=['POST'])
@login_required
def api_raw_data_validate():
    """对已导入的原始数据执行校核"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': '缺少请求参数'}), 400

        sample_numbers = data.get('sample_numbers', [])
        detection_date = data.get('detection_date', None)

        if not sample_numbers:
            return jsonify({'error': '请指定样品编号'}), 400

        results = validate_from_database(sample_numbers, detection_date)

        # 统计
        counts = {'error': 0, 'warning': 0, 'notice': 0}
        for r in results:
            counts[r['level']] += 1

        return jsonify({
            'success': True,
            'total': len(results),
            'counts': counts,
            'results': results,
        })

    except Exception as e:
        return jsonify({'error': f'校核失败: {str(e)}'}), 500


@raw_data_bp.route('/api/raw-data/validate-by-filters', methods=['POST'])
@login_required
def api_raw_data_validate_by_filters():
    """按筛选条件查找样品并执行校核"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': '缺少请求参数'}), 400

        company_name = data.get('company_name', '').strip()
        plant_name = data.get('plant_name', '').strip()
        sample_type = data.get('sample_type', '').strip()
        date_from = data.get('date_from', '').strip()
        date_to = data.get('date_to', '').strip()
        detection_date = data.get('detection_date', None)

        # 构建查询
        conditions = []
        query_params = []
        if company_name:
            conditions.append("company_name LIKE ?")
            query_params.append(f"%{company_name}%")
        if plant_name:
            conditions.append("plant_name LIKE ?")
            query_params.append(f"%{plant_name}%")
        if sample_type:
            conditions.append("sample_type = ?")
            query_params.append(sample_type)
        if date_from:
            conditions.append("sampling_date >= ?")
            query_params.append(date_from)
        if date_to:
            conditions.append("sampling_date <= ?")
            query_params.append(date_to)

        where = " AND ".join(conditions) if conditions else "1=1"
        query = f"SELECT sample_number FROM raw_data_records WHERE {where} ORDER BY sampling_date DESC LIMIT 200"

        with get_db() as conn:
            rows = conn.execute(query, query_params).fetchall()

        sample_numbers = [r['sample_number'] for r in rows]

        if not sample_numbers:
            return jsonify({
                'success': True,
                'total': 0,
                'counts': {'error': 0, 'warning': 0, 'notice': 0},
                'results': [],
                'sample_count': 0,
            })

        results = validate_from_database(sample_numbers, detection_date)

        counts = {'error': 0, 'warning': 0, 'notice': 0}
        for r in results:
            counts[r['level']] += 1

        return jsonify({
            'success': True,
            'total': len(results),
            'counts': counts,
            'results': results,
            'sample_count': len(sample_numbers),
        })

    except Exception as e:
        return jsonify({'error': f'校核失败: {str(e)}'}), 500
