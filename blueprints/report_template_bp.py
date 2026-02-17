"""
报告模版管理 API Blueprint
从 app_v2.py 提取的报告模版管理路由
"""
from flask import Blueprint, request, jsonify, send_file, session
from models_v2 import get_db
from auth import login_required, admin_required, log_operation
import json
import os
import re
import openpyxl
from datetime import datetime

report_template_bp = Blueprint('report_template_bp', __name__)

# ==================== 报告模版管理 API ====================
@report_template_bp.route('/api/report-templates', methods=['GET'])
@login_required
def api_report_templates():
    """获取报告模版列表"""
    with get_db() as conn:

        templates = conn.execute(
            'SELECT t.*, st.name as sample_type_name '
            'FROM excel_report_templates t '
            'LEFT JOIN sample_types st ON t.sample_type_id = st.id '
            'WHERE t.is_active = 1 '
            'ORDER BY t.created_at DESC'
        ).fetchall()


        return jsonify([dict(t) for t in templates])

@report_template_bp.route('/api/report-templates/import', methods=['POST'])
@admin_required
def api_import_report_template():
    """导入报告模版"""
    if 'file' not in request.files:
        return jsonify({'error': '未上传文件'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '文件名为空'}), 400

    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': '只支持Excel文件'}), 400

    # 获取表单数据
    template_name = request.form.get('name')
    sample_type_id = request.form.get('sample_type_id')
    description = request.form.get('description', '')

    if not template_name:
        return jsonify({'error': '模版名称不能为空'}), 400

    try:
        # 保存上传的文件
        os.makedirs('templates/excel_reports', exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{template_name}_{timestamp}.xlsx"
        file_path = os.path.join('templates/excel_reports', filename)
        file.save(file_path)

        # 读取Excel文件
        wb = openpyxl.load_workbook(file_path)

        # 保存到数据库
        with get_db() as conn:
            cursor = conn.cursor()

            # 检查模板名称是否已存在，如果存在则自动添加序号
            final_template_name = template_name
            existing = conn.execute(
                'SELECT COUNT(*) as count FROM excel_report_templates WHERE name = ?',
                (template_name,)
            ).fetchone()

            if existing['count'] > 0:
                # 查找所有相似名称的模板
                similar = conn.execute(
                    'SELECT name FROM excel_report_templates WHERE name LIKE ?',
                    (f'{template_name}%',)
                ).fetchall()

                # 找出最大的序号
                max_num = 0
                for row in similar:
                    name = row['name']
                    # 尝试提取末尾的数字
                    match = re.search(r'_(\d+)$', name)
                    if match:
                        num = int(match.group(1))
                        if num > max_num:
                            max_num = num

                # 使用下一个序号
                final_template_name = f"{template_name}_{max_num + 1}"

            cursor.execute(
                'INSERT INTO excel_report_templates (name, sample_type_id, description, template_file_path) '
                'VALUES (?, ?, ?, ?)',
                (final_template_name, sample_type_id if sample_type_id else None, description, file_path)
            )

            template_id = cursor.lastrowid

            # 分析工作表结构
            for index, sheet_name in enumerate(wb.sheetnames):
                sheet_type = identify_sheet_type(sheet_name)
                page_number = extract_page_number(sheet_name)

                cursor.execute(
                    'INSERT INTO template_sheet_configs '
                    '(template_id, sheet_name, sheet_index, sheet_type, page_number) '
                    'VALUES (?, ?, ?, ?, ?)',
                    (template_id, sheet_name, index, sheet_type, page_number)
                )

            # 解析模板字段（带有[]、()、;标记的单元格）
            from template_field_parser import TemplateFieldParser

            field_count = 0
            try:
                fields = TemplateFieldParser.extract_template_fields(file_path)

                for field in fields:
                    # 插入字段映射
                    cursor.execute(
                        '''INSERT INTO template_field_mappings
                           (template_id, field_name, field_display_name, field_type,
                            sheet_name, cell_address, placeholder, default_value, is_required,
                            original_cell_text, field_code, is_reference, column_mapping)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                        (template_id,
                         field['field_name'],
                         field['display_name'],
                         field.get('field_type', 'text'),  # 使用解析得到的字段类型
                         field['sheet_name'],
                         field['cell_address'],
                         field.get('placeholder', ''),
                         field.get('default_value', ''),
                         1 if field.get('is_required', True) else 0,
                         field.get('original_value', ''),  # 保存原始单元格文本
                         field.get('field_code'),  # 保存字段代号（如 #report_no）
                         1 if field.get('is_reference', False) else 0,  # 是否为引用字段
                         field.get('column_mapping', ''))  # 检测数据列映射
                    )
                    field_count += 1
            except Exception as e:
                print(f"字段解析警告: {e}")
                # 字段解析失败不影响模板导入


            log_operation('导入报告模版', f'导入模版: {template_name}, 解析字段: {field_count}个', conn=conn)

            return jsonify({
                'id': template_id,
                'message': '模版导入成功',
                'sheet_count': len(wb.sheetnames),
                'field_count': field_count
            }), 201

    except Exception as e:
        return jsonify({'error': f'导入失败: {str(e)}'}), 500

@report_template_bp.route('/api/report-templates/<int:id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def api_report_template_detail(id):
    """获取、修改或删除报告模版"""
    with get_db() as conn:

        if request.method == 'PUT':
            # 仅管理员可修改
            if session.get('role') not in ('admin', 'super_admin'):
                return jsonify({'error': '需要管理员权限'}), 403

            data = request.json
            name = data.get('name')
            sample_type_id = data.get('sample_type_id')
            description = data.get('description', '')

            if not name:
                return jsonify({'error': '模版名称不能为空'}), 400

            # 检查是否存在同名的其他模版（包括已删除的模版，因为UNIQUE约束对所有行生效）
            existing = conn.execute(
                'SELECT id, is_active, name FROM excel_report_templates WHERE name = ? AND id != ?',
                (name, id)
            ).fetchone()

            if existing:
                if existing['is_active']:
                    return jsonify({'error': '模版名称已存在，请使用其他名称'}), 400
                else:
                    # 如果是已删除的模版，自动重命名它以释放该名称
                    import datetime
                    timestamp = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
                    new_name_for_deleted = f"{existing['name']}_已删除_{timestamp}"
                    conn.execute(
                        'UPDATE excel_report_templates SET name = ? WHERE id = ?',
                        (new_name_for_deleted, existing['id'])
                    )

            try:
                conn.execute(
                    'UPDATE excel_report_templates SET name = ?, sample_type_id = ?, description = ? WHERE id = ?',
                    (name, sample_type_id, description, id)
                )

                log_operation('修改报告模版', f'修改模版: {name}', conn=conn)

                return jsonify({'message': '模版更新成功'})
            except Exception as e:
                return jsonify({'error': f'更新失败: {str(e)}'}), 500

        if request.method == 'DELETE':
            # 仅管理员可删除
            if session.get('role') not in ('admin', 'super_admin'):
                return jsonify({'error': '需要管理员权限'}), 403

            # 获取模版信息
            template = conn.execute(
                'SELECT * FROM excel_report_templates WHERE id = ?',
                (id,)
            ).fetchone()

            if not template:
                return jsonify({'error': '模版不存在'}), 404

            # 删除文件
            if template['template_file_path'] and os.path.exists(template['template_file_path']):
                try:
                    os.remove(template['template_file_path'])
                except Exception as e:
                    print(f"删除模版文件失败: {e}")

            # 软删除（设置is_active=0）
            conn.execute('UPDATE excel_report_templates SET is_active = 0 WHERE id = ?', (id,))

            log_operation('删除报告模版', f'删除模版: {template["name"]}', conn=conn)

            return jsonify({'message': '模版删除成功'})

        # GET请求 - 获取模版详情
        template = conn.execute(
            'SELECT * FROM excel_report_templates WHERE id = ?',
            (id,)
        ).fetchone()

        if not template:
            return jsonify({'error': '模版不存在'}), 404

        # 获取工作表配置
        sheets = conn.execute(
            'SELECT * FROM template_sheet_configs WHERE template_id = ? ORDER BY sheet_index',
            (id,)
        ).fetchall()

        # 获取字段映射
        fields = conn.execute(
            'SELECT * FROM template_field_mappings WHERE template_id = ?',
            (id,)
        ).fetchall()


        return jsonify({
            'template': dict(template),
            'sheets': [dict(s) for s in sheets],
            'fields': [dict(f) for f in fields]
        })

@report_template_bp.route('/api/report-templates/<int:id>/fields', methods=['GET', 'POST'])
@admin_required
def api_template_fields(id):
    """获取或添加模版字段映射"""
    with get_db() as conn:

        if request.method == 'POST':
            data = request.json
            field_name = data.get('field_name')
            field_type = data.get('field_type')
            sheet_name = data.get('sheet_name')
            cell_address = data.get('cell_address')
            start_row = data.get('start_row')
            start_col = data.get('start_col')
            description = data.get('description', '')
            is_required = data.get('is_required', False)
            default_value = data.get('default_value', '')

            if not all([field_name, field_type, sheet_name]):
                return jsonify({'error': '缺少必填字段'}), 400

            cursor = conn.cursor()
            cursor.execute(
                'INSERT INTO template_field_mappings '
                '(template_id, field_name, field_type, sheet_name, cell_address, '
                'start_row, start_col, description, is_required, default_value) '
                'VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
                (id, field_name, field_type, sheet_name, cell_address,
                 start_row, start_col, description, is_required, default_value)
            )

            field_id = cursor.lastrowid

            log_operation('添加模版字段映射', f'模版ID: {id}, 字段: {field_name}', conn=conn)

            return jsonify({'id': field_id, 'message': '字段映射添加成功'}), 201

        # GET请求
        fields = conn.execute(
            'SELECT * FROM template_field_mappings WHERE template_id = ? ORDER BY id',
            (id,)
        ).fetchall()


        return jsonify([dict(f) for f in fields])

@report_template_bp.route('/api/template-fields/<int:field_id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def api_template_field_detail(field_id):
    """获取、更新或删除单个模板字段"""
    with get_db() as conn:

        if request.method == 'GET':
            # 获取字段详情
            field = conn.execute(
                'SELECT * FROM template_field_mappings WHERE id = ?',
                (field_id,)
            ).fetchone()


            if not field:
                return jsonify({'error': '字段不存在'}), 404

            return jsonify(dict(field))

        if request.method == 'PUT':
            # 仅管理员可修改
            if session.get('role') not in ('admin', 'super_admin'):
                return jsonify({'error': '需要管理员权限'}), 403

            # 检查字段是否存在
            field = conn.execute(
                'SELECT * FROM template_field_mappings WHERE id = ?',
                (field_id,)
            ).fetchone()

            if not field:
                return jsonify({'error': '字段不存在'}), 404

            data = request.json
            field_name = data.get('field_name')
            field_display_name = data.get('field_display_name', '')
            field_type = data.get('field_type')
            sheet_name = data.get('sheet_name')
            cell_address = data.get('cell_address', '')
            placeholder = data.get('placeholder', '')
            default_value = data.get('default_value', '')
            is_required = data.get('is_required', 0)
            description = data.get('description', '')

            if not all([field_name, field_type, sheet_name]):
                return jsonify({'error': '缺少必填字段'}), 400

            try:
                conn.execute(
                    '''UPDATE template_field_mappings
                       SET field_name = ?, field_display_name = ?, field_type = ?,
                           sheet_name = ?, cell_address = ?, placeholder = ?,
                           default_value = ?, is_required = ?, description = ?
                       WHERE id = ?''',
                    (field_name, field_display_name, field_type, sheet_name,
                     cell_address, placeholder, default_value, is_required,
                     description, field_id)
                )

                log_operation('更新模板字段', f'字段ID: {field_id}, 字段名: {field_name}', conn=conn)

                return jsonify({'message': '字段更新成功'})
            except Exception as e:
                return jsonify({'error': f'更新失败: {str(e)}'}), 500

        if request.method == 'DELETE':
            # 仅管理员可删除
            if session.get('role') not in ('admin', 'super_admin'):
                return jsonify({'error': '需要管理员权限'}), 403

            # 检查字段是否存在
            field = conn.execute(
                'SELECT * FROM template_field_mappings WHERE id = ?',
                (field_id,)
            ).fetchone()

            if not field:
                return jsonify({'error': '字段不存在'}), 404

            try:
                conn.execute('DELETE FROM template_field_mappings WHERE id = ?', (field_id,))

                log_operation('删除模板字段', f'字段ID: {field_id}, 字段名: {field["field_name"]}', conn=conn)

                return jsonify({'message': '字段删除成功'})
            except Exception as e:
                return jsonify({'error': f'删除失败: {str(e)}'}), 500

@report_template_bp.route('/api/report-templates/<int:id>/export-config', methods=['GET'])
@login_required
def api_export_template_config(id):
    """导出模板配置为Excel文件"""
    from template_config_excel import TemplateConfigExcel

    try:
        output_path = TemplateConfigExcel.export_template_config(id)

        log_operation('导出模板配置', f'导出模板ID: {id}的配置')

        return send_file(
            output_path,
            as_attachment=True,
            download_name=os.path.basename(output_path),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except ValueError as e:
        return jsonify({'error': str(e)}), 404
    except Exception as e:
        return jsonify({'error': f'导出失败: {str(e)}'}), 500

@report_template_bp.route('/api/report-templates/<int:id>/import-config', methods=['POST'])
@admin_required
def api_import_template_config(id):
    """从Excel文件导入模板配置"""
    if 'file' not in request.files:
        return jsonify({'error': '未上传文件'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '文件名为空'}), 400

    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': '只支持Excel文件'}), 400

    from template_config_excel import TemplateConfigExcel

    try:
        # 保存上传的文件到临时位置
        os.makedirs('temp/config_imports', exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        temp_path = f'temp/config_imports/config_import_{timestamp}.xlsx'
        file.save(temp_path)

        # 导入配置
        result = TemplateConfigExcel.import_template_config(id, temp_path)

        # 删除临时文件
        try:
            os.remove(temp_path)
        except OSError:
            pass

        log_operation('导入模板配置', f'导入模板ID: {id}的配置，共{result["inserted_count"]}个字段')

        return jsonify(result), 200

    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': f'导入失败: {str(e)}'}), 500


def identify_sheet_type(sheet_name):
    """识别工作表类型"""
    sheet_name_lower = sheet_name.lower()

    if '1' in sheet_name or 'cover' in sheet_name_lower or '封面' in sheet_name:
        return 'cover'
    elif '2' in sheet_name or 'info' in sheet_name_lower or '信息' in sheet_name:
        return 'info'
    elif any(x in sheet_name for x in ['3', '4']) or 'data' in sheet_name_lower or '数据' in sheet_name:
        return 'data'
    elif '5' in sheet_name or 'note' in sheet_name_lower or '说明' in sheet_name:
        return 'conclusion'
    else:
        return 'other'

def extract_page_number(sheet_name):
    """从工作表名称提取页码"""
    match = re.search(r'\d+', sheet_name)
    return int(match.group()) if match else 0
