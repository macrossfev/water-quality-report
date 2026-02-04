#!/usr/bin/env python3
"""检查第一页C8和C11单元格"""
import openpyxl

# 检查模板文件
template_path = 'templates/excel_reports/出厂水_20260203_002830.xlsx'
wb = openpyxl.load_workbook(template_path)
ws = wb[wb.sheetnames[0]]  # 第一页

print("模板文件第一页:")
print(f"  C8单元格: {repr(ws['C8'].value)}")
print(f"  C9单元格: {repr(ws['C9'].value)}")
print(f"  C10单元格: {repr(ws['C10'].value)}")
print(f"  C11单元格: {repr(ws['C11'].value)}")

wb.close()

# 检查生成的报告
report_path = 'exports/report_20260203编4_20260203_020743.xlsx'
wb = openpyxl.load_workbook(report_path)
ws = wb[wb.sheetnames[0]]

print("\n生成的报告第一页:")
print(f"  C8单元格: {repr(ws['C8'].value)}")
print(f"  C9单元格: {repr(ws['C9'].value)}")
print(f"  C10单元格: {repr(ws['C10'].value)}")
print(f"  C11单元格: {repr(ws['C11'].value)}")

wb.close()

# 检查数据库中的字段映射
import sqlite3
conn = sqlite3.connect('database/water_quality_v2.db')
conn.row_factory = sqlite3.Row
cursor = conn.cursor()

cursor.execute("""
    SELECT field_name, field_code, cell_address, original_cell_text
    FROM template_field_mappings
    WHERE template_id = 8 AND sheet_name = '1' AND cell_address IN ('C8', 'C9', 'C10', 'C11')
    ORDER BY cell_address
""")

fields = cursor.fetchall()
print("\n数据库中第一页C列的字段映射:")
for field in fields:
    print(f"  {field['cell_address']}: {field['field_name']} [{field['field_code']}]")
    print(f"    原始文本: {repr(field['original_cell_text'])}")

conn.close()
