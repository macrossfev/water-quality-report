"""
生成水质检测报告Excel模板示例
包含完整的字段代号标记和格式
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_example_template():
    """创建Excel示例模板"""
    wb = openpyxl.Workbook()

    # 删除默认工作表
    wb.remove(wb.active)

    # 创建四个工作表
    ws1 = wb.create_sheet("报告信息")
    ws2 = wb.create_sheet("说明")
    ws3 = wb.create_sheet("检测数据表1")
    ws4 = wb.create_sheet("检测数据表2")

    # 定义样式
    title_font = Font(name='微软雅黑', size=16, bold=True)
    header_font = Font(name='微软雅黑', size=11, bold=True)
    normal_font = Font(name='微软雅黑', size=10)
    field_font = Font(name='Consolas', size=10, color='0000FF')  # 蓝色字体用于字段代号

    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    header_fill = PatternFill(start_color='E8F4F8', end_color='E8F4F8', fill_type='solid')
    field_fill = PatternFill(start_color='FFF9E6', end_color='FFF9E6', fill_type='solid')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ========== 第一页：报告信息 ==========
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 30
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['D'].width = 30

    # 标题
    ws1.merge_cells('A1:D1')
    ws1['A1'] = '水质检测报告'
    ws1['A1'].font = title_font
    ws1['A1'].alignment = center_align
    ws1.row_dimensions[1].height = 30

    # 基本信息区
    row = 3
    ws1[f'A{row}'] = '报告编号：'
    ws1[f'B{row}'] = '[#report_no]'
    ws1[f'B{row}'].font = field_font
    ws1[f'B{row}'].fill = field_fill
    ws1[f'C{row}'] = '样品编号：'
    ws1[f'D{row}'] = '[#sample_no]'
    ws1[f'D{row}'].font = field_font
    ws1[f'D{row}'].fill = field_fill

    row += 1
    ws1[f'A{row}'] = '样品类型：'
    ws1[f'B{row}'] = '[#sample_type]'
    ws1[f'B{row}'].font = field_font
    ws1[f'B{row}'].fill = field_fill

    # 委托单位信息
    row += 2
    ws1.merge_cells(f'A{row}:D{row}')
    ws1[f'A{row}'] = '=== 委托单位信息 ==='
    ws1[f'A{row}'].font = header_font
    ws1[f'A{row}'].fill = header_fill
    ws1[f'A{row}'].alignment = center_align

    row += 1
    ws1[f'A{row}'] = '委托单位：'
    ws1[f'B{row}'] = '[#company]'
    ws1[f'B{row}'].font = field_font
    ws1[f'B{row}'].fill = field_fill
    ws1[f'C{row}'] = '被检单位：'
    ws1[f'D{row}'] = '[#customer_unit]'
    ws1[f'D{row}'].font = field_font
    ws1[f'D{row}'].fill = field_fill

    row += 1
    ws1[f'A{row}'] = '被检水厂：'
    ws1[f'B{row}'] = '[#customer_plant]'
    ws1[f'B{row}'].font = field_font
    ws1[f'B{row}'].fill = field_fill
    ws1[f'C{row}'] = '单位地址：'
    ws1[f'D{row}'] = '[#unit_address]'
    ws1[f'D{row}'].font = field_font
    ws1[f'D{row}'].fill = field_fill

    # 采样信息
    row += 2
    ws1.merge_cells(f'A{row}:D{row}')
    ws1[f'A{row}'] = '=== 采样信息 ==='
    ws1[f'A{row}'].font = header_font
    ws1[f'A{row}'].fill = header_fill
    ws1[f'A{row}'].alignment = center_align

    row += 1
    ws1[f'A{row}'] = '采样日期：'
    ws1[f'B{row}'] = '[#sampling_date]'
    ws1[f'B{row}'].font = field_font
    ws1[f'B{row}'].fill = field_fill
    ws1[f'C{row}'] = '采样人：'
    ws1[f'D{row}'] = '[#sampler]'
    ws1[f'D{row}'].font = field_font
    ws1[f'D{row}'].fill = field_fill

    row += 1
    ws1[f'A{row}'] = '采样地点：'
    ws1[f'B{row}'] = '[#sampling_location]'
    ws1[f'B{row}'].font = field_font
    ws1[f'B{row}'].fill = field_fill
    ws1[f'C{row}'] = '采样依据：'
    ws1[f'D{row}'] = '[#sampling_basis]'
    ws1[f'D{row}'].font = field_font
    ws1[f'D{row}'].fill = field_fill

    row += 1
    ws1[f'A{row}'] = '样品来源：'
    ws1[f'B{row}'] = '[#sample_source]'
    ws1[f'B{row}'].font = field_font
    ws1[f'B{row}'].fill = field_fill
    ws1[f'C{row}'] = '样品状态：'
    ws1[f'D{row}'] = '[#sample_status]'
    ws1[f'D{row}'].font = field_font
    ws1[f'D{row}'].fill = field_fill

    row += 1
    ws1[f'A{row}'] = '收样日期：'
    ws1[f'B{row}'] = '[#sample_received]'
    ws1[f'B{row}'].font = field_font
    ws1[f'B{row}'].fill = field_fill

    # 检测信息
    row += 2
    ws1.merge_cells(f'A{row}:D{row}')
    ws1[f'A{row}'] = '=== 检测信息 ==='
    ws1[f'A{row}'].font = header_font
    ws1[f'A{row}'].fill = header_fill
    ws1[f'A{row}'].alignment = center_align

    row += 1
    ws1[f'A{row}'] = '检测日期：'
    ws1[f'B{row}'] = '[#detection_date]'
    ws1[f'B{row}'].font = field_font
    ws1[f'B{row}'].fill = field_fill
    ws1[f'C{row}'] = '检测人员：'
    ws1[f'D{row}'] = '[#detection_person]'
    ws1[f'D{row}'].font = field_font
    ws1[f'D{row}'].fill = field_fill

    row += 1
    ws1[f'A{row}'] = '审核人员：'
    ws1[f'B{row}'] = '[#review_person]'
    ws1[f'B{row}'].font = field_font
    ws1[f'B{row}'].fill = field_fill
    ws1[f'C{row}'] = '报告日期：'
    ws1[f'D{row}'] = '[#report_date]'
    ws1[f'D{row}'].font = field_font
    ws1[f'D{row}'].fill = field_fill

    # 其他信息
    row += 2
    ws1.merge_cells(f'A{row}:D{row}')
    ws1[f'A{row}'] = '=== 其他信息 ==='
    ws1[f'A{row}'].font = header_font
    ws1[f'A{row}'].fill = header_fill
    ws1[f'A{row}'].alignment = center_align

    row += 1
    ws1[f'A{row}'] = '产品标准：'
    ws1.merge_cells(f'B{row}:D{row}')
    ws1[f'B{row}'] = '[#product_standard]'
    ws1[f'B{row}'].font = field_font
    ws1[f'B{row}'].fill = field_fill

    row += 1
    ws1[f'A{row}'] = '检测项目：'
    ws1.merge_cells(f'B{row}:D{row}')
    ws1[f'B{row}'] = '[#detection_items]'
    ws1[f'B{row}'].font = field_font
    ws1[f'B{row}'].fill = field_fill

    row += 1
    ws1[f'A{row}'] = '检测结论：'
    ws1.merge_cells(f'B{row}:D{row}')
    ws1[f'B{row}'] = '[#test_conclusion]'
    ws1[f'B{row}'].font = field_font
    ws1[f'B{row}'].fill = field_fill

    row += 1
    ws1[f'A{row}'] = '附加信息：'
    ws1.merge_cells(f'B{row}:D{row}')
    ws1[f'B{row}'] = '[#additional_info]'
    ws1[f'B{row}'].font = field_font
    ws1[f'B{row}'].fill = field_fill

    row += 1
    ws1[f'A{row}'] = '附件信息：'
    ws1.merge_cells(f'B{row}:D{row}')
    ws1[f'B{row}'] = '[#attachment_info]'
    ws1[f'B{row}'].font = field_font
    ws1[f'B{row}'].fill = field_fill

    # ========== 第二页：说明 ==========
    ws2.column_dimensions['A'].width = 80

    ws2['A1'] = '模板使用说明'
    ws2['A1'].font = title_font
    ws2.row_dimensions[1].height = 25

    instructions = [
        '',
        '本模板使用字段代号系统，蓝色背景的代号会被系统自动替换为实际数据。',
        '',
        '一、字段代号格式',
        '  [#代号] - 标准字段，如 [#report_no]、[#sample_no]',
        '  [*字段名] - 引用字段，从历史报告引用，如 [*被检单位]',
        '',
        '二、检测数据表',
        '  - 第三页和第四页是检测数据表',
        '  - 使用 [#dt_xxx] 代号标记列位置',
        '  - 使用 [#dt_end] 标记数据区结束',
        '  - 如果数据超过第三页容量，会自动填充到第四页',
        '',
        '三、数据容量',
        '  - 第三页：第8行到第29行（共22行数据）',
        '  - 第四页：第8行到第29行（共22行数据）',
        '  - 总容量：44个检测项目',
        '',
        '四、注意事项',
        '  1. 字段代号必须完全匹配（包括方括号和#号）',
        '  2. 不要删除或修改字段代号',
        '  3. 可以调整格式、边框、颜色等',
        '  4. [#dt_end] 标记必须设置，否则默认容量为1000行',
        '',
        '五、下载字段代号完整说明',
        '  在报告模板管理页面点击"下载字段代号说明"按钮',
    ]

    for i, text in enumerate(instructions, start=2):
        ws2[f'A{i}'] = text
        ws2[f'A{i}'].font = normal_font
        ws2[f'A{i}'].alignment = left_align

    # ========== 第三页：检测数据表1 ==========
    create_data_sheet(ws3, "检测数据表（第1页）", thin_border, header_font, normal_font, field_font,
                     center_align, header_fill, field_fill)

    # ========== 第四页：检测数据表2 ==========
    create_data_sheet(ws4, "检测数据表（第2页 - 续）", thin_border, header_font, normal_font, field_font,
                     center_align, header_fill, field_fill)

    # 保存文件
    output_path = 'template_examples/水质检测报告模板示例.xlsx'
    wb.save(output_path)
    print(f"示例模板已创建: {output_path}")
    return output_path

def create_data_sheet(ws, title, border, header_font, normal_font, field_font, center_align, header_fill, field_fill):
    """创建检测数据表格"""
    # 设置列宽
    ws.column_dimensions['A'].width = 8   # 序号
    ws.column_dimensions['B'].width = 20  # 检测项目
    ws.column_dimensions['C'].width = 12  # 单位
    ws.column_dimensions['D'].width = 15  # 检测结果
    ws.column_dimensions['E'].width = 15  # 标准限值
    ws.column_dimensions['F'].width = 25  # 检测方法

    # 标题
    ws.merge_cells('A1:F1')
    ws['A1'] = title
    ws['A1'].font = Font(name='微软雅黑', size=14, bold=True)
    ws['A1'].alignment = center_align
    ws.row_dimensions[1].height = 25

    # 报告信息行
    ws.merge_cells('A3:C3')
    ws['A3'] = '样品编号：[#sample_no]'
    ws['A3'].font = normal_font
    ws['A3'].alignment = Alignment(horizontal='left', vertical='center')

    ws.merge_cells('D3:F3')
    ws['D3'] = '检测日期：[#detection_date]'
    ws['D3'].font = normal_font
    ws['D3'].alignment = Alignment(horizontal='left', vertical='center')

    # 表头
    headers = ['序号', '检测项目', '单位', '检测结果', '标准限值', '检测方法']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col)
        cell.value = header
        cell.font = header_font
        cell.alignment = center_align
        cell.fill = header_fill
        cell.border = border

    # 数据行（第8行）- 放置字段代号
    field_codes = ['[#dt_index]', '[#dt_name]', '[#dt_unit]', '[#dt_result]', '[#dt_limit]', '[#dt_method]']
    for col, code in enumerate(field_codes, start=1):
        cell = ws.cell(row=8, column=col)
        cell.value = code
        cell.font = field_font
        cell.alignment = center_align
        cell.fill = field_fill
        cell.border = border

    # 添加示例说明行（第9行）
    ws.row_dimensions[9].height = 30
    ws.merge_cells('A9:F9')
    ws['A9'] = '↑ 第8行的字段代号标记数据起始位置，系统会从这里开始填充检测数据'
    ws['A9'].font = Font(name='微软雅黑', size=9, italic=True, color='666666')
    ws['A9'].alignment = center_align
    ws['A9'].fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')

    # 数据区结束标记（第30行）
    ws['A30'] = '[#dt_end]'
    ws['A30'].font = field_font
    ws['A30'].fill = field_fill

    ws.row_dimensions[31].height = 30
    ws.merge_cells('A31:F31')
    ws['A31'] = '↑ 第30行标记数据区结束，该页最多填充22行数据（第8-29行）'
    ws['A31'].font = Font(name='微软雅黑', size=9, italic=True, color='666666')
    ws['A31'].alignment = center_align
    ws['A31'].fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')

    # 签名行
    ws.merge_cells('A33:C33')
    ws['A33'] = '检测人员：[#detection_person]'
    ws['A33'].font = normal_font

    ws.merge_cells('D33:F33')
    ws['D33'] = '审核人员：[#review_person]'
    ws['D33'].font = normal_font

if __name__ == '__main__':
    import os
    os.makedirs('template_examples', exist_ok=True)
    create_example_template()
