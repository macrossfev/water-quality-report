"""
Excel批量导入处理器
解析导入的Excel文件并创建报告
"""
import openpyxl
from models_v2 import get_db_connection
from datetime import datetime

class ImportProcessor:
    """导入处理器"""

    def __init__(self, file_path, template_id=None, created_by='system'):
        """
        初始化处理器

        Args:
            file_path: Excel文件路径
            template_id: 报告模板ID（可选）
            created_by: 创建人
        """
        self.file_path = file_path
        self.template_id = template_id
        self.created_by = created_by
        self.workbook = None
        self.results = {
            'success': [],
            'errors': [],
            'warnings': []
        }

    def process(self):
        """
        处理导入

        Returns:
            dict: 处理结果
        """
        try:
            # 打开工作簿
            self.workbook = openpyxl.load_workbook(self.file_path)

            # 1. 解析基本信息
            basic_info_list = self._parse_basic_info()

            # 2. 解析检测数据
            detection_data_dict = self._parse_detection_data()

            # 3. 解析模板字段（如果有）
            template_fields_dict = self._parse_template_fields()

            # 4. 创建报告
            for basic_info in basic_info_list:
                sample_number = basic_info['sample_number']

                try:
                    # 获取该样品的检测数据
                    detection_data = detection_data_dict.get(sample_number, [])

                    # 获取该样品的模板字段
                    template_fields = template_fields_dict.get(sample_number, {})

                    # 创建报告
                    report_id = self._create_report(basic_info, detection_data, template_fields)

                    self.results['success'].append({
                        'sample_number': sample_number,
                        'report_id': report_id,
                        'message': '导入成功'
                    })

                except Exception as e:
                    self.results['errors'].append({
                        'sample_number': sample_number,
                        'message': f'创建报告失败: {str(e)}'
                    })

            return self.results

        except Exception as e:
            self.results['errors'].append({
                'sample_number': 'N/A',
                'message': f'导入失败: {str(e)}'
            })
            return self.results

        finally:
            if self.workbook:
                self.workbook.close()

    def _parse_basic_info(self):
        """解析基本信息sheet"""
        if '基本信息' not in self.workbook.sheetnames:
            raise ValueError("缺少\"基本信息\"工作表")

        ws = self.workbook['基本信息']
        basic_info_list = []

        # 读取标题行
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value).strip() if cell.value else '')

        # 字段映射
        field_map = {
            '样品编号*': 'sample_number',
            '样品编号': 'sample_number',
            '样品类型*': 'sample_type_name',
            '样品类型': 'sample_type_name',
            '委托单位': 'company_name',
            '检测日期': 'detection_date',
            '检测人员': 'detection_person',
            '审核人员': 'review_person',
            '备注': 'remark'
        }

        # 读取数据行
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:  # 跳过空行
                continue

            basic_info = {}
            for idx, value in enumerate(row):
                if idx < len(headers):
                    header = headers[idx]
                    if header in field_map:
                        field_name = field_map[header]
                        basic_info[field_name] = str(value).strip() if value else ''

            # 验证必填字段
            if not basic_info.get('sample_number'):
                self.results['warnings'].append({
                    'sheet': '基本信息',
                    'message': f'跳过空样品编号的行'
                })
                continue

            if not basic_info.get('sample_type_name'):
                self.results['warnings'].append({
                    'sample_number': basic_info['sample_number'],
                    'message': '缺少样品类型'
                })
                continue

            basic_info_list.append(basic_info)

        return basic_info_list

    def _parse_detection_data(self):
        """解析检测数据sheet（简化格式：A列检测项目，B列单位，C列起样品数据）"""
        if '检测数据' not in self.workbook.sheetnames:
            self.results['warnings'].append({
                'sheet': '检测数据',
                'message': '未找到检测数据工作表'
            })
            return {}

        ws = self.workbook['检测数据']
        detection_data_dict = {}

        # 读取首行的样品编号（从C列/第3列开始，A列是检测项目，B列是单位）
        sample_numbers = []
        sample_col_start = 3  # C列开始
        for col in range(sample_col_start, ws.max_column + 1):
            cell_value = ws.cell(1, col).value
            if cell_value:
                sample_number = str(cell_value).strip()
                # 跳过"单位"列（如果B列标题是"单位"）
                if sample_number.lower() in ['单位', 'unit']:
                    continue
                sample_numbers.append((col, sample_number))
                # 初始化样品的数据列表
                if sample_number not in detection_data_dict:
                    detection_data_dict[sample_number] = []
            else:
                break  # 遇到空列就停止

        if not sample_numbers:
            self.results['warnings'].append({
                'sheet': '检测数据',
                'message': '未找到样品编号（首行C列起应包含样品编号）'
            })
            return {}

        # 读取数据（从第2行开始）
        for row in range(2, ws.max_row + 1):
            # 读取检测项目名称（A列/第1列）
            indicator_name = ws.cell(row, 1).value
            if not indicator_name:
                continue  # 跳过空行

            indicator_name = str(indicator_name).strip()

            # 读取每个样品的检测值
            for col_idx, sample_number in sample_numbers:
                measured_value = ws.cell(row, col_idx).value

                # 只有当检测值不为空时才添加
                if measured_value is not None and str(measured_value).strip():
                    data_item = {
                        'indicator_name': indicator_name,
                        'measured_value': str(measured_value).strip(),
                        'remark': ''
                    }
                    detection_data_dict[sample_number].append(data_item)

        return detection_data_dict

    def _parse_template_fields(self):
        """解析模板字段sheet"""
        if '模板字段' not in self.workbook.sheetnames:
            return {}

        ws = self.workbook['模板字段']
        template_fields_dict = {}

        # 读取标题行
        headers = []
        for cell in ws[1]:
            header = str(cell.value).strip() if cell.value else ''
            # 移除*标记
            header = header.rstrip('*')
            headers.append(header)

        # 读取数据行
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:  # 跳过空行
                continue

            sample_number = str(row[0]).strip() if row[0] else ''
            if not sample_number:
                continue

            fields = {}
            for idx, value in enumerate(row[1:], start=1):  # 跳过样品编号列
                if idx < len(headers):
                    field_name = headers[idx]
                    if field_name:
                        fields[field_name] = str(value).strip() if value else ''

            template_fields_dict[sample_number] = fields

        return template_fields_dict

    def _create_report(self, basic_info, detection_data, template_fields):
        """
        创建报告

        Args:
            basic_info: 基本信息
            detection_data: 检测数据列表
            template_fields: 模板字段字典

        Returns:
            int: 报告ID
        """
        conn = get_db_connection()
        cursor = conn.cursor()

        try:
            # 1. 查找样品类型ID
            sample_type = conn.execute(
                'SELECT id FROM sample_types WHERE name = ?',
                (basic_info['sample_type_name'],)
            ).fetchone()

            if not sample_type:
                raise ValueError(f"样品类型不存在: {basic_info['sample_type_name']}")

            sample_type_id = sample_type['id']

            # 2. 查找或创建委托单位
            company_id = None
            if basic_info.get('company_name'):
                company = conn.execute(
                    'SELECT id FROM companies WHERE name = ?',
                    (basic_info['company_name'],)
                ).fetchone()

                if company:
                    company_id = company['id']
                else:
                    # 自动创建委托单位
                    cursor.execute(
                        'INSERT INTO companies (name) VALUES (?)',
                        (basic_info['company_name'],)
                    )
                    company_id = cursor.lastrowid

            # 3. 生成报告编号
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            report_number = f"R{timestamp}_{basic_info['sample_number']}"

            # 4. 创建报告记录
            cursor.execute(
                '''INSERT INTO reports
                   (report_number, sample_number, company_id, sample_type_id,
                    detection_person, review_person, detection_date, remark,
                    template_id, review_status, created_by, created_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                (report_number, basic_info['sample_number'], company_id, sample_type_id,
                 basic_info.get('detection_person', ''), basic_info.get('review_person', ''),
                 basic_info.get('detection_date', ''), basic_info.get('remark', ''),
                 self.template_id, 'pending', self.created_by, datetime.now())
            )

            report_id = cursor.lastrowid

            # 5. 插入检测数据
            for data_item in detection_data:
                # 查找指标ID
                indicator = conn.execute(
                    'SELECT id FROM indicators WHERE name = ?',
                    (data_item['indicator_name'],)
                ).fetchone()

                if not indicator:
                    self.results['warnings'].append({
                        'sample_number': basic_info['sample_number'],
                        'message': f"指标不存在，已跳过: {data_item['indicator_name']}"
                    })
                    continue

                indicator_id = indicator['id']

                # 插入检测数据
                cursor.execute(
                    '''INSERT INTO report_data
                       (report_id, indicator_id, measured_value, remark)
                       VALUES (?, ?, ?, ?)''',
                    (report_id, indicator_id, data_item['measured_value'],
                     data_item.get('remark', ''))
                )

            # 6. 插入模板字段值（如果有）
            if self.template_id and template_fields:
                # 获取模板字段映射
                field_mappings = conn.execute(
                    'SELECT id, field_name, field_display_name FROM template_field_mappings WHERE template_id = ?',
                    (self.template_id,)
                ).fetchall()

                for mapping in field_mappings:
                    field_name = mapping['field_display_name'] or mapping['field_name']
                    field_value = template_fields.get(field_name, '')

                    if field_value:
                        cursor.execute(
                            '''INSERT INTO report_field_values
                               (report_id, field_mapping_id, field_value)
                               VALUES (?, ?, ?)''',
                            (report_id, mapping['id'], field_value)
                        )

            conn.commit()
            return report_id

        except Exception as e:
            conn.rollback()
            raise e

        finally:
            conn.close()


def import_reports_from_excel(file_path, template_id=None, created_by='system'):
    """
    从Excel导入报告的便捷函数

    Args:
        file_path: Excel文件路径
        template_id: 报告模板ID（可选）
        created_by: 创建人

    Returns:
        dict: 处理结果
    """
    processor = ImportProcessor(file_path, template_id, created_by)
    return processor.process()


if __name__ == '__main__':
    import sys

    if len(sys.argv) < 2:
        print("用法: python3 import_processor.py <excel文件路径> [模板ID]")
        sys.exit(1)

    file_path = sys.argv[1]
    template_id = int(sys.argv[2]) if len(sys.argv) > 2 else None

    print("="*60)
    print("批量导入处理器")
    print("="*60)
    print(f"文件: {file_path}")
    print(f"模板ID: {template_id or '未指定'}")
    print()

    results = import_reports_from_excel(file_path, template_id)

    print("\n导入结果:")
    print(f"  成功: {len(results['success'])} 条")
    print(f"  失败: {len(results['errors'])} 条")
    print(f"  警告: {len(results['warnings'])} 条")

    if results['success']:
        print("\n成功列表:")
        for item in results['success']:
            print(f"  ✓ {item['sample_number']}: 报告ID={item['report_id']}")

    if results['errors']:
        print("\n失败列表:")
        for item in results['errors']:
            print(f"  ✗ {item['sample_number']}: {item['message']}")

    if results['warnings']:
        print("\n警告列表:")
        for item in results['warnings']:
            sample_number = item.get('sample_number', 'N/A')
            print(f"  ⚠ {sample_number}: {item['message']}")
