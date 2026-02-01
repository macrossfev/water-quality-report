"""
Excel导入模板生成器
根据报告模板和样品类型生成统一的导入模板
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from models_v2 import get_db_connection
import os
from datetime import datetime

class ImportTemplateGenerator:
    """导入模板生成器"""

    def __init__(self, template_id=None, sample_type_id=None):
        """
        初始化生成器

        Args:
            template_id: 报告模板ID（可选）
            sample_type_id: 样品类型ID（可选）
        """
        self.template_id = template_id
        self.sample_type_id = sample_type_id
        self.template_info = None
        self.sample_type_info = None
        self.indicators = []

    def generate(self, output_path=None):
        """
        生成导入模板

        Args:
            output_path: 输出文件路径

        Returns:
            str: 生成的文件路径
        """
        # 加载数据
        self._load_data()

        # 创建工作簿
        wb = openpyxl.Workbook()

        # 移除默认的Sheet
        wb.remove(wb.active)

        # 1. 创建说明sheet（放在第一个）
        self._create_instruction_sheet(wb)

        # 2. 创建检测数据sheet（新格式）
        self._create_detection_data_sheet_new(wb)

        # 3. 保存文件
        if output_path is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            sample_type_name = self.sample_type_info['name'] if self.sample_type_info else '通用'
            output_path = f"exports/import_template_{sample_type_name}_{timestamp}.xlsx"

        os.makedirs('exports', exist_ok=True)
        wb.save(output_path)
        wb.close()

        return output_path

    def _load_data(self):
        """加载模板和样品类型信息"""
        conn = get_db_connection()

        # 加载模板信息
        if self.template_id:
            template = conn.execute(
                'SELECT * FROM excel_report_templates WHERE id = ?',
                (self.template_id,)
            ).fetchone()

            if template:
                self.template_info = dict(template)

                # 获取字段映射
                fields = conn.execute(
                    'SELECT * FROM template_field_mappings WHERE template_id = ? ORDER BY sort_order',
                    (self.template_id,)
                ).fetchall()
                self.template_info['fields'] = [dict(f) for f in fields]

        # 加载样品类型信息和关联的检测指标
        if self.sample_type_id:
            sample_type = conn.execute(
                'SELECT * FROM sample_types WHERE id = ?',
                (self.sample_type_id,)
            ).fetchone()

            if sample_type:
                self.sample_type_info = dict(sample_type)

                # 获取该样品类型关联的所有检测指标
                indicators = conn.execute('''
                    SELECT i.id, i.group_id, i.name, i.unit, i.default_value,
                           i.description, i.created_at, i.limit_value,
                           i.detection_method, i.remark, ti.sort_order
                    FROM indicators i
                    JOIN template_indicators ti ON i.id = ti.indicator_id
                    WHERE ti.sample_type_id = ?
                    ORDER BY ti.sort_order, i.id
                ''', (self.sample_type_id,)).fetchall()

                self.indicators = [dict(ind) for ind in indicators]

        conn.close()

    def _create_detection_data_sheet_new(self, wb):
        """
        创建检测数据sheet（简化格式）
        格式：
        - A列: 检测项目
        - B列: 单位（参考）
        - C列及以后: 样品编号（首行）及其检测结果
        """
        ws = wb.create_sheet("检测数据")

        # 设置样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        subheader_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
        subheader_font = Font(bold=True, size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 第一行：表头
        headers = ['检测项目', '单位']

        # 添加示例样品编号列（3个样品）
        sample_numbers = ['样品编号1*', '样品编号2', '样品编号3']

        for col_idx, header in enumerate(headers + sample_numbers, start=1):
            cell = ws.cell(1, col_idx)
            cell.value = header

            if col_idx <= len(headers):
                # 项目信息列
                cell.fill = subheader_fill
                cell.font = subheader_font
            else:
                # 样品列
                cell.fill = header_fill
                cell.font = header_font

            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # 如果有样品类型和指标数据，填充检测项目信息
        if self.indicators and len(self.indicators) > 0:
            for row_idx, indicator in enumerate(self.indicators, start=2):
                # 检测项目
                cell = ws.cell(row_idx, 1)
                cell.value = indicator['name']
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = border

                # 单位
                cell = ws.cell(row_idx, 2)
                cell.value = indicator.get('unit', '')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

                # 样品结果列（示例数据在第一个样品）
                if row_idx == 2:
                    # 仅第一行第一个样品填写示例数据
                    cell = ws.cell(row_idx, 3)
                    cell.value = indicator.get('default_value', '')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border
                else:
                    # 其他行留空
                    for col_idx in range(3, 3 + len(sample_numbers)):
                        cell = ws.cell(row_idx, col_idx)
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            # 没有指标数据，提供示例
            example_indicators = [
                {'name': 'pH', 'unit': '无量纲', 'example': '7.2'},
                {'name': '浊度', 'unit': 'NTU', 'example': '0.5'},
                {'name': '余氯', 'unit': 'mg/L', 'example': '0.3'},
                {'name': '色度', 'unit': '度', 'example': '<5'},
                {'name': '臭和味', 'unit': '级', 'example': '无'},
                {'name': '肉眼可见物', 'unit': '', 'example': '无'},
                {'name': '耗氧量', 'unit': 'mg/L', 'example': '1.2'},
                {'name': '总大肠菌群', 'unit': 'CFU/100mL', 'example': '未检出'},
                {'name': '菌落总数', 'unit': 'CFU/mL', 'example': '<1'},
            ]

            for row_idx, ind in enumerate(example_indicators, start=2):
                # 检测项目
                cell = ws.cell(row_idx, 1)
                cell.value = ind['name']
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = border

                # 单位
                cell = ws.cell(row_idx, 2)
                cell.value = ind['unit']
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

                # 第一个样品的示例值
                cell = ws.cell(row_idx, 3)
                cell.value = ind['example']
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

                # 其他样品列留空但设置边框
                for col_idx in range(4, 3 + len(sample_numbers)):
                    cell = ws.cell(row_idx, col_idx)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        # 调整列宽
        ws.column_dimensions['A'].width = 20  # 检测项目
        ws.column_dimensions['B'].width = 12  # 单位

        # 样品列
        for col_letter in ['C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col_letter].width = 15

        # 冻结首行和前2列
        ws.freeze_panes = 'C2'

    def _create_instruction_sheet(self, wb):
        """创建填写说明sheet"""
        ws = wb.create_sheet("填写说明", 0)  # 插入到第一个位置

        # 标题
        ws['A1'] = "导入模板填写说明"
        ws['A1'].font = Font(size=16, bold=True, color="4472C4")

        row = 3

        # 基本说明
        sample_type_name = self.sample_type_info['name'] if self.sample_type_info else '未指定'

        instructions = [
            ("一、模板信息", ""),
            ("", f"样品类型: {sample_type_name}"),
            ("", f"检测项目数量: {len(self.indicators) if self.indicators else '示例数据'}"),
            ("", f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"),
            ("", ""),
            ("二、填写说明", ""),
            ("", "【检测数据】sheet 说明:"),
            ("", "  - 格式: 简化横向表格"),
            ("", "  - A列: 检测项目名称（请勿修改）"),
            ("", "  - B列: 单位（参考用）"),
            ("", "  - C列起: 样品编号（首行）及其检测结果"),
            ("", ""),
            ("", "【填写步骤】:"),
            ("", "  1. 在首行C列起修改样品编号（如 W260105C08, S20250128001...）"),
            ("", "     注意：第一个样品列（C列）为必填"),
            ("", "  2. 在对应列下方填写该样品各检测项目的检测结果"),
            ("", "  3. 如需增加样品，直接在右侧添加新列即可"),
            ("", "  4. 可以删除示例数据，但请保留表头行"),
            ("", ""),
            ("三、注意事项", ""),
            ("", "1. 样品编号必须唯一"),
            ("", "2. 检测结果直接填写数值或文本（如：7.2、未检出、<1）"),
            ("", "3. 不需要的检测项目可以留空，但不要删除该行"),
            ("", "4. 不要修改A列的检测项目名称"),
            ("", "5. 如果是特殊结果（如未检出、无等），直接输入文本即可"),
            ("", "6. 表格已冻结首行和前2列，方便浏览大量数据"),
            ("", ""),
            ("四、导入步骤", ""),
            ("", "1. 在系统中选择对应的报告模板和样品类型"),
            ("", "2. 点击\"下载导入模板\"按钮获取本模板"),
            ("", "3. 按照说明填写检测数据"),
            ("", "4. 点击\"批量导入\"按钮上传填写好的文件"),
            ("", "5. 系统将自动创建报告并填充数据"),
            ("", ""),
            ("五、示例", ""),
            ("", "表格格式示例："),
            ("", "检测项目  | 单位      | W260105C08 | W260105C09 | W260105C10"),
            ("", "pH        | 无量纲    | 7.2        | 7.3        | 7.1"),
            ("", "浊度      | NTU       | 0.5        | 0.6        | 0.4"),
            ("", "余氯      | mg/L      | 0.3        | 0.35       | 0.28"),
        ]

        for title, content in instructions:
            if title:
                cell = ws.cell(row, 1)
                cell.value = title
                cell.font = Font(size=12, bold=True, color="4472C4")
            else:
                cell = ws.cell(row, 1)
                cell.value = content
                cell.alignment = Alignment(wrap_text=True)

            row += 1

        # 调整列宽
        ws.column_dimensions['A'].width = 100


def generate_import_template(template_id=None, sample_type_id=None, output_path=None):
    """
    生成导入模板的便捷函数

    Args:
        template_id: 报告模板ID（可选）
        sample_type_id: 样品类型ID（可选）
        output_path: 输出路径（可选）

    Returns:
        str: 生成的文件路径
    """
    generator = ImportTemplateGenerator(template_id, sample_type_id)
    return generator.generate(output_path)


if __name__ == '__main__':
    print("="*60)
    print("Excel导入模板生成器测试")
    print("="*60)

    # 生成通用模板
    print("\n生成通用导入模板...")
    path = generate_import_template()
    print(f"✓ 生成成功: {path}")

    # 如果有样品类型，生成特定模板
    conn = get_db_connection()
    sample_types = conn.execute('SELECT id, name FROM sample_types LIMIT 1').fetchone()

    if sample_types:
        sample_type_id = sample_types['id']
        sample_type_name = sample_types['name']
        print(f"\n生成样品类型 [{sample_type_name}] 的导入模板...")
        path = generate_import_template(None, sample_type_id)
        print(f"✓ 生成成功: {path}")

    conn.close()
