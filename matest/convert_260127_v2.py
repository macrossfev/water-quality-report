"""
Convert 260127-1-4.xlsx into transposed raw_data import format.

Output layout:
  Row 1: A1="样品编号 →", B1=W260127C01, C1=W260127C02, ...
  Col A: field names (报告编号, 被检单位, ... 检测指标...)
  Data:  intersection values
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import re
import os

SRC = os.path.join(os.path.dirname(__file__), '260127-1-4.xlsx')
DST = os.path.join(os.path.dirname(__file__), '260127-1-4_import_v2.xlsx')


def clean_param(name):
    """Normalize parameter name: collapse whitespace, strip, keep units."""
    if not name:
        return ''
    name = str(name).replace('\n', '')
    name = re.sub(r'\s{2,}', '', name)
    name = name.strip()
    # Fix known typos
    name = name.replace('肉哏可见物', '肉眼可见物')
    # Normalize CaCO2 -> CaCO3
    name = name.replace('CaCO₂', 'CaCO₃')
    # Remove trailing V (volatile marker from instrument)
    name = re.sub(r'V$', '', name)
    # Fix inconsistent spacing before units: "xxx (mg/L)" -> "xxx(mg/L)"
    name = re.sub(r'\s+\(', '(', name)
    # Fix missing closing parenthesis
    if '(' in name and ')' not in name:
        name += ')'
    # Normalize "阴离子合成洗涤剂阴离子表面活性剂" -> "阴离子合成洗涤剂"
    name = name.replace('阴离子合成洗涤剂阴离子表面活性剂', '阴离子合成洗涤剂')
    return name


def sample_type_from_location(loc):
    """Derive 样品类型 from the location description in Sheet1."""
    if not loc:
        return ''
    if '出厂水' in loc:
        return '出厂水'
    if '管网水' in loc:
        return '管网水'
    if '原水' in loc:
        return '原水'
    return ''


# ── 1. Read Sheet1 – sample registration ─────────────────────────────────
wb = openpyxl.load_workbook(SRC, data_only=True)

ws1 = wb['Sheet1']
samples = []
for r in range(4, ws1.max_row + 1):
    sid = ws1.cell(r, 4).value  # column D = 样品编号
    loc = ws1.cell(r, 2).value  # column B = 采样地点/样品类型
    if not sid or not str(sid).strip().startswith('W'):
        continue  # skip blank samples and non-W entries
    samples.append({
        '样品编号': str(sid).strip(),
        '被检单位': '双庆水厂',
        '被检水厂': '双庆水厂',
        '样品类型': sample_type_from_location(loc),
        '采样日期': '2026-01-27',
    })

sample_ids = [s['样品编号'] for s in samples]
print(f"Samples ({len(sample_ids)}): {sample_ids}")

# ── 2. Read data sheets ──────────────────────────────────────────────────
# { sample_id: { param_name: value } }
data = {sid: {} for sid in sample_ids}


def read_sheet_vertical(ws, header_row):
    """Sheets where params are in col A (merged A:C) and sample IDs in header row."""
    col_map = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(header_row, c).value
        if val:
            val = str(val).strip()
            if val in sample_ids:
                col_map[c] = val
    for r in range(header_row + 1, ws.max_row + 1):
        param_raw = ws.cell(r, 1).value
        if not param_raw:
            continue
        param = clean_param(param_raw)
        if not param:
            continue
        for c, sid in col_map.items():
            val = ws.cell(r, c).value
            if val is not None:
                data[sid][param] = val


def read_sheet_horizontal(ws, header_row):
    """Sheets where sample IDs in col A and params in header row."""
    param_map = {}
    for c in range(2, ws.max_column + 1):
        val = ws.cell(header_row, c).value
        if val:
            param_map[c] = clean_param(val)
    for r in range(header_row + 1, ws.max_row + 1):
        sid_raw = ws.cell(r, 1).value
        if not sid_raw:
            continue
        sid = str(sid_raw).strip()
        if sid not in sample_ids:
            continue
        for c, param in param_map.items():
            val = ws.cell(r, c).value
            if val is not None:
                data[sid][param] = val


# Sheet2: vertical (params in col A merged A:C, sample IDs in row 3, cols 4+)
read_sheet_vertical(wb['Sheet2'], header_row=3)

# Sheet3: vertical (params in col A merged A:B, sample IDs in row 2, cols 3+)
read_sheet_vertical(wb['Sheet3'], header_row=2)

# Sheet4: horizontal (sample IDs in col A, params in row 3)
read_sheet_horizontal(wb['Sheet4'], header_row=3)

# Sheet5: horizontal (sample IDs in col A, params in row 3)
read_sheet_horizontal(wb['Sheet5'], header_row=3)

wb.close()

# ── 3. Build ordered parameter list ─────────────────────────────────────
all_params = []
seen = set()
for sid in sample_ids:
    for p in data[sid]:
        if p not in seen:
            all_params.append(p)
            seen.add(p)

print(f"\nParameters ({len(all_params)}):")
for i, p in enumerate(all_params, 1):
    print(f"  {i}. {p}")

# ── 4. Write transposed output ───────────────────────────────────────────
out_wb = openpyxl.Workbook()
ws_out = out_wb.active
ws_out.title = '数据导入'

# Styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
required_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
required_font = Font(bold=True, size=10, color="C00000")
label_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
label_font = Font(bold=True, size=10)
border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Row 1: header with sample IDs
cell_a1 = ws_out.cell(1, 1, "样品编号 →")
cell_a1.fill = header_fill
cell_a1.font = header_font
cell_a1.alignment = Alignment(horizontal='center', vertical='center')
cell_a1.border = border

for col_offset, s in enumerate(samples):
    col_idx = col_offset + 2
    cell = ws_out.cell(1, col_idx, s['样品编号'])
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# Base fields rows
BASE_ROW_FIELDS = ['报告编号', '被检单位', '被检水厂', '样品类型', '采样日期']
all_row_fields = BASE_ROW_FIELDS + all_params

for row_offset, field_name in enumerate(all_row_fields):
    row_idx = row_offset + 2

    # Column A: field label
    label_cell = ws_out.cell(row_idx, 1, field_name)
    if field_name in BASE_ROW_FIELDS:
        label_cell.fill = required_fill
        label_cell.font = required_font
    else:
        label_cell.fill = label_fill
        label_cell.font = label_font
    label_cell.alignment = Alignment(horizontal='left', vertical='center')
    label_cell.border = border

    # Data columns
    for col_offset, s in enumerate(samples):
        col_idx = col_offset + 2
        sid = s['样品编号']

        if field_name in BASE_ROW_FIELDS:
            val = s.get(field_name, '')
        else:
            val = data[sid].get(field_name)

        cell = ws_out.cell(row_idx, col_idx)
        if val is not None:
            cell.value = val
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

# Adjust column widths
ws_out.column_dimensions['A'].width = 28
for col_idx in range(2, len(samples) + 2):
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    ws_out.column_dimensions[col_letter].width = 18

# Freeze panes
ws_out.freeze_panes = 'B2'

out_wb.save(DST)
out_wb.close()
print(f"\nSaved to {DST}")

# ── 5. Verify ────────────────────────────────────────────────────────────
vwb = openpyxl.load_workbook(DST)
vws = vwb.active
print(f"\nOutput: {vws.max_row} rows x {vws.max_column} cols")
print(f"  Row 1 (样品编号): ", end='')
for c in range(1, vws.max_column + 1):
    v = vws.cell(1, c).value
    print(f"{v}  ", end='')
print()
print(f"  Fields (col A):")
for r in range(2, vws.max_row + 1):
    field = vws.cell(r, 1).value
    vals = []
    for c in range(2, vws.max_column + 1):
        v = vws.cell(r, c).value
        vals.append(str(v) if v is not None else '-')
    print(f"    {field}: {' | '.join(vals)}")
vwb.close()
