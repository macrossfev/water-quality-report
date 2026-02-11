#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
水质检测报告文件全面分析脚本
扫描所有 .xlsx / .xls 报告，按类别输出待确认问题清单。
"""

import os, re, sys, traceback
from collections import defaultdict, Counter
from datetime import datetime

import openpyxl
import xlrd

REPORT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(REPORT_DIR, "待确认问题清单.txt")

# ──────────────────────────── helpers ────────────────────────────

def extract_number_prefix(fname):
    """Return the leading numeric string of a filename, e.g. '0001'."""
    m = re.match(r'^(\d+)', fname)
    return m.group(1) if m else None

def classify_water_type(fname):
    """Classify report type from filename."""
    if '二次供水' in fname:
        return '二次供水'
    if '农饮水' in fname or '生活饮用水' in fname:
        return '农饮水'
    if '转供水' in fname:
        return '转供水'
    if '日检九项' in fname:
        return '日检九项'
    if '送检' in fname:
        return '送检'
    if '高锰酸盐指数' in fname:
        return '高锰酸盐指数'
    if '原水' in fname:
        return '原水'
    if '出厂水' in fname:
        return '出厂水'
    if '管网' in fname:
        return '管网水'
    return '未知'

def extract_plant_name(fname):
    """Try to extract water plant name from filename."""
    # Remove prefix number
    name = re.sub(r'^\d+', '', fname)
    # Remove extension
    name = re.sub(r'\.(xlsx?|xls)$', '', name)
    # Remove date suffixes like 01.05
    name = re.sub(r'\s*\d{2}\.\d{2}\s*$', '', name)
    # Remove trailing markers
    for tag in ['-送检', '送检', '-荣昌', '荣昌', '日检九项', '高锰酸盐指数',
                '-地表三类', '地表三类', '应急水样', '-应急水样']:
        name = name.replace(tag, '')
    name = name.strip(' -')
    # Try to get the plant name before the water-type bracket
    # e.g. 北门水厂（出厂水） -> 北门水厂
    m = re.match(r'^(.+?水厂|.+?水库|.+?泵站).*', name)
    if m:
        plant = m.group(1)
        # Normalize: remove 管网水 prefix patterns
        plant = re.sub(r'管网水$', '水厂', plant)
        return plant.strip()
    # For things like 小北海（出厂水）
    m = re.match(r'^([^（(]+)', name)
    if m:
        return m.group(1).strip()
    return name.strip()

# ──────────────────── reading Excel data ────────────────────

def read_xlsx_report_info(filepath):
    """Read key metadata from an xlsx file."""
    info = {}
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        info['sheet_count'] = len(wb.sheetnames)
        info['sheet_names'] = wb.sheetnames

        # Page 1 (cover page) - try first sheet
        ws1 = wb[wb.sheetnames[0]]
        info['total_rows_sheet1'] = ws1.max_row
        info['total_cols_sheet1'] = ws1.max_column

        # Extract report number from B1
        b1 = ws1.cell(1, 2).value
        if b1:
            info['report_number_raw'] = str(b1).strip()
            m = re.search(r'第\s*\(\s*(\d+)\s*\)\s*号', str(b1))
            if m:
                info['report_number'] = m.group(1).strip()

        # Extract page info from B2
        b2 = ws1.cell(2, 2).value
        if b2:
            m = re.search(r'共\s*(\d+)\s*页', str(b2))
            if m:
                info['total_pages'] = int(m.group(1))

        # Sample name from C8 or C9 (row 8)
        for r in range(7, 13):
            cv = ws1.cell(r, 3).value
            if cv and ('水' in str(cv) or '【' in str(cv)):
                info['sample_name'] = str(cv).strip()
                break

        # Company from C9
        for r in range(8, 13):
            cv = ws1.cell(r, 3).value
            if cv and '公司' in str(cv):
                info['company'] = str(cv).strip()
                break

        # Report date from C12 or C11
        for r in range(10, 14):
            bv = ws1.cell(r, 2).value
            cv = ws1.cell(r, 3).value
            if bv and '报告编制日期' in str(bv) and cv:
                info['report_date'] = str(cv).strip()
                break

        # Page 2 (检测结果) - try second sheet
        if len(wb.sheetnames) >= 2:
            ws2 = wb[wb.sheetnames[1]]
            info['total_rows_sheet2'] = ws2.max_row

            # Sample type from C3
            c3 = ws2.cell(3, 3).value
            if c3:
                info['sample_type'] = str(c3).strip()

            # Sampler from C4
            c4 = ws2.cell(4, 3).value
            if c4:
                info['sampler'] = str(c4).strip()

            # Sampling date from E4
            e4 = ws2.cell(4, 5).value
            if e4:
                info['sampling_date'] = str(e4).strip()

            # Receipt date from E5
            e5 = ws2.cell(5, 5).value
            if e5:
                info['receipt_date'] = str(e5).strip()

            # Sampling location from C6
            c6 = ws2.cell(6, 3).value
            if c6:
                info['sampling_location'] = str(c6).strip()

            # Sample ID from C8
            c8 = ws2.cell(8, 3).value
            if c8:
                info['sample_id'] = str(c8).strip()

            # Testing date from E8
            e8 = ws2.cell(8, 5).value
            if e8:
                info['testing_date'] = str(e8).strip()

            # Product standard from C9
            c9 = ws2.cell(9, 3).value
            if c9:
                info['product_standard'] = str(c9).strip()

            # Number of test items from C10
            c10 = ws2.cell(10, 3).value
            if c10:
                info['test_items_desc'] = str(c10).strip()
                m = re.search(r'(\d+)\s*项', str(c10))
                if m:
                    info['test_item_count'] = int(m.group(1))

            # Conclusion from B13
            b13 = ws2.cell(13, 2).value
            if b13:
                info['conclusion'] = str(b13).strip()

        # Page 3+ (检测数据) - collect test items
        test_items = []
        for si in range(2, len(wb.sheetnames)):
            ws = wb[wb.sheetnames[si]]
            for r in range(1, ws.max_row + 1):
                a_val = ws.cell(r, 1).value
                b_val = ws.cell(r, 2).value
                d_val = ws.cell(r, 4).value
                if a_val is not None and b_val is not None:
                    try:
                        seq = int(float(str(a_val)))
                        if 1 <= seq <= 100 and b_val:
                            item = {
                                'seq': seq,
                                'name': str(b_val).strip(),
                                'unit': str(ws.cell(r, 3).value or '').strip(),
                                'result': str(d_val).strip() if d_val is not None else '',
                                'standard': str(ws.cell(r, 5).value or '').strip(),
                                'method': str(ws.cell(r, 6).value or '').strip(),
                            }
                            test_items.append(item)
                    except (ValueError, TypeError):
                        pass
        info['test_items'] = test_items
        wb.close()
    except Exception as e:
        info['read_error'] = f"{type(e).__name__}: {e}"
    return info


def read_xls_report_info(filepath):
    """Read key metadata from an xls file."""
    info = {}
    try:
        wb = xlrd.open_workbook(filepath)
        info['sheet_count'] = wb.nsheets
        info['sheet_names'] = wb.sheet_names()

        ws1 = wb.sheet_by_index(0)
        info['total_rows_sheet1'] = ws1.nrows
        info['total_cols_sheet1'] = ws1.ncols

        # Report number from B1 (row 0, col 1)
        if ws1.nrows > 0 and ws1.ncols > 1:
            b1 = ws1.cell_value(0, 1)
            if b1:
                info['report_number_raw'] = str(b1).strip()
                m = re.search(r'第\s*\(\s*(\d+)\s*\)\s*号', str(b1))
                if m:
                    info['report_number'] = m.group(1).strip()

        # Page info from B2 (row 1, col 1)
        if ws1.nrows > 1 and ws1.ncols > 1:
            b2 = ws1.cell_value(1, 1)
            if b2:
                m = re.search(r'共\s*(\d+)\s*页', str(b2))
                if m:
                    info['total_pages'] = int(m.group(1))

        # Sample name from C8 (row 7, col 2)
        for r in range(6, min(12, ws1.nrows)):
            if ws1.ncols > 2:
                cv = ws1.cell_value(r, 2)
                if cv and ('水' in str(cv) or '【' in str(cv)):
                    info['sample_name'] = str(cv).strip()
                    break

        # Company
        for r in range(7, min(12, ws1.nrows)):
            if ws1.ncols > 2:
                cv = ws1.cell_value(r, 2)
                if cv and '公司' in str(cv):
                    info['company'] = str(cv).strip()
                    break

        # Report date
        for r in range(9, min(14, ws1.nrows)):
            if ws1.ncols > 1:
                bv = ws1.cell_value(r, 1) if ws1.ncols > 1 else ''
                cv = ws1.cell_value(r, 2) if ws1.ncols > 2 else ''
                if bv and '报告编制日期' in str(bv) and cv:
                    info['report_date'] = str(cv).strip()
                    break

        # Page 2
        if wb.nsheets >= 2:
            ws2 = wb.sheet_by_index(1)
            info['total_rows_sheet2'] = ws2.nrows

            def sv(r, c):
                if r < ws2.nrows and c < ws2.ncols:
                    return ws2.cell_value(r, c)
                return None

            c3 = sv(2, 2)
            if c3:
                info['sample_type'] = str(c3).strip()

            c4 = sv(3, 2)
            if c4:
                info['sampler'] = str(c4).strip()

            e4 = sv(3, 4)
            if e4:
                info['sampling_date'] = str(e4).strip()

            e5 = sv(4, 4)
            if e5:
                info['receipt_date'] = str(e5).strip()

            c6 = sv(5, 2)
            if c6:
                info['sampling_location'] = str(c6).strip()

            c8 = sv(7, 2)
            if c8:
                info['sample_id'] = str(c8).strip()

            e8 = sv(7, 4)
            if e8:
                info['testing_date'] = str(e8).strip()

            c9 = sv(8, 2)
            if c9:
                info['product_standard'] = str(c9).strip()

            c10 = sv(9, 2)
            if c10:
                info['test_items_desc'] = str(c10).strip()
                m = re.search(r'(\d+)\s*项', str(c10))
                if m:
                    info['test_item_count'] = int(m.group(1))

            b13 = sv(12, 1)
            if b13:
                info['conclusion'] = str(b13).strip()

        # Test items from page 3+
        test_items = []
        for si in range(2, wb.nsheets):
            ws = wb.sheet_by_index(si)
            for r in range(ws.nrows):
                if ws.ncols >= 6:
                    a_val = ws.cell_value(r, 0)
                    b_val = ws.cell_value(r, 1)
                    d_val = ws.cell_value(r, 3)
                    if a_val not in ('', None) and b_val not in ('', None):
                        try:
                            seq = int(float(str(a_val)))
                            if 1 <= seq <= 100 and b_val:
                                item = {
                                    'seq': seq,
                                    'name': str(b_val).strip(),
                                    'unit': str(ws.cell_value(r, 2) or '').strip(),
                                    'result': str(d_val).strip() if d_val not in ('', None) else '',
                                    'standard': str(ws.cell_value(r, 4) or '').strip(),
                                    'method': str(ws.cell_value(r, 5) or '').strip(),
                                }
                                test_items.append(item)
                        except (ValueError, TypeError):
                            pass
        info['test_items'] = test_items

    except Exception as e:
        info['read_error'] = f"{type(e).__name__}: {e}"
    return info


# ─────────────────── MAIN ANALYSIS ───────────────────

def main():
    files = sorted([f for f in os.listdir(REPORT_DIR)
                    if f.endswith(('.xlsx', '.xls')) and not f.startswith('~')])

    print(f"共找到 {len(files)} 个报告文件，开始分析...")

    # ── Collect all file info ──
    all_info = {}  # fname -> info dict
    for i, fname in enumerate(files):
        filepath = os.path.join(REPORT_DIR, fname)
        if fname.endswith('.xlsx'):
            info = read_xlsx_report_info(filepath)
        else:
            info = read_xls_report_info(filepath)
        info['filename'] = fname
        info['extension'] = os.path.splitext(fname)[1]
        info['prefix'] = extract_number_prefix(fname)
        info['water_type'] = classify_water_type(fname)
        info['plant_name'] = extract_plant_name(fname)
        all_info[fname] = info
        if (i + 1) % 20 == 0:
            print(f"  已处理 {i+1}/{len(files)} ...")

    print(f"文件读取完成，开始问题检测...")

    # ════════════════════════════════════════════════════
    # Issue categories
    # ════════════════════════════════════════════════════
    issues_naming = []       # 一、命名问题
    issues_numbering = []    # 二、编号问题
    issues_data = []         # 三、数据问题
    issues_format = []       # 四、格式/模板问题
    issues_date = []         # 五、日期问题
    issues_consistency = []  # 六、一致性问题
    issues_values = []       # 七、异常值问题
    issues_read_errors = []  # 八、文件读取问题

    # ──────────── 一、命名问题 ────────────
    # 1. Check bracket matching in filenames
    for fname in files:
        name_part = os.path.splitext(fname)[0]
        open_cn = name_part.count('（')
        close_cn = name_part.count('）')
        open_en = name_part.count('(')
        close_en = name_part.count(')')
        if open_cn != close_cn:
            issues_naming.append(f"文件 \"{fname}\" 中文括号不匹配：'（' 出现 {open_cn} 次，'）' 出现 {close_cn} 次")
        if open_en != close_en:
            issues_naming.append(f"文件 \"{fname}\" 英文括号不匹配：'(' 出现 {open_en} 次，')' 出现 {close_en} 次")

    # 2. Check for inconsistent numbering prefix length
    prefixes = [(fname, extract_number_prefix(fname)) for fname in files]
    prefix_lengths = Counter(len(p) for _, p in prefixes if p)
    if len(prefix_lengths) > 1:
        for fname, p in prefixes:
            if p and len(p) != 4:
                issues_naming.append(f"文件 \"{fname}\" 编号前缀位数异常：'{p}' 为 {len(p)} 位，多数文件为 4 位")

    # 3. Number sequence gaps
    nums = sorted(set(int(p) for _, p in prefixes if p))
    expected = set(range(1, max(nums) + 1))
    actual = set(nums)
    missing = sorted(expected - actual)
    if missing:
        # Group consecutive missing numbers
        groups = []
        start = missing[0]
        end = missing[0]
        for n in missing[1:]:
            if n == end + 1:
                end = n
            else:
                groups.append((start, end))
                start = n
                end = n
        groups.append((start, end))
        for s, e in groups:
            if s == e:
                issues_naming.append(f"文件编号序列缺失：{s:04d}")
            else:
                issues_naming.append(f"文件编号序列缺失：{s:04d}-{e:04d}（共 {e-s+1} 个）")

    # 4. Duplicate prefix numbers
    prefix_counter = Counter(int(p) for _, p in prefixes if p)
    for num, cnt in sorted(prefix_counter.items()):
        if cnt > 1:
            dup_files = [f for f, p in prefixes if p and int(p) == num]
            issues_naming.append(f"文件编号重复：编号 {num:04d} 出现 {cnt} 次，涉及文件：{', '.join(dup_files)}")

    # 5. Inconsistent naming patterns
    # Check for extra spaces in filenames
    for fname in files:
        if '  ' in fname:
            issues_naming.append(f"文件 \"{fname}\" 名称中包含连续空格")
        if fname != fname.strip():
            issues_naming.append(f"文件 \"{fname}\" 名称首尾有多余空格")

    # 6. Check for "水厂水厂" duplicated in name
    for fname in files:
        if '水厂水厂' in fname:
            issues_naming.append(f"文件 \"{fname}\" 名称中 '水厂' 重复出现，可能为笔误")

    # 7. Check for inconsistent date suffixes in filenames
    # Some files have date suffix like "01.05" and some don't
    files_with_date = [f for f in files if re.search(r'\d{2}\.\d{2}\.(xlsx?|xls)$', f) or
                       re.search(r'\d{2}\.\d{2}\s*\.(xlsx?|xls)$', f)]
    files_without_date = [f for f in files if f not in files_with_date]
    if files_with_date and files_without_date and len(files_with_date) < len(files_without_date):
        issues_naming.append(
            f"部分文件名含日期后缀（共 {len(files_with_date)} 个），"
            f"而多数文件不含日期后缀（共 {len(files_without_date)} 个），格式不统一。"
            f"含日期的文件：{', '.join(files_with_date)}")

    # 8. Check extension vs water type consistency
    # Expectation: 原水 -> .xls, 出厂水/管网水 -> .xlsx (general pattern)
    for fname, info in all_info.items():
        wt = info['water_type']
        ext = info['extension']
        if wt == '原水' and ext == '.xlsx':
            issues_naming.append(f"文件 \"{fname}\" 为原水报告但使用 .xlsx 格式，"
                                 f"一般原水报告使用 .xls 格式，请确认")
        if wt in ('出厂水',) and ext == '.xls':
            issues_naming.append(f"文件 \"{fname}\" 为出厂水报告但使用 .xls 格式，"
                                 f"一般出厂水报告使用 .xlsx 格式，请确认")
        # 管网水 can be either, but check for xls ones (pattern: 管网水 xls typically only for early files with brackets)
        if wt == '二次供水' and ext == '.xlsx':
            issues_naming.append(f"文件 \"{fname}\" 为二次供水报告但使用 .xlsx 格式，"
                                 f"一般二次供水报告使用 .xls 格式，请确认")

    # 9. Check for inconsistent water type labeling in filename
    # Two patterns exist: "水厂（管网水）" and "水厂管网水" -- flag the inconsistency as a whole
    guanwang_files = [f for f in files if '管网' in f]
    guanwang_bracket = [f for f in guanwang_files if '（管网' in f or '（管网' in f]
    guanwang_no_bracket = [f for f in guanwang_files if f not in guanwang_bracket]
    if guanwang_bracket and guanwang_no_bracket:
        issues_naming.append(
            f"管网水文件命名格式不统一：{len(guanwang_bracket)} 个文件使用括号形式如 '水厂（管网水）'，"
            f"{len(guanwang_no_bracket)} 个文件使用无括号形式如 '水厂管网水'，建议统一命名规范")

    # ──────────── 二、编号问题 ────────────
    # Check report numbers from file content vs filename prefix
    report_nums = {}  # report_number -> [filenames]
    for fname, info in all_info.items():
        rn = info.get('report_number')
        if rn:
            report_nums.setdefault(rn, []).append(fname)
            # Compare with filename prefix
            prefix = info.get('prefix')
            if prefix:
                # Normalize: remove leading zeros for comparison
                try:
                    if int(rn) != int(prefix):
                        issues_numbering.append(
                            f"文件 \"{fname}\" 的文件名编号 ({prefix}) 与报告内编号 ({rn}) 不一致")
                except ValueError:
                    issues_numbering.append(
                        f"文件 \"{fname}\" 的报告内编号 '{rn}' 无法解析为数字")

    # Check for duplicate report numbers in content
    for rn, fnames in sorted(report_nums.items(), key=lambda x: x[0]):
        if len(fnames) > 1:
            issues_numbering.append(
                f"报告编号 {rn} 在多个文件中重复使用：{', '.join(fnames)}")

    # Check for missing report numbers
    for fname, info in all_info.items():
        if 'report_number' not in info and 'read_error' not in info:
            issues_numbering.append(f"文件 \"{fname}\" 未能提取到报告编号")

    # ──────────── 三、数据问题 ────────────
    for fname, info in all_info.items():
        # Missing sample name
        if 'sample_name' not in info and 'read_error' not in info:
            issues_data.append(f"文件 \"{fname}\" 未提取到样品名称")

        # Missing company
        if 'company' not in info and 'read_error' not in info:
            issues_data.append(f"文件 \"{fname}\" 未提取到被检单位名称")

        # Missing sampling date
        if 'sampling_date' not in info and 'read_error' not in info:
            issues_data.append(f"文件 \"{fname}\" 未提取到采样日期")

        # Missing sample ID
        if 'sample_id' not in info and 'read_error' not in info:
            issues_data.append(f"文件 \"{fname}\" 未提取到样品编号")

        # Check test items for blank results
        test_items = info.get('test_items', [])
        blank_items = [item['name'] for item in test_items if not item['result'] or item['result'] == 'None']
        if blank_items:
            issues_data.append(
                f"文件 \"{fname}\" 以下检测项目结果为空：{', '.join(blank_items)}")

        # Check test items for blank methods
        no_method_items = [item['name'] for item in test_items if not item['method'] or item['method'] == 'None']
        if no_method_items:
            issues_data.append(
                f"文件 \"{fname}\" 以下检测项目缺少检测方法：{', '.join(no_method_items)}")

        # Check declared test item count vs actual
        declared = info.get('test_item_count')
        actual_count = len(test_items)
        if declared and actual_count > 0:
            if actual_count != declared:
                issues_data.append(
                    f"文件 \"{fname}\" 声称检测 {declared} 项指标，"
                    f"但实际提取到 {actual_count} 项数据")

        # Missing conclusion
        if 'conclusion' not in info and 'read_error' not in info:
            issues_data.append(f"文件 \"{fname}\" 未提取到检测结论")

        # Missing report date
        if 'report_date' not in info and 'read_error' not in info:
            issues_data.append(f"文件 \"{fname}\" 未提取到报告编制日期")

    # ──────────── 四、格式/模板问题 ────────────
    # Group by water type, compare sheet counts and structures
    type_groups = defaultdict(list)
    for fname, info in all_info.items():
        wt = info['water_type']
        type_groups[wt].append((fname, info))

    for wt, group in type_groups.items():
        if len(group) < 2:
            continue

        # Compare sheet counts
        sheet_counts = Counter(info.get('sheet_count', 0) for _, info in group)
        if len(sheet_counts) > 1:
            most_common_count = sheet_counts.most_common(1)[0][0]
            for fname, info in group:
                sc = info.get('sheet_count', 0)
                if sc != most_common_count:
                    issues_format.append(
                        f"文件 \"{fname}\"（{wt}类）共 {sc} 个工作表，"
                        f"而同类报告多数为 {most_common_count} 个工作表")

        # Compare total pages
        page_counts = Counter(info.get('total_pages', 0) for _, info in group)
        if len(page_counts) > 1:
            most_common_pages = page_counts.most_common(1)[0][0]
            for fname, info in group:
                pc = info.get('total_pages', 0)
                if pc != most_common_pages and pc != 0:
                    issues_format.append(
                        f"文件 \"{fname}\"（{wt}类）报告页数为 {pc} 页，"
                        f"而同类报告多数为 {most_common_pages} 页")

        # Compare product standards
        standards = Counter(info.get('product_standard', '未知') for _, info in group)
        if len(standards) > 1:
            most_common_std = standards.most_common(1)[0][0]
            for fname, info in group:
                std = info.get('product_standard', '未知')
                if std != most_common_std and std != '未知':
                    issues_format.append(
                        f"文件 \"{fname}\"（{wt}类）产品标准为 \"{std}\"，"
                        f"而同类报告多数为 \"{most_common_std}\"")

        # Compare test item counts
        item_counts = [info.get('test_item_count', 0) for _, info in group if info.get('test_item_count')]
        if item_counts:
            common_count = Counter(item_counts).most_common(1)[0][0]
            for fname, info in group:
                ic = info.get('test_item_count', 0)
                if ic and ic != common_count:
                    issues_format.append(
                        f"文件 \"{fname}\"（{wt}类）检测项目数为 {ic} 项，"
                        f"而同类报告多数为 {common_count} 项")

    # Check if 管网水 xls files follow a different template from xlsx ones
    guanwang_xls = [(f, i) for f, i in all_info.items()
                    if i['water_type'] == '管网水' and i['extension'] == '.xls']
    guanwang_xlsx = [(f, i) for f, i in all_info.items()
                     if i['water_type'] == '管网水' and i['extension'] == '.xlsx']
    if guanwang_xls and guanwang_xlsx:
        xls_pages = Counter(i.get('total_pages', 0) for _, i in guanwang_xls)
        xlsx_pages = Counter(i.get('total_pages', 0) for _, i in guanwang_xlsx)
        issues_format.append(
            f"管网水报告中，.xls 文件共 {len(guanwang_xls)} 个（页数分布：{dict(xls_pages)}），"
            f".xlsx 文件共 {len(guanwang_xlsx)} 个（页数分布：{dict(xlsx_pages)}），请确认是否使用不同模板")

    # Check sampler consistency within type groups
    for wt, group in type_groups.items():
        samplers = defaultdict(list)
        for fname, info in group:
            s = info.get('sampler', '未知')
            samplers[s].append(fname)
        if len(samplers) > 1 and '未知' not in samplers:
            # This is informational, not necessarily an issue
            pass

    # ──────────── 五、日期问题 ────────────
    for fname, info in all_info.items():
        # Check sampling date format
        sd = info.get('sampling_date', '')
        if sd:
            # Expected format: 2026.01.05
            if not re.match(r'^20\d{2}\.\d{2}\.\d{2}$', sd):
                issues_date.append(f"文件 \"{fname}\" 采样日期格式异常：'{sd}'")

        # Check receipt date vs sampling date
        rd = info.get('receipt_date', '')
        if sd and rd and sd != rd:
            # Receipt should be >= sampling
            try:
                sd_parsed = datetime.strptime(sd, '%Y.%m.%d')
                rd_parsed = datetime.strptime(rd, '%Y.%m.%d')
                if rd_parsed < sd_parsed:
                    issues_date.append(
                        f"文件 \"{fname}\" 收样日期 ({rd}) 早于采样日期 ({sd})")
                elif (rd_parsed - sd_parsed).days > 7:
                    issues_date.append(
                        f"文件 \"{fname}\" 收样日期 ({rd}) 与采样日期 ({sd}) 间隔超过 7 天")
            except ValueError:
                pass

        # Check testing date
        td = info.get('testing_date', '')
        if td and sd:
            # Testing date format: 2026.01.05~01.16
            m = re.match(r'(20\d{2}\.\d{2}\.\d{2})~(\d{2}\.\d{2})', td)
            if m:
                try:
                    td_start = datetime.strptime(m.group(1), '%Y.%m.%d')
                    sd_parsed = datetime.strptime(sd, '%Y.%m.%d')
                    if td_start < sd_parsed:
                        issues_date.append(
                            f"文件 \"{fname}\" 检测开始日期 ({m.group(1)}) 早于采样日期 ({sd})")
                except ValueError:
                    pass

        # Check report date format
        rpt_date = info.get('report_date', '')
        if rpt_date:
            # Various formats: "2026年 01月23日", "2026 年1 月 26日", "2026年 2月6日"
            m = re.search(r'(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日', rpt_date)
            if m:
                try:
                    rpt_parsed = datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
                    if sd:
                        sd_parsed = datetime.strptime(sd, '%Y.%m.%d')
                        if rpt_parsed < sd_parsed:
                            issues_date.append(
                                f"文件 \"{fname}\" 报告编制日期 ({rpt_date.strip()}) 早于采样日期 ({sd})")
                    # Check if report date is in a reasonable range (2026)
                    if rpt_parsed.year != 2026:
                        issues_date.append(
                            f"文件 \"{fname}\" 报告编制日期年份为 {rpt_parsed.year}，非 2026 年")
                except ValueError:
                    issues_date.append(f"文件 \"{fname}\" 报告编制日期无法解析：'{rpt_date}'")
            elif rpt_date.strip():
                issues_date.append(f"文件 \"{fname}\" 报告编制日期格式异常：'{rpt_date}'")

    # ──────────── 六、一致性问题 ────────────
    # Group files by plant name and check consistency
    plant_groups = defaultdict(list)
    for fname, info in all_info.items():
        plant = info.get('plant_name', '')
        if plant:
            plant_groups[plant].append((fname, info))

    for plant, group in plant_groups.items():
        if len(group) < 2:
            continue

        # Check if sample_name references are consistent
        companies = set()
        for fname, info in group:
            c = info.get('company', '')
            if c:
                companies.add(c)
        if len(companies) > 1:
            issues_consistency.append(
                f"水厂 \"{plant}\" 的相关报告中被检单位名称不一致：{', '.join(companies)}，"
                f"涉及文件：{', '.join(f for f, _ in group)}")

    # Check for similar plant names that might be the same plant (typos)
    plant_names = list(plant_groups.keys())
    for i in range(len(plant_names)):
        for j in range(i + 1, len(plant_names)):
            a, b = plant_names[i], plant_names[j]
            # Check if one is substring of the other or differ by just "水厂"
            if a in b or b in a:
                if a != b and abs(len(a) - len(b)) <= 2:
                    issues_consistency.append(
                        f"水厂名称疑似重复/不一致：\"{a}\" 与 \"{b}\"，请确认是否为同一水厂")

    # Check sample_name vs filename consistency
    for fname, info in all_info.items():
        sn = info.get('sample_name', '')
        if sn and '【' in sn and '】' in sn:
            m = re.search(r'【(.+?)】', sn)
            if m:
                sample_plant = m.group(1)
                # Remove trailing info like /地表水
                sample_plant = sample_plant.split('/')[0].strip()
                file_plant = info.get('plant_name', '')
                # Simple check: the sample plant name should appear in filename
                name_no_ext = os.path.splitext(fname)[0]
                prefix_removed = re.sub(r'^\d+', '', name_no_ext)
                if sample_plant not in prefix_removed and file_plant not in sample_plant:
                    # More lenient check
                    if sample_plant.replace('水厂', '') not in prefix_removed:
                        issues_consistency.append(
                            f"文件 \"{fname}\" 内样品名称为 \"{sn}\"，"
                            f"与文件名中的水厂名称不一致")

    # Check sample_type vs filename water type
    for fname, info in all_info.items():
        st = info.get('sample_type', '')
        wt = info['water_type']
        if st and wt != '未知':
            if wt == '出厂水' and '出厂水' not in st:
                issues_consistency.append(
                    f"文件 \"{fname}\" 文件名标注为出厂水，但内容样品类型为 \"{st}\"")
            elif wt == '原水' and '原水' not in st:
                issues_consistency.append(
                    f"文件 \"{fname}\" 文件名标注为原水，但内容样品类型为 \"{st}\"")
            elif wt == '管网水' and '管网' not in st:
                issues_consistency.append(
                    f"文件 \"{fname}\" 文件名标注为管网水，但内容样品类型为 \"{st}\"")
            elif wt == '二次供水' and '二次供水' not in st:
                issues_consistency.append(
                    f"文件 \"{fname}\" 文件名标注为二次供水，但内容样品类型为 \"{st}\"")

    # Check for plants that have 出厂水 but no 原水 or vice versa
    plant_types = defaultdict(set)
    for fname, info in all_info.items():
        plant = info.get('plant_name', '')
        wt = info['water_type']
        if plant and wt in ('出厂水', '原水', '管网水'):
            plant_types[plant].add(wt)
    for plant, types in sorted(plant_types.items()):
        if '出厂水' in types and '原水' not in types:
            issues_consistency.append(
                f"水厂 \"{plant}\" 有出厂水报告但缺少原水报告")
        if '原水' in types and '出厂水' not in types:
            # Could be just a water source, not necessarily an issue
            # Only flag if it looks like a water plant (not a reservoir)
            if '水厂' in plant:
                issues_consistency.append(
                    f"水厂 \"{plant}\" 有原水报告但缺少出厂水报告")

    # ──────────── 七、异常值问题 ────────────
    issues_values_critical = []  # Serious exceedances (>2x standard or toxic indicators)
    issues_values_normal = []    # Normal exceedances
    for fname, info in all_info.items():
        test_items = info.get('test_items', [])
        for item in test_items:
            result = item['result']
            name = item['name']
            standard = item['standard']

            if not result or result == 'None':
                continue

            # Skip water temperature -- the standard describes temperature CHANGE limits,
            # not absolute temperature. Absolute values like 14C are completely normal.
            if '水温' in name:
                continue

            # Try to parse numeric results
            numeric_result = None
            if result.startswith('<') or result.startswith('＜'):
                # Below detection limit - generally OK
                continue
            try:
                numeric_result = float(result)
            except (ValueError, TypeError):
                if result in ('无', '未检出', '无异臭、异味', '0', '0.0'):
                    continue
                continue

            if numeric_result is not None:
                exceeded = False
                std_limit_val = None

                if standard:
                    # Pattern: "≤X(II类)" or "≤X"
                    std_match = re.search(r'[≤<]\s*([\d.]+)', standard)
                    if std_match:
                        try:
                            std_limit_val = float(std_match.group(1))
                            if numeric_result > std_limit_val:
                                exceeded = True
                        except ValueError:
                            pass
                    # Pattern: just a number as limit
                    elif re.match(r'^[\d.]+$', str(standard)):
                        try:
                            std_limit_val = float(standard)
                            if numeric_result > std_limit_val:
                                exceeded = True
                        except ValueError:
                            pass
                    # Pattern: range like "0.1~0.8" or "0.02-0.8"
                    range_match = re.match(r'([\d.]+)\s*[~\-～]\s*([\d.]+)', standard)
                    if range_match:
                        try:
                            lo = float(range_match.group(1))
                            hi = float(range_match.group(2))
                            if numeric_result < lo or numeric_result > hi:
                                exceeded = True
                                std_limit_val = hi
                        except ValueError:
                            pass

                if exceeded:
                    # Determine severity
                    ratio_str = ""
                    is_critical = False
                    if std_limit_val and std_limit_val > 0:
                        ratio = numeric_result / std_limit_val
                        ratio_str = f"（为标准限值的 {ratio:.1f} 倍）"
                        if ratio >= 2.0:
                            is_critical = True
                    # Toxic heavy metals are always critical
                    if name in ('铅', '汞', '镉', '砷', '铬(六价)') and exceeded:
                        is_critical = True

                    msg = (f"文件 \"{fname}\" 检测项目 \"{name}\" "
                           f"结果 {result} 超出标准限值 {standard} {ratio_str}")
                    if is_critical:
                        issues_values_critical.append("[严重] " + msg)
                    else:
                        issues_values_normal.append(msg)

                # Check for suspicious values
                if name == 'pH' and numeric_result is not None:
                    if numeric_result < 5 or numeric_result > 10:
                        issues_values_critical.append(
                            f"[严重] 文件 \"{fname}\" pH 值 {result} 异常（通常范围 6-9）")

                # Negative values
                if numeric_result < 0:
                    issues_values_critical.append(
                        f"[严重] 文件 \"{fname}\" 检测项目 \"{name}\" 结果为负值 {result}")

    issues_values = issues_values_critical + issues_values_normal

    # ──────────── 八、文件读取问题 ────────────
    for fname, info in all_info.items():
        if 'read_error' in info:
            issues_read_errors.append(f"文件 \"{fname}\" 读取异常：{info['read_error']}")

    # ════════════════════════════════════════════════════
    # Output
    # ════════════════════════════════════════════════════
    lines = []
    lines.append("=" * 72)
    lines.append("        水质检测报告 —— 待确认问题清单")
    lines.append("=" * 72)
    lines.append(f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"扫描目录：{REPORT_DIR}")
    lines.append(f"扫描文件总数：{len(files)}")
    lines.append("")

    total_issues = (len(issues_naming) + len(issues_numbering) + len(issues_data)
                    + len(issues_format) + len(issues_date) + len(issues_consistency)
                    + len(issues_values) + len(issues_read_errors))
    lines.append(f"共发现 {total_issues} 项待确认问题，分类如下：")
    lines.append(f"  一、命名问题：{len(issues_naming)} 项")
    lines.append(f"  二、编号问题：{len(issues_numbering)} 项")
    lines.append(f"  三、数据问题：{len(issues_data)} 项")
    lines.append(f"  四、格式/模板问题：{len(issues_format)} 项")
    lines.append(f"  五、日期问题：{len(issues_date)} 项")
    lines.append(f"  六、一致性问题：{len(issues_consistency)} 项")
    lines.append(f"  七、异常值问题：{len(issues_values)} 项（其中严重 {len(issues_values_critical)} 项，一般 {len(issues_values_normal)} 项）")
    lines.append(f"  八、文件读取问题：{len(issues_read_errors)} 项")
    lines.append("")

    global_counter = 0

    def write_section(title, issues):
        nonlocal global_counter
        lines.append("-" * 72)
        lines.append(f"{title}（共 {len(issues)} 项）")
        lines.append("-" * 72)
        if not issues:
            lines.append("  （无）")
        for issue in issues:
            global_counter += 1
            lines.append(f"  {global_counter}. {issue}")
        lines.append("")

    write_section("一、命名问题", issues_naming)
    write_section("二、编号问题", issues_numbering)
    write_section("三、数据问题", issues_data)
    write_section("四、格式/模板问题", issues_format)
    write_section("五、日期问题", issues_date)
    write_section("六、一致性问题", issues_consistency)
    write_section("七、异常值问题", issues_values)
    write_section("八、文件读取问题", issues_read_errors)

    lines.append("=" * 72)
    lines.append("以上问题均为程序自动检测，可能存在误报，请人工逐项核实。")
    lines.append("=" * 72)

    output_text = '\n'.join(lines)

    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write(output_text)

    print(f"\n分析完成！共发现 {total_issues} 项待确认问题。")
    print(f"结果已写入：{OUTPUT_FILE}")

    return output_text


if __name__ == '__main__':
    main()
