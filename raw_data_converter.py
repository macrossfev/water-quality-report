"""
通用原始检测数据转换器

将实验室多 Sheet 原始检测 Excel 转换为系统可导入的标准转置格式。
自动识别：
  - Sheet1 样品登记表 → 提取样品编号、采样地点等元信息
  - 数据 Sheet → 自动判断纵向/横向布局，处理合并单元格
  - 参数名清洗 → 修正错别字、规范单位格式、去除换行符

输出格式（与 RawDataImporter 兼容）：
  Row 1:  A1="样品编号 →", B1=样品1, C1=样品2, ...
  Col A:  报告编号, 被检单位, 被检水厂, 样品类型, 采样日期, 检测指标...
  数据区: 各交叉点的值
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import re
import os
from datetime import datetime


# ── 样品编号识别 ──────────────────────────────────────────────────────────

# 常见的样品编号模式：W260105C01, K260105C01, S260105C01 等
SAMPLE_ID_RE = re.compile(r'^[A-Z]\d{5,6}C\d{1,3}$')


def is_sample_id(val):
    """判断一个单元格值是否为样品编号"""
    if val is None:
        return False
    s = str(val).strip()
    return bool(SAMPLE_ID_RE.match(s))


# ── 参数名清洗 ────────────────────────────────────────────────────────────

# 已知错别字映射
TYPO_MAP = {
    '肉哏可见物': '肉眼可见物',
    'CaCO₂': 'CaCO₃',
    'CaCO2': 'CaCO3',
    '阴离子合成洗涤剂阴离子表面活性剂': '阴离子合成洗涤剂',
}


def clean_param_name(name):
    """
    规范化检测指标名称：
    - 去除换行符和多余空格
    - 修正已知错别字
    - 统一单位括号格式
    - 去除末尾的 V（仪器标记）
    """
    if not name:
        return ''
    name = str(name)
    # 去除换行符
    name = name.replace('\n', '').replace('\r', '')
    # 压缩多余空格
    name = re.sub(r'\s{2,}', '', name)
    name = name.strip()

    # 修正错别字
    for wrong, right in TYPO_MAP.items():
        name = name.replace(wrong, right)

    # 去除末尾仪器标记 V
    name = re.sub(r'V$', '', name)

    # 修复单位在名称前面的异常格式:
    # "(CFU/100mL大肠埃希氏菌)" → "大肠埃希氏菌(CFU/100mL)"
    m = re.match(r'^\(([^)]*?)([\u4e00-\u9fff][\u4e00-\u9fff\w]*)\s*\)$', name)
    if m:
        unit_part = m.group(1).strip()
        name_part = m.group(2).strip()
        name = f'{name_part}({unit_part})'
        return name

    # 统一括号前无空格: "xxx (mg/L)" -> "xxx(mg/L)"
    name = re.sub(r'\s+\(', '(', name)
    # 修复缺失的右括号
    if '(' in name and ')' not in name:
        name += ')'

    return name


# ── 样品类型推断 ──────────────────────────────────────────────────────────

def infer_sample_type(location_text):
    """根据采样地点描述推断样品类型"""
    if not location_text:
        return ''
    text = str(location_text)
    if '出厂水' in text:
        return '出厂水'
    if '管网水' in text or '管网末梢' in text:
        return '管网水'
    if '原水' in text or '水源' in text or '水库' in text:
        return '原水'
    if '二次供水' in text:
        return '二次供水'
    if '空白' in text:
        return '空白样'
    return ''


def infer_company_and_plant(location_text):
    """
    从采样地点文本中推断被检单位和被检水厂。
    例: "王家坪水厂出厂水" → ("王家坪水厂", "王家坪水厂")
         "夔州水厂管网水"   → ("夔州水厂", "夔州水厂")
         "王家坪水厂原水黄井水库" → ("王家坪水厂", "王家坪水厂")
    """
    if not location_text:
        return '', ''
    text = str(location_text).strip()
    # 尝试提取 "XX水厂" 部分
    m = re.search(r'(.+?水厂)', text)
    if m:
        plant = m.group(1)
        return plant, plant
    return text, text


# ── 合并单元格解析 ────────────────────────────────────────────────────────

def build_merged_cell_map(ws):
    """
    构建合并单元格映射表。
    返回 dict: {(row, col): 合并区域左上角的值}
    """
    merged_map = {}
    for merge_range in ws.merged_cells.ranges:
        top_left_value = ws.cell(merge_range.min_row, merge_range.min_col).value
        for row in range(merge_range.min_row, merge_range.max_row + 1):
            for col in range(merge_range.min_col, merge_range.max_col + 1):
                merged_map[(row, col)] = top_left_value
    return merged_map


def get_cell_value(ws, row, col, merged_map):
    """获取单元格值，优先使用合并单元格映射"""
    val = ws.cell(row, col).value
    if val is not None:
        return val
    return merged_map.get((row, col))


# ── Sheet1 解析（样品登记表） ─────────────────────────────────────────────

def parse_registration_sheet(ws, merged_map):
    """
    解析样品登记表 (Sheet1)，提取样品元信息。
    自动检测包含样品编号的列。

    返回: [
        {
            '样品编号': 'W260105C01',
            '被检单位': '王家坪水厂',
            '被检水厂': '王家坪水厂',
            '样品类型': '出厂水',
            '采样地点': '王家坪水厂出厂水',
        },
        ...
    ]
    """
    samples = []

    # 扫描前几行，找到包含样品编号的列
    sample_id_col = None
    location_col = None
    header_row = None

    for r in range(1, min(ws.max_row + 1, 10)):
        for c in range(1, ws.max_column + 1):
            val = get_cell_value(ws, r, c, merged_map)
            if val is None:
                continue
            val_str = str(val).strip()
            if val_str == '样品编号' or (val_str.endswith('样品编号') and '采样' not in val_str):
                sample_id_col = c
                header_row = r
            elif '采样地点' in val_str or ('样品类型' in val_str and '样品编号' not in val_str):
                location_col = c

    if sample_id_col is None:
        # 备选：扫描所有单元格找样品编号模式
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                val = get_cell_value(ws, r, c, merged_map)
                if is_sample_id(val):
                    sample_id_col = c
                    header_row = r - 1 if r > 1 else r
                    break
            if sample_id_col:
                break

    if sample_id_col is None:
        return []

    # 在 location_col 未找到时，猜测为 sample_id_col 前一列或前两列
    if location_col is None:
        for candidate in [sample_id_col - 2, sample_id_col - 1]:
            if candidate >= 1:
                location_col = candidate
                break

    start_row = (header_row + 1) if header_row else 2

    for r in range(start_row, ws.max_row + 1):
        sid_val = get_cell_value(ws, r, sample_id_col, merged_map)
        if not is_sample_id(sid_val):
            continue

        sid = str(sid_val).strip()
        location = ''
        if location_col:
            loc_val = get_cell_value(ws, r, location_col, merged_map)
            if loc_val:
                location = str(loc_val).strip()

        company, plant = infer_company_and_plant(location)
        sample_type = infer_sample_type(location)

        samples.append({
            '样品编号': sid,
            '被检单位': company,
            '被检水厂': plant,
            '样品类型': sample_type,
            '采样地点': location,
        })

    return samples


# ── 数据 Sheet 解析 ──────────────────────────────────────────────────────

def find_sample_ids_in_row(ws, row, merged_map, known_ids):
    """在指定行中查找样品编号，返回 {col: sample_id}"""
    col_map = {}
    for c in range(1, ws.max_column + 1):
        val = get_cell_value(ws, row, c, merged_map)
        if val is None:
            continue
        s = str(val).strip()
        if s in known_ids or is_sample_id(s):
            col_map[c] = s
    return col_map


def find_sample_ids_in_col(ws, col, merged_map, known_ids):
    """在指定列中查找样品编号，返回 {row: sample_id}"""
    row_map = {}
    for r in range(1, ws.max_row + 1):
        val = get_cell_value(ws, r, col, merged_map)
        if val is None:
            continue
        s = str(val).strip()
        if s in known_ids or is_sample_id(s):
            row_map[r] = s
    return row_map


def detect_and_parse_data_sheet(ws, merged_map, known_ids):
    """
    自动检测数据 Sheet 的布局并提取数据。

    策略：
    1. 扫描前 5 行，查找包含最多样品编号的行 → 纵向布局（样品在列头）
    2. 扫描前 3 列，查找包含最多样品编号的列 → 横向布局（样品在行头）
    3. 选择匹配数量最多的方案

    返回: {sample_id: {param_name: value}}
    """
    data = {}

    # 尝试在前5行找样品编号（纵向布局）
    best_row = None
    best_row_map = {}
    for r in range(1, min(ws.max_row + 1, 6)):
        col_map = find_sample_ids_in_row(ws, r, merged_map, known_ids)
        if len(col_map) > len(best_row_map):
            best_row = r
            best_row_map = col_map

    # 尝试在前3列找样品编号（横向布局）
    best_col = None
    best_col_map = {}
    for c in range(1, min(ws.max_column + 1, 4)):
        row_map = find_sample_ids_in_col(ws, c, merged_map, known_ids)
        if len(row_map) > len(best_col_map):
            best_col = c
            best_col_map = row_map

    if len(best_row_map) >= len(best_col_map) and len(best_row_map) > 0:
        # 纵向布局：样品编号在行头，参数在列
        data = _parse_vertical(ws, merged_map, best_row, best_row_map)
    elif len(best_col_map) > 0:
        # 横向布局：样品编号在列头，参数在行头
        data = _parse_horizontal(ws, merged_map, best_col, best_col_map)

    return data


def _parse_vertical(ws, merged_map, header_row, col_map):
    """
    纵向布局：参数名在第一列区域，样品编号在 header_row 的各列。
    逐行读取参数名，逐列读取对应的值。
    """
    data = {}
    for sid in col_map.values():
        if sid not in data:
            data[sid] = {}

    # 确定参数名所在的列（通常是第 1 列，可能跨合并单元格）
    param_col = 1

    for r in range(header_row + 1, ws.max_row + 1):
        # 尝试从第 1 列或合并区域获取参数名
        param_raw = get_cell_value(ws, r, param_col, merged_map)
        if param_raw is None:
            # 尝试第 2、3 列（部分 Sheet 参数名在 B 或 C 列）
            for alt_c in range(2, min(ws.max_column + 1, 4)):
                param_raw = get_cell_value(ws, r, alt_c, merged_map)
                if param_raw is not None:
                    break

        if param_raw is None:
            continue

        param = clean_param_name(param_raw)
        if not param or is_sample_id(param):
            continue

        # 跳过明显是页眉/页脚的行
        if any(kw in param for kw in ['汇总表', '分析结果', '检测结果', '第', '页']):
            continue

        for c, sid in col_map.items():
            val = get_cell_value(ws, r, c, merged_map)
            if val is not None:
                val_str = str(val).strip()
                if val_str and not is_sample_id(val_str):
                    data[sid][param] = val_str

    return data


def _parse_horizontal(ws, merged_map, sid_col, row_map):
    """
    横向布局：样品编号在某列，参数名在表头行。
    """
    data = {}
    for sid in row_map.values():
        if sid not in data:
            data[sid] = {}

    # 找表头行（样品编号所在列的第一个非样品编号行，通常是 row_map 中最小行 - 1）
    min_data_row = min(row_map.keys())
    header_row = max(1, min_data_row - 1)

    # 构建列→参数名映射
    param_map = {}
    for c in range(1, ws.max_column + 1):
        if c == sid_col:
            continue
        val = get_cell_value(ws, header_row, c, merged_map)
        if val is not None:
            param = clean_param_name(val)
            if param and not is_sample_id(param):
                param_map[c] = param

    for r, sid in row_map.items():
        for c, param in param_map.items():
            val = get_cell_value(ws, r, c, merged_map)
            if val is not None:
                val_str = str(val).strip()
                if val_str:
                    data[sid][param] = val_str

    return data


# ── 采样日期推断 ──────────────────────────────────────────────────────────

def infer_sampling_date(sample_id):
    """
    从样品编号推断采样日期。
    W260105C01 → 2026-01-05
    """
    m = re.match(r'^[A-Z](\d{6})C', sample_id)
    if m:
        date_part = m.group(1)
        try:
            # 格式: YYMMDD
            dt = datetime.strptime(date_part, '%y%m%d')
            return dt.strftime('%Y-%m-%d')
        except ValueError:
            pass
    return ''


# ── 主转换函数 ────────────────────────────────────────────────────────────

def convert_raw_excel(file_path, output_path=None, skip_blank_samples=True):
    """
    将实验室原始检测 Excel 转换为系统可导入的标准格式。

    参数:
        file_path: 源 Excel 路径
        output_path: 输出 Excel 路径（None 则自动生成）
        skip_blank_samples: 是否跳过空白样（K 开头）

    返回: {
        'success': bool,
        'message': str,
        'output_path': str,
        'samples': list,       # 样品元信息列表
        'parameters': list,    # 检测指标列表
        'data': dict,          # {sample_id: {param: value}}
        'sample_count': int,
        'param_count': int,
    }
    """
    if not os.path.exists(file_path):
        return {'success': False, 'message': f'文件不存在: {file_path}'}

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        return {'success': False, 'message': f'无法打开 Excel 文件: {e}'}

    sheet_names = wb.sheetnames
    if not sheet_names:
        wb.close()
        return {'success': False, 'message': 'Excel 文件无工作表'}

    # ── Step 1: 解析 Sheet1（样品登记表） ─────────────────────────────

    ws1 = wb[sheet_names[0]]
    merged_map1 = build_merged_cell_map(ws1)
    samples = parse_registration_sheet(ws1, merged_map1)

    if not samples:
        # 如果 Sheet1 解析失败，扫描所有 Sheet 收集样品编号
        all_ids = set()
        for sn in sheet_names:
            ws = wb[sn]
            mm = build_merged_cell_map(ws)
            for r in range(1, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    val = get_cell_value(ws, r, c, mm)
                    if is_sample_id(val):
                        all_ids.add(str(val).strip())
        samples = [{'样品编号': sid, '被检单位': '', '被检水厂': '',
                     '样品类型': '', '采样地点': ''} for sid in sorted(all_ids)]

    if not samples:
        wb.close()
        return {'success': False, 'message': '未找到任何样品编号'}

    # 可选：跳过空白样（K 开头）
    if skip_blank_samples:
        samples = [s for s in samples if not s['样品编号'].startswith('K')]

    known_ids = set(s['样品编号'] for s in samples)

    # ── Step 2: 解析所有数据 Sheet ────────────────────────────────────

    all_data = {s['样品编号']: {} for s in samples}

    for sn in sheet_names[1:]:  # 跳过 Sheet1（登记表）
        ws = wb[sn]
        merged_map = build_merged_cell_map(ws)
        sheet_data = detect_and_parse_data_sheet(ws, merged_map, known_ids)
        # 合并数据（后面的 Sheet 不覆盖已有值）
        for sid, params in sheet_data.items():
            if sid not in all_data:
                continue
            for param, value in params.items():
                if param not in all_data[sid]:
                    all_data[sid][param] = value

    wb.close()

    # ── Step 3: 构建有序参数列表 ──────────────────────────────────────

    all_params = []
    seen = set()
    for sid in (s['样品编号'] for s in samples):
        for p in all_data.get(sid, {}):
            if p not in seen:
                all_params.append(p)
                seen.add(p)

    if not all_params:
        return {'success': False, 'message': '未提取到任何检测数据'}

    # ── Step 4: 推断缺失的元信息 ─────────────────────────────────────

    for s in samples:
        if not s.get('采样日期'):
            s['采样日期'] = infer_sampling_date(s['样品编号'])

    # ── Step 5: 生成输出 Excel ────────────────────────────────────────

    if output_path is None:
        base = os.path.splitext(os.path.basename(file_path))[0]
        output_dir = os.path.join(os.path.dirname(file_path) or '.', 'converted')
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, f'{base}_import.xlsx')

    _write_import_excel(output_path, samples, all_params, all_data)

    return {
        'success': True,
        'message': f'转换成功: {len(samples)} 个样品, {len(all_params)} 项指标',
        'output_path': output_path,
        'samples': samples,
        'parameters': all_params,
        'data': all_data,
        'sample_count': len(samples),
        'param_count': len(all_params),
    }


def _write_import_excel(output_path, samples, params, data):
    """写入标准导入格式的 Excel"""
    out_wb = openpyxl.Workbook()
    ws = out_wb.active
    ws.title = '数据导入'

    # 样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    required_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    required_font = Font(bold=True, size=10, color="C00000")
    label_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    label_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    center = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    BASE_FIELDS = ['报告编号', '被检单位', '被检水厂', '样品类型', '采样日期']

    # Row 1: 样品编号表头
    cell = ws.cell(1, 1, '样品编号 →')
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border

    for i, s in enumerate(samples):
        cell = ws.cell(1, i + 2, s['样品编号'])
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # 行字段
    all_fields = BASE_FIELDS + params
    for row_off, field in enumerate(all_fields):
        row_idx = row_off + 2

        # A 列：字段名
        label_cell = ws.cell(row_idx, 1, field)
        if field in BASE_FIELDS:
            label_cell.fill = required_fill
            label_cell.font = required_font
        else:
            label_cell.fill = label_fill
            label_cell.font = label_font
        label_cell.alignment = left_align
        label_cell.border = border

        # 数据列
        for col_off, s in enumerate(samples):
            col_idx = col_off + 2
            sid = s['样品编号']

            if field in BASE_FIELDS:
                val = s.get(field, '')
            else:
                val = data.get(sid, {}).get(field)

            cell = ws.cell(row_idx, col_idx)
            if val is not None:
                cell.value = val
            cell.alignment = center
            cell.border = border

    # 列宽
    ws.column_dimensions['A'].width = 28
    for col_idx in range(2, len(samples) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    ws.freeze_panes = 'B2'

    out_wb.save(output_path)
    out_wb.close()


# ── 预览（不生成文件，仅返回解析结果） ────────────────────────────────────

def preview_raw_excel(file_path, skip_blank_samples=True):
    """
    预览原始 Excel，返回解析结果但不生成文件。
    用于前端预览确认。
    """
    result = convert_raw_excel(file_path, output_path=None, skip_blank_samples=skip_blank_samples)
    # 删除临时生成的文件（preview 不需要保留）
    if result.get('success') and result.get('output_path'):
        try:
            os.remove(result['output_path'])
        except OSError:
            pass
        result.pop('output_path', None)
    return result


# ── CLI 测试入口 ──────────────────────────────────────────────────────────

if __name__ == '__main__':
    import sys
    import json

    if len(sys.argv) < 2:
        print("用法: python raw_data_converter.py <input.xlsx> [output.xlsx]")
        sys.exit(1)

    src = sys.argv[1]
    dst = sys.argv[2] if len(sys.argv) > 2 else None

    result = convert_raw_excel(src, dst)
    if result['success']:
        print(f"✓ {result['message']}")
        print(f"  输出: {result['output_path']}")
        print(f"  样品: {[s['样品编号'] for s in result['samples']]}")
        print(f"  指标 ({result['param_count']}):")
        for i, p in enumerate(result['parameters'], 1):
            print(f"    {i}. {p}")
    else:
        print(f"✗ {result['message']}")
        sys.exit(1)
