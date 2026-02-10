"""
原始数据导入模板生成器
根据已固化的字段配置生成Excel导入模板（转置布局）

生成格式：
  - 第一行：A1为空，B1起为样品编号占位
  - 第一列：A2起为字段名（报告编号、被检单位、...、检测指标...）
  - 数据区：各样品对应字段的值
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from models_v2 import get_db_connection
import os
from datetime import datetime


class RawDataTemplateGenerator:
    """原始数据导入模板生成器"""

    # 基础行字段名称（系统预定义，不含样品编号）
    BASE_ROW_FIELDS = ['报告编号', '被检单位', '被检水厂', '样品类型', '采样日期']

    def __init__(self):
        """初始化生成器"""
        self.fields = []
        self.has_schema = False

    def generate(self, output_path=None):
        """
        生成导入模板

        Args:
            output_path: 输出文件路径（可选）

        Returns:
            str: 生成的文件路径
        """
        # 加载字段配置
        self._load_fields()

        # 创建工作簿
        wb = openpyxl.Workbook()

        # 移除默认的Sheet
        wb.remove(wb.active)

        # 1. 创建说明sheet（放在第一个）
        self._create_instruction_sheet(wb)

        # 2. 创建数据导入sheet
        self._create_data_sheet(wb)

        # 3. 保存文件
        if output_path is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = f"exports/raw_data_import_template_{timestamp}.xlsx"

        os.makedirs('exports', exist_ok=True)
        wb.save(output_path)
        wb.close()

        return output_path

    def _load_fields(self):
        """加载字段配置"""
        conn = get_db_connection()
        cursor = conn.cursor()

        # 尝试获取已固化的字段配置
        cursor.execute('''
            SELECT column_name, column_order, data_type, is_base_field
            FROM raw_data_column_schema
            ORDER BY column_order
        ''')

        rows = cursor.fetchall()
        conn.close()

        if rows:
            # 使用已固化的字段（过滤掉样品编号，它在表头行）
            self.fields = [row[0] for row in rows if row[0] != '样品编号']
            self.has_schema = True
        else:
            # 使用默认字段（仅包含基础字段 + 示例检测指标）
            self.fields = self.BASE_ROW_FIELDS + [
                'pH', '浊度', '余氯', '色度', '臭和味',
                '肉眼可见物', '耗氧量', '总大肠菌群', '菌落总数'
            ]
            self.has_schema = False

    def _create_data_sheet(self, wb):
        """
        创建数据导入sheet（转置布局）
        格式：第一行为样品编号，第一列为字段名
        """
        ws = wb.create_sheet("数据导入")

        # 设置样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        required_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
        required_font = Font(bold=True, size=10, color="C00000")
        label_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        label_font = Font(bold=True, size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # A1单元格：标签
        cell_a1 = ws.cell(1, 1)
        cell_a1.value = "样品编号 →"
        cell_a1.fill = header_fill
        cell_a1.font = header_font
        cell_a1.alignment = Alignment(horizontal='center', vertical='center')
        cell_a1.border = border

        # 第一行（B1起）：示例样品编号
        example_samples = ['W260129C001', 'W260129C002']
        for col_offset, sample_id in enumerate(example_samples):
            col_idx = col_offset + 2  # B列=2, C列=3
            cell = ws.cell(1, col_idx)
            cell.value = sample_id
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # 添加更多空列供填写（D-K列，共8列）
        for col_idx in range(len(example_samples) + 2, len(example_samples) + 10):
            cell = ws.cell(1, col_idx)
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 第一列（A2起）：字段名；数据区：示例数据
        example_data_col1 = {
            '报告编号': 'BGH-2026-001',
            '被检单位': '某某水务公司',
            '被检水厂': '第一水厂',
            '样品类型': '出厂水',
            '采样日期': '2026-01-29',
            'pH': '7.2',
            '浊度': '0.5',
            '余氯': '0.3',
            '色度': '<5',
            '臭和味': '无',
            '肉眼可见物': '无',
            '耗氧量': '1.2',
            '总大肠菌群': '未检出',
            '菌落总数': '<1'
        }
        example_data_col2 = {
            '报告编号': 'BGH-2026-002',
            '被检单位': '某某水务公司',
            '被检水厂': '第二水厂',
            '样品类型': '管网水',
            '采样日期': '2026-01-29',
            'pH': '7.0',
            '浊度': '0.3',
            '余氯': '0.5',
            '色度': '<5',
            '臭和味': '无',
            '肉眼可见物': '无',
            '耗氧量': '0.8',
            '总大肠菌群': '未检出',
            '菌落总数': '<1'
        }

        total_sample_cols = len(example_samples) + 8  # 示例 + 空白列

        for row_offset, field_name in enumerate(self.fields):
            row_idx = row_offset + 2  # 从第2行开始

            # A列：字段名标签
            label_cell = ws.cell(row_idx, 1)
            label_cell.value = field_name

            if field_name in self.BASE_ROW_FIELDS:
                label_cell.fill = required_fill
                label_cell.font = required_font
            else:
                label_cell.fill = label_fill
                label_cell.font = label_font

            label_cell.alignment = Alignment(horizontal='left', vertical='center')
            label_cell.border = border

            # B列：第一个示例样品的数据
            cell_b = ws.cell(row_idx, 2)
            cell_b.value = example_data_col1.get(field_name, '')
            cell_b.alignment = Alignment(horizontal='center', vertical='center')
            cell_b.border = border

            # C列：第二个示例样品的数据
            cell_c = ws.cell(row_idx, 3)
            cell_c.value = example_data_col2.get(field_name, '')
            cell_c.alignment = Alignment(horizontal='center', vertical='center')
            cell_c.border = border

            # D列起：空白数据区域
            for col_idx in range(4, total_sample_cols + 2):
                cell = ws.cell(row_idx, col_idx)
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # 调整列宽
        ws.column_dimensions['A'].width = 16
        for col_idx in range(2, total_sample_cols + 2):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 16

        # 冻结A列和第一行（滚动时始终可见字段名和样品编号）
        ws.freeze_panes = 'B2'

    def _create_instruction_sheet(self, wb):
        """创建填写说明sheet"""
        ws = wb.create_sheet("填写说明", 0)  # 插入到第一个位置

        # 标题
        ws['A1'] = "原始数据导入模板填写说明"
        ws['A1'].font = Font(size=16, bold=True, color="4472C4")

        row = 3

        # 基本说明
        status_text = "已固化" if self.has_schema else "默认示例"
        field_count = len(self.fields)

        instructions = [
            ("一、模板信息", ""),
            ("", f"字段配置状态: {status_text}"),
            ("", f"字段数量: {field_count} 个（不含样品编号）"),
            ("", f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"),
            ("", ""),
            ("二、布局说明", ""),
            ("", "本模板采用转置布局："),
            ("", "  • 第一行（B1起）：填写样品编号，每列一个样品"),
            ("", "  • 第一列（A2起）：字段名称（不可修改）"),
            ("", "  • 数据区：在对应的行列交叉处填写数据"),
            ("", ""),
            ("三、字段说明", ""),
            ("", f"当前模板包含以下字段（按顺序）："),
        ]

        # 添加字段列表
        for idx, field_name in enumerate(self.fields, start=1):
            is_required = "【必填】" if field_name in self.BASE_ROW_FIELDS else "【选填】"
            instructions.append(("", f"  {idx}. {field_name} {is_required}"))

        instructions.extend([
            ("", ""),
            ("四、填写说明", ""),
            ("", "1. 样品编号（第一行）："),
            ("", "   • 从B1单元格开始，每列填写一个样品编号"),
            ("", "   • 样品编号不能重复，必须唯一"),
            ("", ""),
            ("", "2. 必填字段说明："),
            ("", "   • 报告编号：检测报告编号（如：BGH-2026-001）"),
            ("", "   • 被检单位：样品所属的被检单位名称"),
            ("", "   • 被检水厂：样品所属的被检水厂名称"),
            ("", "   • 样品类型：如出厂水、管网水、原水等"),
            ("", "   • 采样日期：必须为YYYY-MM-DD格式（如：2026-01-29）"),
            ("", ""),
            ("", "3. 检测指标字段："),
            ("", "   • 直接填写检测结果（数值或文本）"),
            ("", "   • 可以是数字（如：7.2、0.5）"),
            ("", "   • 可以是文本（如：未检出、无、<5）"),
            ("", "   • 留空表示未检测该指标"),
            ("", ""),
            ("", "4. 格式要求："),
            ("", "   • 采样日期严格按照YYYY-MM-DD格式"),
            ("", "   • 不要修改第一列（A列）的字段名"),
            ("", "   • 不要修改字段的顺序"),
            ("", "   • 不要删除或增加字段行"),
        ])

        if self.has_schema:
            instructions.extend([
                ("", "   • 系统已固化字段配置，必须与此模板完全一致"),
            ])
        else:
            instructions.extend([
                ("", "   • 首次导入将固化字段配置"),
                ("", "   • 后续导入必须与首次导入的字段一致"),
            ])

        instructions.extend([
            ("", ""),
            ("", "5. 示例数据："),
            ("", "   • B列和C列提供了示例数据，可以参考填写"),
            ("", "   • 导入前请替换或删除示例数据"),
            ("", ""),
            ("五、重复数据处理", ""),
            ("", "导入时可以选择重复样品编号的处理方式："),
            ("", "• 跳过重复记录（推荐）：重复的样品编号会被跳过，不影响其他数据导入"),
            ("", "• 覆盖已有记录：重复的样品编号会覆盖数据库中的旧数据"),
            ("", "• 终止导入：遇到重复样品编号立即停止导入"),
            ("", ""),
            ("六、导入步骤", ""),
            ("", "1. 下载本模板"),
            ("", "2. 在\"数据导入\"sheet中填写数据"),
            ("", "3. 第一行从B1起填写样品编号"),
            ("", "4. 在对应行列交叉处填写各字段值"),
            ("", "5. 保存文件"),
            ("", "6. 在系统中点击\"选择Excel文件\"上传"),
            ("", "7. 选择重复处理方式"),
            ("", "8. 点击\"开始导入\"按钮"),
            ("", "9. 查看导入结果报告"),
            ("", ""),
            ("七、常见问题", ""),
            ("", "Q: 为什么导入失败提示\"字段不匹配\"？"),
            ("", "A: 系统已固化字段配置，必须使用系统生成的最新模板，不能修改字段名或顺序。"),
            ("", ""),
            ("", "Q: 采样日期应该如何填写？"),
            ("", "A: 必须严格按照YYYY-MM-DD格式，如2026-01-29，不能是其他格式。"),
            ("", ""),
            ("", "Q: 检测指标可以留空吗？"),
            ("", "A: 可以，留空表示该样品未检测此指标。"),
            ("", ""),
            ("", "Q: 可以添加自己的检测指标吗？"),
        ])

        if self.has_schema:
            instructions.extend([
                ("", "A: 不可以，系统已固化字段配置，不能增加或删除字段行。如需修改，请联系管理员。"),
            ])
        else:
            instructions.extend([
                ("", "A: 可以，首次导入时可以添加任意检测指标行，导入后字段将被固化。"),
            ])

        # 写入说明内容
        for title, content in instructions:
            if title:
                cell = ws.cell(row, 1)
                cell.value = title
                cell.font = Font(size=12, bold=True, color="4472C4")
            else:
                cell = ws.cell(row, 1)
                cell.value = content
                cell.alignment = Alignment(wrap_text=True)

            row += 1

        # 调整列宽
        ws.column_dimensions['A'].width = 100


def generate_raw_data_template(output_path=None):
    """
    生成原始数据导入模板的便捷函数

    Args:
        output_path: 输出路径（可选）

    Returns:
        str: 生成的文件路径
    """
    generator = RawDataTemplateGenerator()
    return generator.generate(output_path)


if __name__ == '__main__':
    print("="*60)
    print("原始数据导入模板生成器测试")
    print("="*60)

    # 生成模板
    print("\n生成原始数据导入模板...")
    path = generate_raw_data_template()
    print(f"生成成功: {path}")
