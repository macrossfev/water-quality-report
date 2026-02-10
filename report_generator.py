"""
报告生成器
按照Excel模版生成水质检测报告
"""
import os
import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime
from models_v2 import get_db_connection
import shutil

class ReportGenerator:
    """按模版生成Excel报告"""

    def __init__(self, template_id, report_data, report_id=None):
        """
        初始化报告生成器

        Args:
            template_id: 报告模版ID
            report_data: 报告数据字典，包含基本信息和检测数据
            report_id: 报告ID（用于从数据库加载完整数据）
        """
        self.template_id = template_id
        self.report_data = report_data
        self.report_id = report_id
        self.template_info = None
        self.workbook = None

    def generate(self, output_path=None, filename_template=None, export_format='xlsx'):
        """
        生成报告

        Args:
            output_path: 输出文件路径，如果为None则自动生成
            filename_template: 文件名模板，支持变量：{report_number}, {sampling_location}, {timestamp}
                              默认: "{sampling_location}{report_number}"
            export_format: 导出格式，支持 'xlsx' 或 'pdf'，默认 'xlsx'

        Returns:
            str: 生成的文件路径
        """
        # 1. 加载模版信息
        self._load_template_info()

        # 2. 从数据库加载完整数据（包括report_field_values）
        self._load_complete_data()

        # 3. 复制模版文件
        if output_path is None:
            # 生成文件名
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

            # 如果没有指定模板，使用默认模板
            if filename_template is None:
                filename_template = "{sampling_location}{report_number}"

            # 准备变量替换
            variables = {
                'report_number': self.report_data.get('report_number', 'report'),
                'sampling_location': self.report_data.get('sampling_location', ''),
                'timestamp': timestamp
            }

            # 替换变量
            filename = filename_template.format(**variables)

            # 清理文件名中的非法字符
            filename = self._sanitize_filename(filename)

            # 先生成Excel文件
            output_path = f"exports/{filename}.xlsx"

        os.makedirs('exports', exist_ok=True)
        shutil.copy2(self.template_info['template_file_path'], output_path)

        # 4. 打开工作簿
        self.workbook = openpyxl.load_workbook(output_path)

        # 5. 填充数据
        self._fill_data()

        # 6. 保存Excel文件
        self.workbook.save(output_path)
        self.workbook.close()

        # 7. 如果需要PDF格式，进行转换
        if export_format.lower() == 'pdf':
            pdf_path = self._convert_to_pdf(output_path)
            if pdf_path:
                return pdf_path
            else:
                print("⚠ PDF转换失败，返回Excel文件")
                return output_path

        return output_path

    def _load_complete_data(self):
        """从数据库加载完整的报告数据（包括report_field_values）"""
        if not self.report_id:
            return

        conn = get_db_connection()

        # 1. 加载报告基本信息
        report = conn.execute('''
            SELECT r.*, st.name as sample_type_name, st.code as sample_type_code,
                   c.name as company_name
            FROM reports r
            LEFT JOIN sample_types st ON r.sample_type_id = st.id
            LEFT JOIN companies c ON r.company_id = c.id
            WHERE r.id = ?
        ''', (self.report_id,)).fetchone()

        if report:
            # 合并所有基本信息到report_data
            self.report_data['report_number'] = report['report_number'] or ''
            self.report_data['sample_number'] = report['sample_number'] or ''
            self.report_data['sample_type'] = report['sample_type_name'] or ''
            self.report_data['sample_type_name'] = report['sample_type_name'] or ''
            self.report_data['company_name'] = report['company_name'] or ''
            self.report_data['detection_date'] = report['detection_date'] or ''
            self.report_data['detection_person'] = report['detection_person'] or ''
            self.report_data['review_person'] = report['review_person'] or ''
            self.report_data['remark'] = report['remark'] or ''

            # 添加更多字段
            self.report_data['sampling_date'] = report['sampling_date'] or ''
            self.report_data['sampler'] = report['sampler'] or ''
            self.report_data['sampling_location'] = report['sampling_location'] or ''
            self.report_data['sampling_basis'] = report['sampling_basis'] or ''
            self.report_data['sample_source'] = report['sample_source'] or ''
            self.report_data['sample_status'] = report['sample_status'] or ''
            self.report_data['sample_received_date'] = report['sample_received_date'] or ''
            self.report_data['report_date'] = report['report_date'] or ''
            self.report_data['product_standard'] = report['product_standard'] or ''
            self.report_data['test_conclusion'] = report['test_conclusion'] or ''
            self.report_data['additional_info'] = report['additional_info'] or ''

            # 添加检测项目描述和附件信息
            detection_items_desc = report['detection_items_description'] if 'detection_items_description' in report.keys() else None
            attachment_info_val = report['attachment_info'] if 'attachment_info' in report.keys() else None
            # 确保 None 转换为空字符串，而不是字符串 'None'
            self.report_data['detection_items_description'] = detection_items_desc if detection_items_desc is not None else ''
            self.report_data['attachment_info'] = attachment_info_val if attachment_info_val is not None else ''

            # 从remark中提取客户信息
            if report['remark']:
                try:
                    import json
                    remark_data = json.loads(report['remark'])
                    self.report_data['customer_unit'] = remark_data.get('customer_unit', '')
                    self.report_data['customer_plant'] = remark_data.get('customer_plant', '')
                    # 注意：remark JSON中的键名是 customer_address，需要映射到 unit_address
                    self.report_data['unit_address'] = remark_data.get('customer_address', '') or remark_data.get('unit_address', '')
                except:
                    pass

            print(f"已加载报告数据，字段数量: {len(self.report_data)}")
            print(f"报告数据键: {list(self.report_data.keys())}")

        # 2. 加载模板字段值（关键！之前缺失的部分）
        field_values = conn.execute('''
            SELECT rfv.field_value, tfm.field_name, tfm.field_display_name
            FROM report_field_values rfv
            JOIN template_field_mappings tfm ON rfv.field_mapping_id = tfm.id
            WHERE rfv.report_id = ?
        ''', (self.report_id,)).fetchall()

        for fv in field_values:
            # 使用field_name作为键
            field_key = fv['field_name']
            self.report_data[field_key] = fv['field_value']
            # 同时使用display_name作为备用键
            if fv['field_display_name']:
                self.report_data[fv['field_display_name']] = fv['field_value']

        # 3. 加载检测数据
        if 'detection_items' not in self.report_data or not self.report_data['detection_items']:
            detection_items = conn.execute('''
                SELECT rd.measured_value, i.name, i.unit, i.limit_value, i.detection_method
                FROM report_data rd
                JOIN indicators i ON rd.indicator_id = i.id
                LEFT JOIN indicator_groups g ON i.group_id = g.id
                JOIN reports r ON rd.report_id = r.id
                LEFT JOIN template_indicators ti
                    ON ti.indicator_id = rd.indicator_id AND ti.sample_type_id = r.sample_type_id
                WHERE rd.report_id = ?
                ORDER BY ti.sort_order, g.sort_order, i.sort_order, i.name
            ''', (self.report_id,)).fetchall()

            self.report_data['detection_items'] = [
                {
                    'name': item['name'],
                    'unit': item['unit'] or '',
                    'result': item['measured_value'] or '',
                    'limit': item['limit_value'] or '',
                    'method': item['detection_method'] or ''
                }
                for item in detection_items
            ]

        conn.close()

    def _load_template_info(self):
        """加载模版信息"""
        conn = get_db_connection()

        template = conn.execute(
            'SELECT * FROM excel_report_templates WHERE id = ?',
            (self.template_id,)
        ).fetchone()

        if not template:
            conn.close()
            raise ValueError(f"模版不存在: ID={self.template_id}")

        # 获取字段映射
        fields = conn.execute(
            'SELECT * FROM template_field_mappings WHERE template_id = ?',
            (self.template_id,)
        ).fetchall()

        conn.close()

        self.template_info = dict(template)
        self.template_info['fields'] = [dict(f) for f in fields]

    def _fill_data(self):
        """填充报告数据"""
        fields = self.template_info.get('fields', [])

        print(f"\n=== 开始填充数据 ===")
        print(f"模板字段数量: {len(fields)}")
        print(f"报告数据键: {list(self.report_data.keys())}")

        # 分组字段：普通字段、检测数据列、数据区结束标记
        detection_columns = {}  # {sheet_name: {column_mapping: cell_address}}
        data_region_ends = {}   # {sheet_name: end_row}

        # 按单元格分组字段，用于处理同一单元格包含多个字段标记的情况
        cell_fields = {}  # {(sheet_name, cell_address): [fields]}

        for field in fields:
            field_name = field['field_name']
            field_type = field['field_type']
            sheet_name = field['sheet_name']
            cell_address = field['cell_address']
            is_reference = field.get('is_reference', False)

            # 如果没有字段映射配置，跳过
            if not cell_address or sheet_name not in self.workbook.sheetnames:
                print(f"跳过字段 {field_name}: cell_address={cell_address}, sheet存在={sheet_name in self.workbook.sheetnames if self.workbook else False}")
                continue

            # 获取工作表
            ws = self.workbook[sheet_name]

            # 检测数据列标记：收集列映射信息
            if field_type == 'detection_column':
                column_mapping = field.get('column_mapping')
                if column_mapping:
                    if sheet_name not in detection_columns:
                        detection_columns[sheet_name] = {}
                    detection_columns[sheet_name][column_mapping] = cell_address
                    print(f"检测数据列: [{field_name}] -> {column_mapping} 在 {cell_address}")
                continue

            # 控制标记：数据区结束标记
            # 修复：数据库中没有control_type字段，通过field_type和field_name识别
            if field_type == 'control_mark' and field_name == 'data_region_end':
                from openpyxl.utils.cell import coordinate_from_string
                _, end_row = coordinate_from_string(cell_address)
                data_region_ends[sheet_name] = end_row
                print(f"✓ 数据区结束标记: 工作表 {sheet_name}, 结束行 {end_row}, 单元格 {cell_address}")

                # 清除控制标记单元格内容和格式，避免在报告中显示
                cell = ws[cell_address]
                cell.value = None
                # 清除背景色
                from openpyxl.styles import PatternFill
                cell.fill = PatternFill(fill_type=None)
                print(f"  已清除控制标记显示和背景色")

                continue

            # 根据字段类型填充数据
            if field_type == 'table_data':
                # 表格数据特殊处理
                print(f"填充表格数据: {field_name}")
                self._fill_table_data(ws, field)
            else:
                # 对于普通字段，按单元格分组处理（支持同一单元格多个字段标记）
                cell_key = (sheet_name, cell_address)
                if cell_key not in cell_fields:
                    cell_fields[cell_key] = []
                cell_fields[cell_key].append(field)

        # 批量处理单元格填充（处理同一单元格的多个字段）
        for (sheet_name, cell_address), fields_in_cell in cell_fields.items():
            ws = self.workbook[sheet_name]

            # 获取第一个字段的原始文本作为基准
            original_text = fields_in_cell[0].get('original_cell_text', '')

            if original_text and original_text.strip():
                # 有原始文本，进行字符串替换
                filled_text = original_text

                for field in fields_in_cell:
                    field_name = field['field_name']
                    is_reference = field.get('is_reference', False)

                    # 获取字段值
                    if is_reference:
                        value = self._get_reference_value(field_name)
                        print(f"引用字段 [{field_name}] = {value} (来自已审核报告)")
                    else:
                        value = self.report_data.get(field_name, field.get('default_value', ''))
                        print(f"普通字段 [{field_name}] = {value}")

                    if value is not None:
                        # 构建字段标记
                        field_code = field.get('field_code')
                        if field_code:
                            field_marker = f"[{field_code}]"
                        elif is_reference:
                            field_marker = f"[*{field_name}]"
                        else:
                            field_marker = f"[{field_name}]"

                        # 日期格式转换：将YYYY-MM-DD转换为YYYY年MM月DD日
                        if field.get('field_type') == 'date' or 'date' in field_name.lower() or '日期' in field_name:
                            value = self._format_date_chinese(value)

                        # 累积替换
                        filled_text = filled_text.replace(field_marker, str(value))
                        print(f"  ✓ 替换 '{field_marker}' -> '{value}'")

                ws[cell_address] = filled_text
                print(f"  ✓ 完成填充到 {sheet_name}!{cell_address}: {repr(filled_text)}")
            else:
                # 没有原始文本，直接填充第一个字段的值
                field = fields_in_cell[0]
                field_name = field['field_name']
                is_reference = field.get('is_reference', False)

                if is_reference:
                    value = self._get_reference_value(field_name)
                else:
                    value = self.report_data.get(field_name, field.get('default_value', ''))

                if value is not None:
                    # 日期格式转换
                    if field.get('field_type') == 'date' or 'date' in field_name.lower() or '日期' in field_name:
                        value = self._format_date_chinese(value)

                    ws[cell_address] = value
                    print(f"  ✓ 已填充到 {sheet_name}!{cell_address}: {value}")

        # 填充检测数据（使用动态列位置，支持跨页填充）
        if detection_columns:
            self._fill_detection_data_by_columns(detection_columns, data_region_ends)

        print(f"=== 数据填充完成 ===\n")

    def _format_date_chinese(self, date_value):
        """
        将日期转换为中文格式

        Args:
            date_value: 日期值，可能是字符串或datetime对象

        Returns:
            str: 中文格式的日期，如 "2026年02月02日"
        """
        if not date_value:
            return ''

        try:
            # 如果是字符串，尝试解析
            if isinstance(date_value, str):
                # 支持多种日期格式
                from datetime import datetime

                # 尝试常见格式
                for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y%m%d']:
                    try:
                        dt = datetime.strptime(date_value, fmt)
                        return f"{dt.year}年{dt.month:02d}月{dt.day:02d}日"
                    except ValueError:
                        continue

                # 如果已经是中文格式，直接返回
                if '年' in date_value and '月' in date_value:
                    return date_value

                # 无法解析，返回原值
                return str(date_value)

            # 如果是datetime对象
            elif hasattr(date_value, 'year'):
                return f"{date_value.year}年{date_value.month:02d}月{date_value.day:02d}日"

            return str(date_value)
        except Exception as e:
            print(f"  ⚠ 日期格式转换失败: {e}")
            return str(date_value)

    def _get_reference_value(self, field_name):
        """
        从已审核报告中获取引用字段的值

        Args:
            field_name: 字段名（如：被检单位、采样日期等）

        Returns:
            str: 字段值，如果找不到返回空字符串
        """
        conn = get_db_connection()

        # 字段名到数据库字段的映射
        field_mapping = {
            '报告编号': 'report_number',
            '样品编号': 'sample_number',
            '样品类型': 'sample_type_name',
            '被检单位': 'customer_unit',
            '被检水厂': 'customer_plant',
            '单位地址': 'unit_address',
            '委托单位': 'company_name',
            '采样人': 'sampler',
            '采样日期': 'sampling_date',
            '采样地点': 'sampling_location',
            '采样依据': 'sampling_basis',
            '样品来源': 'sample_source',
            '样品状态': 'sample_status',
            '收样日期': 'sample_received_date',
            '检测日期': 'detection_date',
            '检测人': 'detection_person',
            '检测人员': 'detection_person',
            '审核人': 'review_person',
            '审核人员': 'review_person',
            '报告编制日期': 'report_date',
            '产品标准': 'product_standard',
            '检测项目': 'detection_items_description',
            '检测结论': 'test_conclusion',
            '附加信息': 'additional_info',
            '附件信息': 'attachment_info',
            '备注': 'remark'
        }

        db_field = field_mapping.get(field_name)

        if not db_field:
            print(f"警告: 未知的引用字段 '{field_name}'")
            conn.close()
            return ''

        try:
            # 从最近的已审核报告中查询数据
            # 优先查找相同样品编号的已审核报告
            sample_number = self.report_data.get('sample_number', '')

            if sample_number:
                # 先尝试查找相同样品编号的已审核报告
                query = f'''
                    SELECT r.{db_field}, r.remark
                    FROM reports r
                    LEFT JOIN sample_types st ON r.sample_type_id = st.id
                    WHERE r.review_status = 'approved'
                    AND r.sample_number = ?
                    ORDER BY r.created_at DESC
                    LIMIT 1
                '''
                result = conn.execute(query, (sample_number,)).fetchone()

                if result and result[db_field]:
                    value = result[db_field]

                    # 如果字段为空但在remark的JSON中，尝试从remark中提取
                    if not value and result['remark']:
                        try:
                            import json
                            remark_data = json.loads(result['remark'])
                            # 检查是否有customer_unit和customer_plant
                            if field_name == '被检单位' and 'customer_unit' in remark_data:
                                value = remark_data['customer_unit']
                            elif field_name == '被检水厂' and 'customer_plant' in remark_data:
                                value = remark_data['customer_plant']
                            elif field_name == '单位地址' and 'unit_address' in remark_data:
                                value = remark_data['unit_address']
                        except:
                            pass

                    conn.close()
                    return value or ''

            # 如果找不到相同样品编号的，查找最近的已审核报告
            query = f'''
                SELECT r.{db_field}, r.remark
                FROM reports r
                WHERE r.review_status = 'approved'
                ORDER BY r.created_at DESC
                LIMIT 1
            '''
            result = conn.execute(query).fetchone()

            if result:
                value = result[db_field] or ''

                # 如果字段为空但在remark的JSON中，尝试从remark中提取
                if not value and result['remark']:
                    try:
                        import json
                        remark_data = json.loads(result['remark'])
                        if field_name == '被检单位' and 'customer_unit' in remark_data:
                            value = remark_data['customer_unit']
                        elif field_name == '被检水厂' and 'customer_plant' in remark_data:
                            value = remark_data['customer_plant']
                        elif field_name == '单位地址' and 'unit_address' in remark_data:
                            value = remark_data['unit_address']
                    except:
                        pass

                conn.close()
                return value

            conn.close()
            return ''

        except Exception as e:
            print(f"查询引用字段失败 {field_name}: {e}")
            conn.close()
            return ''

    def _fill_table_data(self, worksheet, field):
        """
        填充表格数据（检测结果）

        Args:
            worksheet: 工作表对象
            field: 字段配置
        """
        detection_items = self.report_data.get('detection_items', [])

        if not detection_items:
            return

        start_row = field.get('start_row', 8)  # 默认从第8行开始
        start_col = field.get('start_col', 1)  # 默认从第1列开始

        # 填充每一行数据
        for idx, item in enumerate(detection_items):
            row = start_row + idx

            # 序号
            worksheet.cell(row, start_col).value = idx + 1

            # 项目名称
            worksheet.cell(row, start_col + 1).value = item.get('name', '')

            # 单位
            worksheet.cell(row, start_col + 2).value = item.get('unit', '')

            # 检测结果
            worksheet.cell(row, start_col + 3).value = item.get('result', '')

            # 标准限值
            worksheet.cell(row, start_col + 4).value = item.get('limit', '')

            # 检测方法
            worksheet.cell(row, start_col + 5).value = item.get('method', '')

    def _format_detection_method(self, method_text):
        """
        格式化检测方法，自动在标准编号和方法名称之间添加换行

        Args:
            method_text: 原始检测方法文本

        Returns:
            str: 格式化后的检测方法（包含换行符）

        示例：
            输入: "GB/T 5750.5-2023 4.2 离子色谱法"
            输出: "GB/T 5750.5-2023 4.2\n离子色谱法"
        """
        if not method_text or '\n' in method_text:
            return method_text  # 已经包含换行符，直接返回

        import re

        # 匹配标准编号模式: GB/T xxxx-xxxx 或 GB xxxx-xxxx
        # 策略：先匹配标准编号，然后在剩余部分找第一个汉字位置进行分割
        pattern = r'^((?:GB/?T?|HJ|CJ)\s*\d+(?:\.\d+)?-\d+)\s*(.+)$'

        match = re.match(pattern, method_text, re.IGNORECASE)
        if match:
            standard = match.group(1).strip()  # 标准编号部分 (如 GB/T 5750.4-2023)
            suffix = match.group(2).strip()     # 剩余部分 (如 "7.1直接观察法" 或 "4.2 离子色谱法")

            # 在suffix中找第一个汉字的位置
            chinese_match = re.search(r'[\u4e00-\u9fff]', suffix)
            if chinese_match:
                pos = chinese_match.start()
                chapter = suffix[:pos].strip()   # 章节号 (如 "7.1" 或 "4.2")
                method_name = suffix[pos:].strip()  # 方法名 (如 "直接观察法" 或 "离子色谱法")

                if chapter:
                    # 有章节号，标准编号+章节号一行，方法名另一行
                    return f"{standard} {chapter}\n{method_name}"
                else:
                    # 没有章节号，标准编号一行，方法名另一行
                    return f"{standard}\n{method_name}"
            else:
                # suffix中没有汉字（纯数字/字母），不分割
                return f"{standard}\n{suffix}"

        # 如果没有匹配到标准模式，尝试其他常见分隔
        # 例如：按最后一个数字后的空格分隔
        parts = method_text.rsplit(' ', 1)
        if len(parts) == 2 and parts[1] and not parts[1][0].isdigit():
            return f"{parts[0]}\n{parts[1]}"

        return method_text  # 无法识别模式，返回原文

    def _auto_fit_font_size(self, cell, col_width, max_row_height=None):
        """
        当单元格文本可能超出列宽时，自动缩小字体以确保内容可见。
        仅在文本超长时缩小，其余单元格保持原字体大小不变。

        Args:
            cell: openpyxl单元格对象
            col_width: 列宽（Excel字符单位）
            max_row_height: 最大行高(pt)，用于估算可用行数
        """
        from openpyxl.styles import Font, Alignment
        from copy import copy

        value = cell.value
        if not value or not isinstance(value, str):
            return

        # 获取当前字体大小
        original_size = cell.font.size or 9.0

        # Excel列宽单位与字符宽度的对应关系（9pt字体下）：
        # - 中文字符约占 1.2 列宽单位
        # - ASCII字符约占 0.6 列宽单位
        CJK_WIDTH = 1.2  # 列宽单位 per Chinese char at 9pt
        ASCII_WIDTH = 0.5  # 列宽单位 per ASCII char at 9pt

        # 计算一行文本占用的列宽单位
        def calc_line_width(line, font_size):
            ratio = font_size / 9.0
            width = 0
            for ch in line:
                if ord(ch) > 127:
                    width += CJK_WIDTH * ratio
                else:
                    width += ASCII_WIDTH * ratio
            return width

        # 计算文本显示所需的行数
        def calc_lines_needed(text, font_size):
            lines = text.split('\n')
            total = 0
            for line in lines:
                line_w = calc_line_width(line, font_size)
                if line_w <= col_width:
                    total += 1
                else:
                    total += max(1, -(-int(line_w * 100) // int(col_width * 100)))  # ceiling division
            return total

        # 估算最大可用行数（基于行高限制，+1容差允许轻微溢出）
        max_lines = 4  # 默认最大4行
        if max_row_height:
            line_height_pt = original_size * 1.4
            max_lines = max(2, int(max_row_height / line_height_pt))

        # 尝试当前字体大小
        lines = calc_lines_needed(value, original_size)

        if lines <= max_lines:
            return  # 当前字体大小可以容纳，无需缩小

        print(f"  [AUTO-FIT] Cell {cell.coordinate}: lines_needed={lines}, max_lines={max_lines}, col_w={col_width}, value='{value[:50]}...'")

        # 逐步缩小字体，最小到5pt
        for size in [original_size - 1, original_size - 2, 7, 6.5, 6, 5.5, 5]:
            if size < 5:
                break
            lines = calc_lines_needed(value, size)
            if lines <= max_lines:
                # 应用缩小的字体
                new_font = copy(cell.font)
                cell.font = Font(
                    name=new_font.name,
                    size=size,
                    bold=new_font.bold,
                    italic=new_font.italic,
                    color=new_font.color,
                    underline=new_font.underline,
                    strikethrough=new_font.strikethrough
                )
                # 确保启用自动换行
                current_alignment = cell.alignment
                cell.alignment = Alignment(
                    horizontal=current_alignment.horizontal or 'center',
                    vertical=current_alignment.vertical or 'center',
                    wrap_text=True
                )
                return

        # 即使最小字体也放不下，仍使用最小字体+换行
        new_font = copy(cell.font)
        cell.font = Font(
            name=new_font.name,
            size=5,
            bold=new_font.bold,
            italic=new_font.italic,
            color=new_font.color,
            underline=new_font.underline,
            strikethrough=new_font.strikethrough
        )
        current_alignment = cell.alignment
        cell.alignment = Alignment(
            horizontal=current_alignment.horizontal or 'center',
            vertical=current_alignment.vertical or 'center',
            wrap_text=True
        )

    def _auto_adjust_row_height(self, sheet_name, start_row, row_count):
        """
        根据单元格内容自动调整行高

        Args:
            sheet_name: 工作表名称
            start_row: 起始行号
            row_count: 需要调整的行数
        """
        if sheet_name not in self.workbook.sheetnames:
            return

        ws = self.workbook[sheet_name]

        for i in range(row_count):
            row_num = start_row + i
            max_lines = 1

            # 检查该行所有单元格，找出最大行数
            for cell in ws[row_num]:
                if cell.value:
                    cell_text = str(cell.value)
                    lines = cell_text.count('\n') + 1
                    max_lines = max(max_lines, lines)

            # 根据行数计算高度
            # 优化为适应A4打印：22行数据需要控制在约600点以内
            # 单行：17点，双行：26点，三行：35点
            base_height = 8
            line_height = 9
            calculated_height = base_height + (max_lines * line_height)

            # 限制在合理范围内
            min_height = 15
            max_height = 80  # 降低最大高度限制
            final_height = max(min_height, min(calculated_height, max_height))

            # 设置行高
            ws.row_dimensions[row_num].height = final_height

    def _fill_detection_data_by_columns(self, detection_columns, data_region_ends):
        """
        使用动态列位置填充检测数据，支持跨页填充

        Args:
            detection_columns: {sheet_name: {column_mapping: cell_address}}
                例如: {'Sheet1': {'name': 'B8', 'unit': 'C8', 'result': 'D8'}}
            data_region_ends: {sheet_name: end_row}
                例如: {'Sheet1': 30, 'Sheet2': 30}
        """
        detection_items = self.report_data.get('detection_items', [])

        if not detection_items:
            print("没有检测数据需要填充")
            return

        print(f"\n=== 使用动态列位置填充检测数据（支持跨页） ===")
        print(f"检测项目数量: {len(detection_items)}")

        from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

        # 按工作表名称排序，确保按顺序填充（Sheet1, Sheet2, ...）
        sorted_sheets = sorted(detection_columns.keys())

        # 构建数据页信息列表
        data_pages = []
        for sheet_name in sorted_sheets:
            if sheet_name not in self.workbook.sheetnames:
                print(f"警告: 工作表 '{sheet_name}' 不存在")
                continue

            columns = detection_columns[sheet_name]

            # 解析起始行
            first_cell = list(columns.values())[0]
            _, start_row = coordinate_from_string(first_cell)

            # 获取结束行（如果有标记）
            end_row = data_region_ends.get(sheet_name)
            if end_row:
                capacity = end_row - start_row  # 实际可用行数（不包括结束标记行）
            else:
                # 如果没有结束标记，默认容量为1000行（实际上无限制）
                capacity = 1000

            data_pages.append({
                'sheet_name': sheet_name,
                'columns': columns,
                'start_row': start_row,
                'end_row': end_row,
                'capacity': capacity
            })

            print(f"数据页: {sheet_name}, 起始行: {start_row}, 结束行: {end_row or '无限制'}, 容量: {capacity}行")

        if not data_pages:
            print("警告: 没有有效的数据页")
            return

        # 跨页填充检测数据
        item_index = 0  # 当前检测项目索引

        for page in data_pages:
            sheet_name = page['sheet_name']
            columns = page['columns']
            start_row = page['start_row']
            capacity = page['capacity']

            ws = self.workbook[sheet_name]

            # 计算本页要填充的数据量
            remaining_items = len(detection_items) - item_index
            items_to_fill = min(capacity, remaining_items)

            if items_to_fill <= 0:
                print(f"  ℹ {sheet_name}: 无需填充（所有数据已填充完）")
                break

            print(f"\n工作表: {sheet_name}")
            print(f"  本页填充: {items_to_fill} 行 (从第 {item_index + 1} 项到第 {item_index + items_to_fill} 项)")

            # 填充本页数据
            for i in range(items_to_fill):
                item = detection_items[item_index + i]
                current_row = start_row + i

                # 根据列映射填充数据
                for mapping, cell_address in columns.items():
                    col_letter, _ = coordinate_from_string(cell_address)
                    col_index = column_index_from_string(col_letter)

                    # 根据映射类型获取数据
                    if mapping == 'index':
                        value = item_index + i + 1  # 全局序号
                    elif mapping == 'name':
                        value = item.get('name', '')
                    elif mapping == 'unit':
                        value = item.get('unit', '')
                    elif mapping == 'result':
                        value = item.get('result', '')
                    elif mapping == 'limit':
                        value = item.get('limit', '')
                    elif mapping == 'method':
                        raw_method = item.get('method', '')
                        value = self._format_detection_method(raw_method)  # 格式化检测方法
                    elif mapping == 'judgment':
                        value = item.get('judgment', '')
                    else:
                        value = ''

                    try:
                        cell = ws.cell(row=current_row, column=col_index)
                        cell.value = value

                        # 如果值包含换行符，仅启用wrap_text，保留模板原有对齐格式
                        if value and isinstance(value, str) and '\n' in value:
                            from copy import copy
                            new_align = copy(cell.alignment)
                            new_align.wrap_text = True
                            cell.alignment = new_align

                        if i == 0:  # 只打印第一行的详细信息
                            print(f"  列 {mapping}: {col_letter}{current_row} = '{value}'")
                    except Exception as e:
                        print(f"  ✗ 填充失败 {col_letter}{current_row}: {e}")

            print(f"  ✓ 已填充 {items_to_fill} 行到 {sheet_name}")

            # 保留模板原始行高，不做自动调整

            item_index += items_to_fill

            # 如果所有数据都已填充完，退出循环
            if item_index >= len(detection_items):
                break

        print(f"\n=== 检测数据填充完成 ===")
        print(f"总计填充: {item_index} / {len(detection_items)} 项")

        if item_index < len(detection_items):
            print(f"⚠ 警告: 有 {len(detection_items) - item_index} 项数据未填充（数据页容量不足）")

    def _sanitize_filename(self, filename):
        """
        清理文件名中的非法字符

        Args:
            filename: 原始文件名

        Returns:
            str: 清理后的文件名
        """
        import re

        # Windows和Linux文件名非法字符
        illegal_chars = r'[<>:"/\\|?*]'

        # 替换非法字符为下划线
        filename = re.sub(illegal_chars, '_', filename)

        # 替换中文括号为英文括号（可选，保持兼容性）
        filename = filename.replace('（', '(').replace('）', ')')

        # 去除首尾空格
        filename = filename.strip()

        # 如果文件名为空，使用默认值
        if not filename:
            filename = 'report'

        return filename

    def _convert_to_pdf(self, excel_path):
        """
        将Excel文件转换为PDF

        Args:
            excel_path: Excel文件路径

        Returns:
            str: PDF文件路径，如果转换失败返回None
        """
        import subprocess
        import platform

        # 生成PDF文件路径
        pdf_path = excel_path.replace('.xlsx', '.pdf')

        try:
            # 检查LibreOffice是否可用
            system = platform.system()

            if system == 'Linux':
                # Linux系统使用libreoffice
                cmd = [
                    'libreoffice',
                    '--headless',
                    '--convert-to', 'pdf',
                    '--outdir', os.path.dirname(excel_path),
                    excel_path
                ]
            elif system == 'Windows':
                # Windows系统使用soffice.exe
                cmd = [
                    'soffice.exe',
                    '--headless',
                    '--convert-to', 'pdf',
                    '--outdir', os.path.dirname(excel_path),
                    excel_path
                ]
            else:
                print(f"⚠ 不支持的操作系统: {system}")
                return None

            # 执行转换
            print(f"正在转换为PDF: {excel_path}")
            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                timeout=30  # 30秒超时
            )

            if result.returncode == 0 and os.path.exists(pdf_path):
                print(f"✓ PDF转换成功: {pdf_path}")
                # 删除原Excel文件（可选）
                # os.remove(excel_path)
                return pdf_path
            else:
                print(f"✗ PDF转换失败: {result.stderr}")
                return None

        except FileNotFoundError:
            print("⚠ LibreOffice未安装，无法转换为PDF")
            print("  安装方法: sudo apt-get install libreoffice")
            return None
        except subprocess.TimeoutExpired:
            print("✗ PDF转换超时")
            return None
        except Exception as e:
            print(f"✗ PDF转换异常: {str(e)}")
            return None

def generate_simple_report(report_id):
    """
    简化版报告生成（不依赖模版）
    直接从数据库读取报告数据并生成Excel

    Args:
        report_id: 报告ID

    Returns:
        str: 生成的文件路径
    """
    conn = get_db_connection()

    # 获取报告基本信息
    report = conn.execute(
        'SELECT r.*, st.name as sample_type_name, c.name as company_name '
        'FROM reports r '
        'LEFT JOIN sample_types st ON r.sample_type_id = st.id '
        'LEFT JOIN companies c ON r.company_id = c.id '
        'WHERE r.id = ?',
        (report_id,)
    ).fetchone()

    if not report:
        conn.close()
        raise ValueError(f"报告不存在: ID={report_id}")

    # 获取检测数据
    data_items = conn.execute(
        'SELECT rd.*, i.name as indicator_name, i.unit, i.limit_value, i.detection_method '
        'FROM report_data rd '
        'JOIN indicators i ON rd.indicator_id = i.id '
        'JOIN reports r ON rd.report_id = r.id '
        'LEFT JOIN template_indicators ti '
        '    ON ti.indicator_id = rd.indicator_id AND ti.sample_type_id = r.sample_type_id '
        'WHERE rd.report_id = ? '
        'ORDER BY ti.sort_order, i.sort_order, i.name',
        (report_id,)
    ).fetchall()

    conn.close()

    # 创建Excel文件
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "检测报告"

    # 设置标题
    ws['A1'] = "水质检测报告"
    ws['A1'].font = Font(size=18, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:F1')

    # 基本信息
    row = 3
    ws[f'A{row}'] = "报告编号："
    ws[f'B{row}'] = report['report_number']
    ws[f'D{row}'] = "样品编号："
    ws[f'E{row}'] = report['sample_number']

    row += 1
    ws[f'A{row}'] = "样品类型："
    ws[f'B{row}'] = report['sample_type_name']
    ws[f'D{row}'] = "委托单位："
    ws[f'E{row}'] = report['company_name'] or '-'

    row += 1
    ws[f'A{row}'] = "检测日期："
    ws[f'B{row}'] = report['detection_date'] or '-'
    ws[f'D{row}'] = "检测人员："
    ws[f'E{row}'] = report['detection_person'] or '-'

    # 检测数据表格
    row += 2
    headers = ['序号', '检测项目', '单位', '检测结果', '标准限值', '检测方法']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row, col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # 填充检测数据
    for idx, item in enumerate(data_items, start=1):
        row += 1
        ws.cell(row, 1).value = idx
        ws.cell(row, 2).value = item['indicator_name']
        ws.cell(row, 3).value = item['unit'] or '-'
        ws.cell(row, 4).value = item['measured_value'] or '-'
        ws.cell(row, 5).value = item['limit_value'] or '-'
        ws.cell(row, 6).value = item['detection_method'] or '-'

    # 调整列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 40

    # 保存文件
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = f"exports/report_{report['report_number']}_{timestamp}.xlsx"
    os.makedirs('exports', exist_ok=True)
    wb.save(output_path)

    return output_path

if __name__ == '__main__':
    # 测试简化版报告生成
    print("报告生成器已就绪")
    print("使用方法：")
    print("1. 按模版生成：ReportGenerator(template_id, report_data).generate()")
    print("2. 简化版生成：generate_simple_report(report_id)")
