"""
报告模板导出器
用于导出报告基本信息填写模板
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from models_v2 import get_db_connection
import os
from datetime import datetime


class ReportTemplateExporter:
    """报告模板导出器"""

    def __init__(self, template_id):
        """
        初始化导出器

        Args:
            template_id: 报告模板ID
        """
        self.template_id = template_id
        self.template_info = None
        self.fields = []

    def export(self, output_path=None):
        """
        导出报告填写模板

        Args:
            output_path: 输出文件路径

        Returns:
            str: 生成的文件路径
        """
        # 加载模板数据
        self._load_template_data()

        # 创建工作簿
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # 1. 创建说明sheet
        self._create_instruction_sheet(wb)

        # 2. 创建报告基本信息sheet
        self._create_report_info_sheet(wb)

        # 保存文件
        if output_path is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            template_name = self.template_info.get('name', '报告模板')
            output_path = f"exports/report_template_{template_name}_{timestamp}.xlsx"

        os.makedirs('exports', exist_ok=True)
        wb.save(output_path)
        wb.close()

        return output_path

    def _load_template_data(self):
        """加载模板数据"""
        conn = get_db_connection()

        # 加载模板基本信息
        template = conn.execute(
            'SELECT * FROM excel_report_templates WHERE id = ?',
            (self.template_id,)
        ).fetchone()

        if not template:
            conn.close()
            raise ValueError(f'模板ID {self.template_id} 不存在')

        self.template_info = dict(template)

        # 加载字段映射
        fields = conn.execute(
            'SELECT * FROM template_field_mappings WHERE template_id = ? ORDER BY id',
            (self.template_id,)
        ).fetchall()

        self.fields = [dict(f) for f in fields]

        conn.close()

    def _create_instruction_sheet(self, wb):
        """创建填写说明sheet"""
        ws = wb.create_sheet("填写说明", 0)

        ws['A1'] = "报告基本信息填写说明"
        ws['A1'].font = Font(size=16, bold=True, color="4472C4")

        row = 3

        template_name = self.template_info.get('name', '未知')

        instructions = [
            ("一、模板信息", ""),
            ("", f"报告模板: {template_name}"),
            ("", f"字段数量: {len(self.fields)}"),
            ("", f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"),
            ("", ""),
            ("二、填写说明", ""),
            ("", "【报告基本信息】sheet 说明:"),
            ("", "  - 每一行代表一份报告的基本信息"),
            ("", "  - 必填字段标记了 * 号，请务必填写"),
            ("", "  - 样品编号是报告的唯一标识，建议格式：S+日期+序号"),
            ("", "  - 填写完成后保存此文件，在系统中导入"),
            ("", ""),
            ("三、字段说明", ""),
        ]

        # 添加字段说明
        for field in self.fields:
            field_name = field.get('field_display_name') or field.get('field_name')
            is_required = field.get('is_required', 0)
            description = field.get('description', '')
            required_mark = '*' if is_required else ''

            instructions.append(("", f"  【{field_name}{required_mark}】: {description}"))

        instructions.extend([
            ("", ""),
            ("四、注意事项", ""),
            ("", "1. 样品编号必须唯一"),
            ("", "2. 必填字段不能为空"),
            ("", "3. 日期格式统一使用 YYYY-MM-DD"),
            ("", "4. 删除示例数据行，填入实际数据"),
            ("", "5. 可以填写多份报告的数据"),
            ("", ""),
            ("五、导入步骤", ""),
            ("", "1. 填写完成此Excel文件"),
            ("", "2. 在系统【报告填写】页面点击【导入报告基本信息】"),
            ("", "3. 选择此文件上传"),
            ("", "4. 系统将自动创建报告记录"),
        ])

        for title, content in instructions:
            if title:
                cell = ws.cell(row, 1)
                cell.value = title
                cell.font = Font(size=12, bold=True, color="4472C4")
            else:
                cell = ws.cell(row, 1)
                cell.value = content
                cell.alignment = Alignment(wrap_text=True)

            row += 1

        ws.column_dimensions['A'].width = 100

    def _create_report_info_sheet(self, wb):
        """创建报告基本信息sheet（横向格式：字段为行，样品为列）"""
        ws = wb.create_sheet("报告基本信息")

        # 设置样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        subheader_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
        subheader_font = Font(bold=True, size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # A1: 字段名称标题
        cell = ws.cell(1, 1)
        cell.value = "字段名称"
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

        # 首行：样品编号列（3个示例）
        sample_numbers = ['样品编号1*', '样品编号2', '样品编号3']
        for col_idx, sample_number in enumerate(sample_numbers, start=2):
            cell = ws.cell(1, col_idx)
            cell.value = sample_number
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # 第一列：字段名称
        row_idx = 2

        # 样品编号字段（必填）
        cell = ws.cell(row_idx, 1)
        cell.value = "样品编号*"
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = border

        # 示例样品编号
        cell = ws.cell(row_idx, 2)
        cell.value = "S20260128001"
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

        # 其他样品列留空
        for col_idx in range(3, 2 + len(sample_numbers)):
            cell = ws.cell(row_idx, col_idx)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        row_idx += 1

        # 添加模板字段
        for field in self.fields:
            field_name = field.get('field_display_name') or field.get('field_name')
            is_required = field.get('is_required', 0)
            if is_required:
                field_name += '*'

            # 字段名称
            cell = ws.cell(row_idx, 1)
            cell.value = field_name
            cell.fill = subheader_fill
            cell.font = subheader_font
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = border

            # 第一个样品的示例值
            default_value = field.get('default_value', '')
            placeholder = field.get('placeholder', '')
            example_value = default_value if default_value else placeholder if placeholder else ''

            if row_idx == 2:  # 仅第一个字段填示例
                cell = ws.cell(row_idx, 2)
                cell.value = example_value
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            # 其他样品列留空
            for col_idx in range(2 if row_idx > 2 else 3, 2 + len(sample_numbers)):
                cell = ws.cell(row_idx, col_idx)
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')

            row_idx += 1

        # 调整列宽
        ws.column_dimensions['A'].width = 25  # 字段名称列
        for col_letter in ['B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col_letter].width = 20  # 样品数据列

        # 冻结首行首列
        ws.freeze_panes = 'B2'


def export_report_template(template_id, output_path=None):
    """
    导出报告填写模板的便捷函数

    Args:
        template_id: 报告模板ID
        output_path: 输出路径（可选）

    Returns:
        str: 生成的文件路径
    """
    exporter = ReportTemplateExporter(template_id)
    return exporter.export(output_path)


if __name__ == '__main__':
    print("="*60)
    print("报告模板导出器测试")
    print("="*60)

    conn = get_db_connection()
    templates = conn.execute('SELECT id, name FROM excel_report_templates WHERE is_active = 1 LIMIT 1').fetchone()
    conn.close()

    if templates:
        template_id = templates['id']
        template_name = templates['name']
        print(f"\n导出模板 [{template_name}] 的填写模板...")
        path = export_report_template(template_id)
        print(f"✓ 生成成功: {path}")
    else:
        print("\n没有找到活动的报告模板")
