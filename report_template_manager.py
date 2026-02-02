"""
报告模版管理器
用于导入、识别和管理Excel报告模版
"""
import os
import shutil
import openpyxl
from datetime import datetime
from models_v2 import get_db_connection
from models_report_template import get_standard_field_definitions

class ReportTemplateManager:
    """报告模版管理器"""

    def __init__(self):
        self.template_dir = 'templates/excel_reports'
        os.makedirs(self.template_dir, exist_ok=True)

    def import_template(self, template_file_path, template_name, sample_type_id=None, description=''):
        """
        导入Excel报告模版

        Args:
            template_file_path: Excel模版文件路径
            template_name: 模版名称
            sample_type_id: 关联的样品类型ID（可选）
            description: 模版描述

        Returns:
            template_id: 创建的模版ID
        """
        if not os.path.exists(template_file_path):
            raise FileNotFoundError(f"模版文件不存在: {template_file_path}")

        # 检查文件是否为Excel文件
        if not template_file_path.endswith(('.xlsx', '.xls')):
            raise ValueError("模版文件必须是Excel格式")

        try:
            # 读取Excel文件
            wb = openpyxl.load_workbook(template_file_path)

            # 复制模版文件到模版目录
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            new_filename = f"{template_name}_{timestamp}.xlsx"
            new_file_path = os.path.join(self.template_dir, new_filename)
            shutil.copy2(template_file_path, new_file_path)

            # 保存到数据库
            conn = get_db_connection()
            cursor = conn.cursor()

            cursor.execute(
                'INSERT INTO excel_report_templates (name, sample_type_id, description, template_file_path) '
                'VALUES (?, ?, ?, ?)',
                (template_name, sample_type_id, description, new_file_path)
            )

            template_id = cursor.lastrowid

            # 分析工作表结构
            self._analyze_template_structure(template_id, wb)

            # 扫描并保存字段映射
            field_count = self._scan_and_save_fields(template_id, new_file_path)

            conn.commit()
            conn.close()

            print(f"✓ 模版导入成功: {template_name}")
            print(f"  工作表数量: {len(wb.sheetnames)}")
            print(f"  识别字段数量: {field_count}")
            print(f"  文件路径: {new_file_path}")

            return template_id

        except Exception as e:
            print(f"✗ 导入模版失败: {str(e)}")
            raise

    def _analyze_template_structure(self, template_id, workbook):
        """
        分析模版结构并保存工作表配置

        Args:
            template_id: 模版ID
            workbook: openpyxl workbook对象
        """
        conn = get_db_connection()
        cursor = conn.cursor()

        for index, sheet_name in enumerate(workbook.sheetnames):
            # 判断工作表类型
            sheet_type = self._identify_sheet_type(sheet_name)

            cursor.execute(
                'INSERT INTO template_sheet_configs '
                '(template_id, sheet_name, sheet_index, sheet_type, page_number) '
                'VALUES (?, ?, ?, ?, ?)',
                (template_id, sheet_name, index, sheet_type, self._extract_page_number(sheet_name))
            )

        conn.commit()
        conn.close()

    def _identify_sheet_type(self, sheet_name):
        """
        根据工作表名称识别类型

        Args:
            sheet_name: 工作表名称

        Returns:
            sheet_type: 工作表类型（cover/info/data/conclusion）
        """
        sheet_name_lower = sheet_name.lower()

        if '1' in sheet_name or 'cover' in sheet_name_lower or '封面' in sheet_name:
            return 'cover'  # 封面页
        elif '2' in sheet_name or 'info' in sheet_name_lower or '信息' in sheet_name:
            return 'info'  # 信息页
        elif any(x in sheet_name for x in ['3', '4']) or 'data' in sheet_name_lower or '数据' in sheet_name:
            return 'data'  # 数据页
        elif '5' in sheet_name or 'note' in sheet_name_lower or '说明' in sheet_name:
            return 'conclusion'  # 说明页
        else:
            return 'other'  # 其他

    def _extract_page_number(self, sheet_name):
        """
        从工作表名称中提取页码

        Args:
            sheet_name: 工作表名称

        Returns:
            page_number: 页码数字
        """
        import re
        match = re.search(r'\d+', sheet_name)
        return int(match.group()) if match else 0

    def _scan_and_save_fields(self, template_id, template_file_path):
        """
        扫描模板文件中的所有字段并保存到数据库

        Args:
            template_id: 模板ID
            template_file_path: 模板文件路径

        Returns:
            int: 识别到的字段数量
        """
        from template_field_parser import TemplateFieldParser

        # 提取所有字段
        fields = TemplateFieldParser.extract_template_fields(template_file_path)

        if not fields:
            print("  ⚠ 未在模板中发现任何字段标记")
            return 0

        conn = get_db_connection()
        cursor = conn.cursor()

        saved_count = 0
        reference_count = 0

        for field in fields:
            try:
                # 检查是否为引用字段
                is_reference = field.get('is_reference', False)

                # 保存字段映射
                cursor.execute('''
                    INSERT INTO template_field_mappings
                    (template_id, field_name, field_display_name, field_type,
                     sheet_name, cell_address, placeholder, default_value,
                     is_required, is_reference, description)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    template_id,
                    field['field_name'],
                    field.get('display_name', field['field_name']),
                    'text',  # 默认类型
                    field['sheet_name'],
                    field['cell_address'],
                    field.get('placeholder', ''),
                    field.get('default_value', ''),
                    1 if field.get('is_required', False) else 0,
                    1 if is_reference else 0,
                    f"工作表: {field['sheet_name']}, 位置: {field['cell_address']}"
                ))

                saved_count += 1
                if is_reference:
                    reference_count += 1
                    print(f"  ✓ 引用字段: [*{field['field_name']}] 在 {field['sheet_name']}!{field['cell_address']}")

            except Exception as e:
                print(f"  ✗ 保存字段失败 {field.get('field_name', 'unknown')}: {e}")

        conn.commit()
        conn.close()

        if reference_count > 0:
            print(f"  ℹ 其中引用字段（[*xx]格式）: {reference_count} 个")

        return saved_count

    def add_field_mapping(self, template_id, field_name, field_type, sheet_name,
                         cell_address=None, start_row=None, start_col=None,
                         description='', is_required=False, default_value=''):
        """
        添加字段映射配置

        Args:
            template_id: 模版ID
            field_name: 字段名称
            field_type: 字段类型
            sheet_name: 工作表名称
            cell_address: 单元格地址（如：B1）
            start_row: 起始行（用于表格数据）
            start_col: 起始列（用于表格数据）
            description: 字段描述
            is_required: 是否必填
            default_value: 默认值

        Returns:
            mapping_id: 映射ID
        """
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute(
            'INSERT INTO template_field_mappings '
            '(template_id, field_name, field_type, sheet_name, cell_address, '
            'start_row, start_col, description, is_required, default_value) '
            'VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
            (template_id, field_name, field_type, sheet_name, cell_address,
             start_row, start_col, description, is_required, default_value)
        )

        mapping_id = cursor.lastrowid
        conn.commit()
        conn.close()

        return mapping_id

    def get_template_info(self, template_id):
        """
        获取模版信息

        Args:
            template_id: 模版ID

        Returns:
            dict: 模版信息
        """
        conn = get_db_connection()

        template = conn.execute(
            'SELECT * FROM excel_report_templates WHERE id = ?',
            (template_id,)
        ).fetchone()

        if not template:
            conn.close()
            return None

        # 获取工作表配置
        sheets = conn.execute(
            'SELECT * FROM template_sheet_configs WHERE template_id = ? ORDER BY sheet_index',
            (template_id,)
        ).fetchall()

        # 获取字段映射
        fields = conn.execute(
            'SELECT * FROM template_field_mappings WHERE template_id = ?',
            (template_id,)
        ).fetchall()

        conn.close()

        return {
            'template': dict(template),
            'sheets': [dict(s) for s in sheets],
            'fields': [dict(f) for f in fields]
        }

    def list_templates(self):
        """
        列出所有报告模版

        Returns:
            list: 模版列表
        """
        conn = get_db_connection()

        templates = conn.execute(
            'SELECT t.*, st.name as sample_type_name '
            'FROM excel_report_templates t '
            'LEFT JOIN sample_types st ON t.sample_type_id = st.id '
            'ORDER BY t.created_at DESC'
        ).fetchall()

        conn.close()

        return [dict(t) for t in templates]

    def delete_template(self, template_id):
        """
        删除报告模版

        Args:
            template_id: 模版ID
        """
        conn = get_db_connection()

        # 获取模版文件路径
        template = conn.execute(
            'SELECT template_file_path FROM excel_report_templates WHERE id = ?',
            (template_id,)
        ).fetchone()

        if template and template['template_file_path']:
            # 删除文件
            try:
                if os.path.exists(template['template_file_path']):
                    os.remove(template['template_file_path'])
            except Exception as e:
                print(f"删除模版文件失败: {e}")

        # 删除数据库记录（会级联删除相关配置）
        conn.execute('DELETE FROM excel_report_templates WHERE id = ?', (template_id,))
        conn.commit()
        conn.close()

        print(f"✓ 模版已删除: ID={template_id}")

def main():
    """示例：导入报告模版"""
    manager = ReportTemplateManager()

    # 检查示例模版文件
    sample_template = 'sample/报告模版.xlsx'

    if os.path.exists(sample_template):
        print("发现示例报告模版，正在导入...")
        try:
            template_id = manager.import_template(
                sample_template,
                '水质检测标准模版',
                description='从sample目录导入的标准报告模版，包含出厂水和原水的报告格式'
            )

            print(f"\n模版导入成功！模版ID: {template_id}")

            # 显示模版信息
            info = manager.get_template_info(template_id)
            print(f"\n工作表列表：")
            for sheet in info['sheets']:
                print(f"  - {sheet['sheet_name']} (类型: {sheet['sheet_type']}, 页码: {sheet['page_number']})")

        except Exception as e:
            print(f"导入失败: {e}")
    else:
        print("未找到示例模版文件")

    # 列出所有模版
    print("\n当前所有模版：")
    templates = manager.list_templates()
    for t in templates:
        print(f"  - {t['name']} (ID: {t['id']}, 样品类型: {t.get('sample_type_name', '未指定')})")

if __name__ == '__main__':
    main()
