"""
样品类型导出器
用于导出检测项目数据填写模板
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from models_v2 import get_db_connection
import os
from datetime import datetime


class SampleTypeExporter:
    """样品类型导出器"""

    def __init__(self, sample_type_id):
        """
        初始化导出器

        Args:
            sample_type_id: 样品类型ID
        """
        self.sample_type_id = sample_type_id
        self.sample_type_info = None
        self.indicators = []

    def export(self, output_path=None):
        """
        导出检测项目填写模板

        Args:
            output_path: 输出文件路径

        Returns:
            str: 生成的文件路径
        """
        # 加载数据
        self._load_sample_type_data()

        # 创建工作簿
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # 1. 创建说明sheet
        self._create_instruction_sheet(wb)

        # 2. 创建检测数据sheet
        self._create_detection_data_sheet(wb)

        # 保存文件
        if output_path is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            sample_type_name = self.sample_type_info.get('name', '样品类型')
            output_path = f"exports/sample_type_{sample_type_name}_{timestamp}.xlsx"

        os.makedirs('exports', exist_ok=True)
        wb.save(output_path)
        wb.close()

        return output_path

    def _load_sample_type_data(self):
        """加载样品类型数据"""
        conn = get_db_connection()

        # 加载样品类型信息
        sample_type = conn.execute(
            'SELECT * FROM sample_types WHERE id = ?',
            (self.sample_type_id,)
        ).fetchone()

        if not sample_type:
            conn.close()
            raise ValueError(f'样品类型ID {self.sample_type_id} 不存在')

        self.sample_type_info = dict(sample_type)

        # 获取该样品类型关联的所有检测指标
        indicators = conn.execute('''
            SELECT i.id, i.group_id, i.name, i.unit, i.default_value,
                   i.description, i.created_at, i.limit_value,
                   i.detection_method, i.remark, ti.sort_order
            FROM indicators i
            JOIN template_indicators ti ON i.id = ti.indicator_id
            WHERE ti.sample_type_id = ?
            ORDER BY ti.sort_order, i.id
        ''', (self.sample_type_id,)).fetchall()

        self.indicators = [dict(ind) for ind in indicators]

        conn.close()

    def _create_instruction_sheet(self, wb):
        """创建填写说明sheet"""
        ws = wb.create_sheet("填写说明", 0)

        ws['A1'] = "检测项目数据填写说明"
        ws['A1'].font = Font(size=16, bold=True, color="4472C4")

        row = 3

        sample_type_name = self.sample_type_info.get('name', '未知')

        instructions = [
            ("一、模板信息", ""),
            ("", f"样品类型: {sample_type_name}"),
            ("", f"检测项目数量: {len(self.indicators)}"),
            ("", f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"),
            ("", ""),
            ("二、填写说明", ""),
            ("", "【检测数据】sheet 说明:"),
            ("", "  - 格式: 简化横向表格"),
            ("", "  - A列: 检测项目名称（请勿修改）"),
            ("", "  - B列: 单位（参考用）"),
            ("", "  - C列起: 样品编号（首行）及其检测结果"),
            ("", ""),
            ("", "【填写步骤】:"),
            ("", "  1. 在首行C列起修改样品编号（如 W260105C08, S20260128001...）"),
            ("", "     注意：第一个样品列（C列）为必填"),
            ("", "  2. 在对应列下方填写该样品各检测项目的检测结果"),
            ("", "  3. 如需增加样品，直接在右侧添加新列即可"),
            ("", "  4. 可以删除示例数据，但请保留表头行"),
            ("", ""),
            ("三、注意事项", ""),
            ("", "1. 样品编号必须唯一"),
            ("", "2. 检测结果直接填写数值或文本（如：7.2、未检出、<1）"),
            ("", "3. 不需要的检测项目可以留空，但不要删除该行"),
            ("", "4. 不要修改A列的检测项目名称"),
            ("", "5. 如果是特殊结果（如未检出、无等），直接输入文本即可"),
            ("", "6. 表格已冻结首行和前2列，方便浏览大量数据"),
            ("", ""),
            ("四、导入步骤", ""),
            ("", "1. 填写完成此Excel文件"),
            ("", "2. 在系统中点击【解析Excel】上传"),
            ("", "3. 系统将自动解析检测数据"),
            ("", ""),
            ("五、示例", ""),
            ("", "表格格式示例："),
            ("", "检测项目  | 单位      | W260105C08 | W260105C09 | W260105C10"),
            ("", "pH        | 无量纲    | 7.2        | 7.3        | 7.1"),
            ("", "浊度      | NTU       | 0.5        | 0.6        | 0.4"),
            ("", "余氯      | mg/L      | 0.3        | 0.35       | 0.28"),
        ]

        for title, content in instructions:
            if title:
                cell = ws.cell(row, 1)
                cell.value = title
                cell.font = Font(size=12, bold=True, color="4472C4")
            else:
                cell = ws.cell(row, 1)
                cell.value = content
                cell.alignment = Alignment(wrap_text=True)

            row += 1

        ws.column_dimensions['A'].width = 100

    def _create_detection_data_sheet(self, wb):
        """创建检测数据sheet（简化格式）"""
        ws = wb.create_sheet("检测数据")

        # 设置样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        subheader_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
        subheader_font = Font(bold=True, size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 第一行：表头（简化格式：检测项目、单位、样品数据）
        headers = ['检测项目', '单位']

        # 添加示例样品编号列（3个样品）
        sample_numbers = ['样品编号1*', '样品编号2', '样品编号3']

        for col_idx, header in enumerate(headers + sample_numbers, start=1):
            cell = ws.cell(1, col_idx)
            cell.value = header

            if col_idx <= len(headers):
                # 项目信息列
                cell.fill = subheader_fill
                cell.font = subheader_font
            else:
                # 样品列
                cell.fill = header_fill
                cell.font = header_font

            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # 填充检测项目信息
        for row_idx, indicator in enumerate(self.indicators, start=2):
            # 检测项目（A列）
            cell = ws.cell(row_idx, 1)
            cell.value = indicator['name']
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = border

            # 单位（B列）
            cell = ws.cell(row_idx, 2)
            cell.value = indicator.get('unit', '')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

            # 样品结果列（C列起，第一个样品填写示例数据）
            if row_idx == 2:
                # 仅第一行第一个样品填写示例数据
                cell = ws.cell(row_idx, 3)
                cell.value = indicator.get('default_value', '')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            # 其他样品列留空但设置边框
            for col_idx in range(3 if row_idx > 2 else 4, 3 + len(sample_numbers)):
                cell = ws.cell(row_idx, col_idx)
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # 调整列宽
        ws.column_dimensions['A'].width = 20  # 检测项目
        ws.column_dimensions['B'].width = 12  # 单位

        # 样品列
        for col_letter in ['C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col_letter].width = 15

        # 冻结首行和前2列
        ws.freeze_panes = 'C2'


def export_sample_type_template(sample_type_id, output_path=None):
    """
    导出样品类型检测模板的便捷函数

    Args:
        sample_type_id: 样品类型ID
        output_path: 输出路径（可选）

    Returns:
        str: 生成的文件路径
    """
    exporter = SampleTypeExporter(sample_type_id)
    return exporter.export(output_path)


if __name__ == '__main__':
    print("="*60)
    print("样品类型导出器测试")
    print("="*60)

    conn = get_db_connection()
    sample_types = conn.execute('SELECT id, name FROM sample_types LIMIT 1').fetchone()
    conn.close()

    if sample_types:
        sample_type_id = sample_types['id']
        sample_type_name = sample_types['name']
        print(f"\n导出样品类型 [{sample_type_name}] 的检测模板...")
        path = export_sample_type_template(sample_type_id)
        print(f"✓ 生成成功: {path}")
    else:
        print("\n没有找到样品类型")
