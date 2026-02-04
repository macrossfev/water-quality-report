"""
分析报告模板Excel结构
提取字段映射规则（[]、()、;符号规则）
"""
import openpyxl
import re
from openpyxl.utils import get_column_letter

def analyze_template(template_path):
    """分析模板文件，提取所有包含特殊标记的单元格"""
    wb = openpyxl.load_workbook(template_path)

    results = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n{'='*60}")
        print(f"工作表: {sheet_name}")
        print(f"{'='*60}")

        # 遍历所有单元格
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row, col)
                cell_value = str(cell.value) if cell.value else ""

                # 检查是否包含特殊标记
                has_bracket = '[' in cell_value and ']' in cell_value
                has_paren = '(' in cell_value and ')' in cell_value
                has_semicolon = ';' in cell_value

                if has_bracket or has_paren or has_semicolon:
                    cell_address = f"{get_column_letter(col)}{row}"

                    result = {
                        'sheet': sheet_name,
                        'cell': cell_address,
                        'row': row,
                        'col': col,
                        'value': cell_value,
                        'has_bracket': has_bracket,
                        'has_paren': has_paren,
                        'has_semicolon': has_semicolon
                    }

                    results.append(result)

                    # 打印详细信息
                    print(f"\n位置: {cell_address}")
                    print(f"内容: {cell_value}")

                    # 解析字段名（方括号内容）
                    if has_bracket:
                        field_names = re.findall(r'\[(.*?)\]', cell_value)
                        print(f"字段名: {field_names}")

                    # 解析待填内容（圆括号内容）
                    if has_paren:
                        fill_contents = re.findall(r'\((.*?)\)', cell_value)
                        print(f"待填内容: {fill_contents}")

                    # 解析分号规则
                    if has_semicolon:
                        parts = cell_value.split(';')
                        print(f"分号前: '{parts[0].strip()}'")
                        if len(parts) > 1:
                            print(f"分号后: '{parts[1].strip()}'")

                            # 判断规则
                            if not parts[0].strip():
                                print("规则: 必填项（分号前为空）")
                            else:
                                print("规则: 默认值可编辑（分号前有内容）")

    wb.close()

    print(f"\n\n{'='*60}")
    print(f"汇总: 共找到 {len(results)} 个包含特殊标记的单元格")
    print(f"{'='*60}")

    return results

if __name__ == '__main__':
    template_path = '/home/macrossfev/water-quality-report/sample/报告模版.xlsx'

    try:
        results = analyze_template(template_path)

        # 按工作表分组统计
        sheets = {}
        for r in results:
            sheet = r['sheet']
            if sheet not in sheets:
                sheets[sheet] = []
            sheets[sheet].append(r)

        print("\n按工作表统计:")
        for sheet, items in sheets.items():
            print(f"  {sheet}: {len(items)} 个字段")

    except Exception as e:
        print(f"分析失败: {e}")
        import traceback
        traceback.print_exc()
