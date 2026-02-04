#!/usr/bin/env python3
"""检查生成的Excel报告"""
import openpyxl

file_path = 'exports/report_20260203编4_20260203_014733.xlsx'

wb = openpyxl.load_workbook(file_path)

print(f"Excel文件: {file_path}")
print(f"工作表数量: {len(wb.sheetnames)}\n")

print("工作表列表:")
for i, sheet_name in enumerate(wb.sheetnames, 1):
    ws = wb[sheet_name]
    print(f"  {i}. {sheet_name} (最大行: {ws.max_row}, 最大列: {ws.max_column})")

# 检查第三页（通常是检测数据页）
if len(wb.sheetnames) >= 3:
    sheet_name = wb.sheetnames[2]  # 第三页
    ws = wb[sheet_name]

    print(f"\n第三页 '{sheet_name}' 的内容（前20行）:")
    print("-" * 80)

    for row in range(1, min(21, ws.max_row + 1)):
        row_data = []
        for col in range(1, min(8, ws.max_column + 1)):
            cell = ws.cell(row, col)
            value = cell.value
            if value:
                row_data.append(f"{value}")

        if row_data:
            print(f"第{row}行: {' | '.join(row_data)}")

    # 特别检查第8-9行
    print(f"\n重点检查第8-9行（应该包含pH和菌落总数）:")
    for row in [8, 9]:
        print(f"第{row}行:")
        for col in range(1, 8):
            cell = ws.cell(row, col)
            print(f"  列{col}: {repr(cell.value)}")

wb.close()
