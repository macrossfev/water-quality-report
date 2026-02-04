#!/usr/bin/env python3
"""检查实际的Excel模板文件内容"""
import openpyxl

template_path = 'templates/excel_reports/出厂水_20260203_002830.xlsx'

wb = openpyxl.load_workbook(template_path)

print(f"模板文件: {template_path}")
print(f"工作表: {wb.sheetnames}\n")

# 检查第三页
if len(wb.sheetnames) >= 3:
    sheet_name = wb.sheetnames[2]
    ws = wb[sheet_name]

    print(f"第三页 '{sheet_name}' 的第6-10行内容:")
    print("-" * 80)

    for row in range(6, 11):
        print(f"\n第{row}行:")
        for col in range(1, 8):
            cell = ws.cell(row, col)
            if cell.value:
                print(f"  {chr(64+col)}{row}: {repr(cell.value)}")

    # 搜索包含 #dt 的单元格
    print(f"\n\n搜索包含 '#dt' 的单元格:")
    found = False
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            if cell.value and '#dt' in str(cell.value):
                print(f"  {chr(64+col) if col <= 26 else 'col'+str(col)}{row}: {repr(cell.value)}")
                found = True

    if not found:
        print("  ❌ 没有找到任何包含 '#dt' 的单元格！")

    # 搜索包含 [ 和 ] 的单元格
    print(f"\n搜索包含 '[...]' 的单元格（前20个）:")
    count = 0
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            if cell.value and '[' in str(cell.value) and ']' in str(cell.value):
                print(f"  {chr(64+col) if col <= 26 else 'col'+str(col)}{row}: {repr(cell.value)}")
                count += 1
                if count >= 20:
                    break
        if count >= 20:
            break

    if count == 0:
        print("  ❌ 没有找到任何包含 '[...]' 的单元格！")

wb.close()
