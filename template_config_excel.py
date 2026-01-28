"""
模板配置Excel导入导出
用于导入和导出模板字段配置
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from models_v2 import get_db_connection
from datetime import datetime
import os


class TemplateConfigExcel:
    """模板配置Excel处理器"""

    @staticmethod
    def export_template_config(template_id, output_path=None):
        """
        导出模板配置到Excel文件

        Args:
            template_id: 模板ID
            output_path: 输出文件路径，如果为None则自动生成

        Returns:
            str: 生成的文件路径
        """
        conn = get_db_connection()

        # 获取模板信息
        template = conn.execute(
            'SELECT * FROM excel_report_templates WHERE id = ?',
            (template_id,)
        ).fetchone()

        if not template:
            conn.close()
            raise ValueError(f'模板不存在: ID={template_id}')

        # 获取字段映射
        fields = conn.execute(
            'SELECT * FROM template_field_mappings WHERE template_id = ? ORDER BY id',
            (template_id,)
        ).fetchall()

        conn.close()

        # 创建Excel工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '模板配置'

        # 设置样式
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        border_style = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # 设置标题行
        headers = [
            'ID', '字段名称', '显示名称', '字段类型', '工作表名称',
            '单元格地址', '占位符', '默认值', '是否必填', '描述'
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border_style

        # 设置列宽
        column_widths = [8, 20, 20, 15, 15, 15, 30, 20, 10, 30]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        # 填充数据
        data_font = Font(name='微软雅黑', size=10)
        data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        for row_idx, field in enumerate(fields, 2):
            field_dict = dict(field)

            row_data = [
                field_dict.get('id', ''),
                field_dict.get('field_name', ''),
                field_dict.get('field_display_name', ''),
                field_dict.get('field_type', ''),
                field_dict.get('sheet_name', ''),
                field_dict.get('cell_address', ''),
                field_dict.get('placeholder', ''),
                field_dict.get('default_value', ''),
                '是' if field_dict.get('is_required', 0) == 1 else '否',
                field_dict.get('description', '')
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border_style

        # 添加说明sheet
        ws_info = wb.create_sheet('导入说明')

        instructions = [
            ['模板配置导入说明', ''],
            ['', ''],
            ['文件格式：', 'Excel文件（.xlsx或.xls）'],
            ['工作表名称：', '必须包含名为"模板配置"的工作表'],
            ['', ''],
            ['字段说明：', ''],
            ['ID', '字段ID（导入时此列会被忽略，系统自动生成）'],
            ['字段名称', '必填，字段的唯一标识符'],
            ['显示名称', '可选，用于界面显示的名称'],
            ['字段类型', '必填，可选值：text/date/selection/table_data/signature/constant/formula'],
            ['工作表名称', '必填，该字段所在的Excel工作表名称'],
            ['单元格地址', '可选，单元格位置，如：A1, B2'],
            ['占位符', '可选，输入框的提示文本'],
            ['默认值', '可选，字段的默认值'],
            ['是否必填', '必填，可选值：是/否'],
            ['描述', '可选，字段的详细描述'],
            ['', ''],
            ['注意事项：', ''],
            ['1. 导入时会删除该模板的所有现有字段配置，然后重新创建', ''],
            ['2. 请确保字段名称、字段类型、工作表名称等必填字段不为空', ''],
            ['3. 字段类型必须是系统支持的类型之一', ''],
            ['4. 是否必填列只能填"是"或"否"', ''],
            ['', ''],
            [f'导出时间：', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            [f'模板名称：', template['name']],
            [f'模板ID：', str(template['id'])],
        ]

        for row_idx, (label, value) in enumerate(instructions, 1):
            cell_a = ws_info.cell(row=row_idx, column=1, value=label)
            cell_b = ws_info.cell(row=row_idx, column=2, value=value)

            if row_idx == 1:
                cell_a.font = Font(name='微软雅黑', size=14, bold=True, color='4472C4')
            elif label in ['文件格式：', '工作表名称：', '字段说明：', '注意事项：', '导出时间：', '模板名称：', '模板ID：']:
                cell_a.font = Font(name='微软雅黑', size=10, bold=True)
            else:
                cell_a.font = Font(name='微软雅黑', size=10)

            cell_b.font = Font(name='微软雅黑', size=10)

        ws_info.column_dimensions['A'].width = 20
        ws_info.column_dimensions['B'].width = 60

        # 生成输出文件路径
        if output_path is None:
            os.makedirs('exports/template_configs', exist_ok=True)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            template_name = template['name'].replace('/', '_').replace('\\', '_')
            output_path = f'exports/template_configs/{template_name}_配置_{timestamp}.xlsx'

        # 保存文件
        wb.save(output_path)

        return output_path

    @staticmethod
    def import_template_config(template_id, excel_path):
        """
        从Excel文件导入模板配置

        Args:
            template_id: 模板ID
            excel_path: Excel文件路径

        Returns:
            dict: 导入结果统计
        """
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f'Excel文件不存在: {excel_path}')

        # 读取Excel文件
        try:
            wb = openpyxl.load_workbook(excel_path)
        except Exception as e:
            raise ValueError(f'无法读取Excel文件: {str(e)}')

        # 检查是否包含"模板配置"工作表
        if '模板配置' not in wb.sheetnames:
            raise ValueError('Excel文件中未找到"模板配置"工作表')

        ws = wb['模板配置']

        # 读取标题行
        headers = []
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col_idx).value
            headers.append(cell_value)

        # 验证必需的列
        required_columns = ['字段名称', '字段类型', '工作表名称', '是否必填']
        for col in required_columns:
            if col not in headers:
                raise ValueError(f'缺少必需的列: {col}')

        # 读取数据行
        fields_data = []
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            for col_idx, header in enumerate(headers, 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                row_data[header] = cell_value if cell_value is not None else ''

            # 跳过空行（字段名称为空）
            if not row_data.get('字段名称'):
                continue

            # 验证必填字段
            if not row_data.get('字段名称') or not row_data.get('字段类型') or not row_data.get('工作表名称'):
                raise ValueError(f'第{row_idx}行缺少必填字段（字段名称/字段类型/工作表名称）')

            # 验证字段类型
            valid_types = ['text', 'date', 'selection', 'table_data', 'signature', 'constant', 'formula']
            if row_data.get('字段类型') not in valid_types:
                raise ValueError(f'第{row_idx}行字段类型无效: {row_data.get("字段类型")}，必须是以下之一: {", ".join(valid_types)}')

            # 转换是否必填
            is_required_text = str(row_data.get('是否必填', '否')).strip()
            if is_required_text == '是':
                is_required = 1
            elif is_required_text == '否':
                is_required = 0
            else:
                raise ValueError(f'第{row_idx}行"是否必填"列的值无效: {is_required_text}，必须是"是"或"否"')

            fields_data.append({
                'field_name': row_data.get('字段名称', '').strip(),
                'field_display_name': row_data.get('显示名称', '').strip() or row_data.get('字段名称', '').strip(),
                'field_type': row_data.get('字段类型', '').strip(),
                'sheet_name': row_data.get('工作表名称', '').strip(),
                'cell_address': row_data.get('单元格地址', '').strip(),
                'placeholder': row_data.get('占位符', '').strip(),
                'default_value': row_data.get('默认值', '').strip(),
                'is_required': is_required,
                'description': row_data.get('描述', '').strip()
            })

        if not fields_data:
            raise ValueError('Excel文件中没有有效的字段数据')

        # 开始导入到数据库
        conn = get_db_connection()
        cursor = conn.cursor()

        try:
            # 验证模板是否存在
            template = cursor.execute(
                'SELECT * FROM excel_report_templates WHERE id = ?',
                (template_id,)
            ).fetchone()

            if not template:
                raise ValueError(f'模板不存在: ID={template_id}')

            # 删除该模板的所有现有字段配置
            cursor.execute(
                'DELETE FROM template_field_mappings WHERE template_id = ?',
                (template_id,)
            )

            # 插入新的字段配置
            inserted_count = 0
            for field in fields_data:
                cursor.execute(
                    '''INSERT INTO template_field_mappings
                       (template_id, field_name, field_display_name, field_type,
                        sheet_name, cell_address, placeholder, default_value, is_required, description)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                    (template_id,
                     field['field_name'],
                     field['field_display_name'],
                     field['field_type'],
                     field['sheet_name'],
                     field['cell_address'],
                     field['placeholder'],
                     field['default_value'],
                     field['is_required'],
                     field['description'])
                )
                inserted_count += 1

            # 提交事务
            conn.commit()

            return {
                'success': True,
                'message': f'成功导入{inserted_count}个字段配置',
                'inserted_count': inserted_count,
                'template_name': template['name']
            }

        except Exception as e:
            conn.rollback()
            raise e
        finally:
            conn.close()


def main():
    """测试导入导出功能"""
    from report_template_manager import ReportTemplateManager

    manager = ReportTemplateManager()
    templates = manager.list_templates()

    if not templates:
        print('暂无模板，无法测试')
        return

    # 测试导出
    template_id = templates[0]['id']
    print(f'\n测试导出模板配置：{templates[0]["name"]} (ID: {template_id})')

    try:
        output_path = TemplateConfigExcel.export_template_config(template_id)
        print(f'✓ 导出成功: {output_path}')

        # 测试导入
        print(f'\n测试导入模板配置...')
        result = TemplateConfigExcel.import_template_config(template_id, output_path)
        print(f'✓ 导入成功: {result["message"]}')

    except Exception as e:
        print(f'✗ 测试失败: {str(e)}')


if __name__ == '__main__':
    main()
