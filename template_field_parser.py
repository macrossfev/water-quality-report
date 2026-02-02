"""
模板字段解析器
解析[]、()、;格式，生成填写表单配置

字段格式规则:
1. [字段名];(占位符说明) - 必填字段，无默认值
2. [字段名]默认值;(占位符说明) - 有默认值，可编辑
3. [字段名](占位符说明) - 必填字段（简化格式）
4. [字段名] - 必填字段（最简格式）
5. [*字段名] - 引用字段，从已审核报告中查找数据（不可编辑）
6. [#代号] - 标准字段代号，如 [#report_no], [#dt_name]（推荐使用）

示例:
- [报告编号];(请输入报告编号) - 必填
- [检测日期]2025-01-15;(检测日期) - 默认值为2025-01-15，可编辑
- [检测人]张三;() - 默认值为张三，可编辑
- [*被检单位] - 从已审核报告中引用被检单位数据
- [*采样日期] - 从已审核报告中引用采样日期
- [#report_no] - 报告编号（使用标准代号）
- [#dt_name] - 检测项目名称列（检测数据）
- [#dt_end] - 数据区结束标记
"""
import re
from typing import Dict, List, Optional, Tuple
from field_code_mapping import FieldCodeMapping

class TemplateFieldParser:
    """模板字段解析器"""

    @staticmethod
    def parse_field(field_text: str) -> Dict:
        """
        解析单个字段文本

        Args:
            field_text: 字段文本，如 "[报告编号];(请输入报告编号)" 或 "[*被检单位]"

        Returns:
            dict: {
                'field_name': '报告编号',
                'display_name': '报告编号',
                'default_value': None,
                'placeholder': '请输入报告编号',
                'is_required': True,
                'is_editable': True,
                'is_reference': False  # 是否为引用字段
            }
        """
        result = {
            'field_name': '',
            'display_name': '',
            'default_value': None,
            'placeholder': '',
            'is_required': True,
            'is_editable': True,
            'is_reference': False,  # 新增：标记是否为引用字段
            'field_code': None,     # 字段代号（如果使用代号）
            'field_type': 'text'    # 字段类型
        }

        if not field_text:
            return result

        # 1. 提取字段名（方括号内容）
        field_name_match = re.search(r'\[(.*?)\]', field_text)
        if field_name_match:
            raw_field_name = field_name_match.group(1).strip()

            # 检查是否为字段代号（以 # 开头）
            if raw_field_name.startswith('#'):
                code_info = FieldCodeMapping.get_field_info(f'[{raw_field_name}]')
                if code_info:
                    result['field_code'] = raw_field_name
                    result['display_name'] = code_info['display_name']
                    result['placeholder'] = code_info['description']
                    result['is_editable'] = False  # 代号字段由系统填充
                    result['is_required'] = False

                    # 根据代号类型设置字段信息
                    if code_info['type'] == 'basic_field':
                        # 基本字段
                        result['field_name'] = code_info['db_field']
                        result['field_type'] = 'text'
                    elif code_info['type'] == 'detection_column':
                        # 检测数据列
                        result['field_name'] = code_info['display_name']
                        result['field_type'] = 'detection_column'
                        result['column_mapping'] = code_info['column_mapping']
                    elif code_info['type'] == 'data_region_end':
                        # 数据区结束标记
                        result['field_name'] = 'data_region_end'
                        result['field_type'] = 'control_mark'
                        result['control_type'] = 'data_region_end'
                    else:
                        # 其他控制标记
                        result['field_name'] = code_info.get('type', 'unknown')
                        result['field_type'] = 'control_mark'
                        result['control_type'] = code_info.get('type', 'unknown')

                    return result
                else:
                    # 未知代号，当作普通字段处理
                    result['field_name'] = raw_field_name
                    result['display_name'] = raw_field_name
                    return result

            # 检查是否为引用字段（以 * 开头）
            elif raw_field_name.startswith('*'):
                result['is_reference'] = True
                result['is_editable'] = False  # 引用字段不可编辑
                result['is_required'] = False  # 引用字段由系统自动填充
                # 去掉 * 号，得到实际字段名
                result['field_name'] = raw_field_name[1:].strip()
                result['display_name'] = result['field_name']
                result['placeholder'] = f"从已审核报告中引用{result['field_name']}"
                # 引用字段不需要继续解析其他部分，直接返回
                return result
            else:
                result['field_name'] = raw_field_name
                result['display_name'] = result['field_name']
        else:
            # 没有方括号，使用整个文本作为字段名
            result['field_name'] = field_text.strip()
            result['display_name'] = field_text.strip()
            return result

        # 2. 检查是否有分号
        if ';' in field_text:
            # 有分号，需要解析默认值和占位符
            # 分号前的内容（除去[字段名]部分）
            before_semicolon = field_text.split(';')[0]
            after_semicolon = field_text.split(';')[1] if len(field_text.split(';')) > 1 else ''

            # 提取默认值（方括号后、分号前的内容）
            default_value_match = re.search(r'\](.+)', before_semicolon)
            if default_value_match:
                default_val = default_value_match.group(1).strip()
                if default_val:
                    result['default_value'] = default_val
                    result['is_required'] = False  # 有默认值，非必填
            else:
                # 分号前为空（只有[字段名];），表示必填
                result['is_required'] = True

            # 提取占位符（圆括号内容）
            placeholder_match = re.search(r'\((.*?)\)', after_semicolon)
            if placeholder_match:
                result['placeholder'] = placeholder_match.group(1).strip()

        else:
            # 没有分号，检查是否有圆括号
            placeholder_match = re.search(r'\((.*?)\)', field_text)
            if placeholder_match:
                result['placeholder'] = placeholder_match.group(1).strip()

            # 检查方括号后是否有内容作为默认值
            after_bracket = field_text.split(']')[1] if ']' in field_text else ''
            # 移除圆括号部分
            after_bracket = re.sub(r'\(.*?\)', '', after_bracket).strip()
            if after_bracket:
                result['default_value'] = after_bracket
                result['is_required'] = False

        return result

    @staticmethod
    def parse_cell_value(cell_value: str) -> List[Dict]:
        """
        解析单元格值中的所有字段

        Args:
            cell_value: 单元格值，可能包含多个字段

        Returns:
            list: 字段列表
        """
        if not cell_value:
            return []

        fields = []

        # 查找所有方括号标记的字段
        field_matches = re.finditer(r'\[([^\]]+)\]', cell_value)

        for match in field_matches:
            # 获取字段及其周围的上下文
            start = match.start()
            end = match.end()

            # 向后查找直到遇到下一个[或字符串结束
            context_end = cell_value.find('[', end)
            if context_end == -1:
                context_end = len(cell_value)

            field_text = cell_value[start:context_end].strip()
            field_info = TemplateFieldParser.parse_field(field_text)

            if field_info['field_name']:
                fields.append(field_info)

        return fields

    @staticmethod
    def extract_template_fields(template_path: str) -> List[Dict]:
        """
        从模板文件中提取所有字段配置

        Args:
            template_path: 模板文件路径

        Returns:
            list: 字段配置列表
        """
        import openpyxl
        from openpyxl.utils import get_column_letter

        wb = openpyxl.load_workbook(template_path)
        all_fields = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            for row in range(1, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row, col)
                    cell_value = str(cell.value) if cell.value else ""

                    # 检查是否包含字段标记
                    if '[' in cell_value and ']' in cell_value:
                        fields = TemplateFieldParser.parse_cell_value(cell_value)

                        for field in fields:
                            field['sheet_name'] = sheet_name
                            field['cell_address'] = f"{get_column_letter(col)}{row}"
                            field['row'] = row
                            field['col'] = col
                            field['original_value'] = cell_value
                            all_fields.append(field)

        wb.close()
        return all_fields

    @staticmethod
    def generate_form_config(template_id: int, fields: List[Dict]) -> List[Dict]:
        """
        生成表单配置

        Args:
            template_id: 模板ID
            fields: 字段列表

        Returns:
            list: 表单配置
        """
        form_config = []

        # 兼容性映射：对于没有使用代号的旧模板，仍然支持中文字段名识别
        legacy_detection_columns = {
            '序号': 'index',
            '检测项目': 'name',
            '项目名称': 'name',
            '检测项': 'name',
            '单位': 'unit',
            '检测值': 'result',
            '检测结果': 'result',
            '测定值': 'result',
            '结果': 'result',
            '限值': 'limit',
            '标准限值': 'limit',
            '限定值': 'limit',
            '检测方法': 'method',
            '方法': 'method',
            '检测依据': 'method',
            '单项判定': 'judgment',
            '判定': 'judgment',
        }

        for field in fields:
            config = {
                'template_id': template_id,
                'field_name': field['field_name'],
                'field_display_name': field['display_name'],
                'field_type': field.get('field_type', 'text'),  # 使用解析出的类型
                'sheet_name': field['sheet_name'],
                'cell_address': field['cell_address'],
                'placeholder': field['placeholder'],
                'default_value': field['default_value'],
                'is_required': field['is_required'],
                'description': f"工作表:{field['sheet_name']}, 位置:{field['cell_address']}"
            }

            # 如果字段已经在解析阶段确定了类型（如使用代号），直接使用
            if field.get('field_type') == 'detection_column':
                config['column_mapping'] = field.get('column_mapping')
                config['description'] = f"检测数据列: {field['display_name']} -> {field.get('column_mapping')}"
            elif field.get('field_type') == 'control_mark':
                config['control_type'] = field.get('control_type')
                config['description'] = f"控制标记: {field.get('control_type')}"
            # 兼容旧模板：检查是否为检测数据列标记（中文字段名）
            elif field['field_name'] in legacy_detection_columns:
                config['field_type'] = 'detection_column'
                config['column_mapping'] = legacy_detection_columns[field['field_name']]
                config['description'] = f"检测数据列: {field['field_name']} -> {legacy_detection_columns[field['field_name']]}"
            else:
                # 根据字段名判断类型（对于普通字段）
                if config['field_type'] == 'text':  # 只有当类型还是默认的text时才判断
                    field_name_lower = field['field_name'].lower()
                    if '日期' in field['field_name'] or 'date' in field_name_lower:
                        config['field_type'] = 'date'
                    elif '时间' in field['field_name'] or 'time' in field_name_lower:
                        config['field_type'] = 'datetime'
                    elif '数量' in field['field_name'] or '编号' in field['field_name']:
                        config['field_type'] = 'number'
                    elif '备注' in field['field_name'] or '说明' in field['field_name']:
                        config['field_type'] = 'textarea'

            form_config.append(config)

        return form_config


def test_parser():
    """测试解析器"""
    test_cases = [
        "[报告编号];(请输入报告编号)",
        "[检测日期]2025-01-15;(检测日期)",
        "[检测人]张三;()",
        "[样品编号](样品编号)",
        "[委托单位]",
        "检测项目: [总大肠菌群]0;(CFU/100mL)",
    ]

    print("="*60)
    print("模板字段解析器测试")
    print("="*60)

    for test_text in test_cases:
        print(f"\n输入: {test_text}")
        result = TemplateFieldParser.parse_field(test_text)
        print(f"结果:")
        print(f"  字段名: {result['field_name']}")
        print(f"  显示名: {result['display_name']}")
        print(f"  默认值: {result['default_value']}")
        print(f"  占位符: {result['placeholder']}")
        print(f"  必填: {result['is_required']}")
        print(f"  可编辑: {result['is_editable']}")


if __name__ == '__main__':
    test_parser()
