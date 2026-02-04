#!/usr/bin/env python3
"""测试生成报告20260203编4"""
from report_generator import ReportGenerator
import os

# 报告参数
report_id = 40
template_id = 8

print(f"生成报告:")
print(f"  报告ID: {report_id}")
print(f"  模板ID: {template_id} (出厂水)")

try:
    # 生成报告
    generator = ReportGenerator(template_id, {}, report_id=report_id)
    output_path = generator.generate()

    print(f"\n✅ 报告生成成功！")
    print(f"输出路径: {output_path}")
    print(f"文件大小: {os.path.getsize(output_path)} 字节")

    # 验证生成的Excel
    import openpyxl
    wb = openpyxl.load_workbook(output_path)

    print(f"\n检查第三页第8-10行:")
    ws = wb[wb.sheetnames[2]]

    for row in [8, 9, 10]:
        print(f"\n第{row}行:")
        for col in range(1, 7):
            cell = ws.cell(row, col)
            if cell.value:
                print(f"  列{col}: {cell.value}")

    wb.close()

except Exception as e:
    import traceback
    print(f"\n❌ 生成失败:")
    traceback.print_exc()
